"""Phase 6 D7 #3 — LLM Output Spot Check (Kimi as Independent Judge)

对 D4 (trustpilot) / D5 (zendesk + amazon) 三批 LLM 重打输出做抽样质量验证。

策略：
  1. 每个数据源分层抽 N 条（默认 50）
  2. 对每条，用 Kimi 独立判定每个 tag_id 是否被 review text 支撑
  3. 输出：tag-level precision (Kimi-validated / pred), 整批 precision, 样本审计表

  注意：这是 precision 评估（Kimi 是独立判官），不评估 recall。

设计：
  - 输入：F4/F3 的 *_relabel.jsonl + 原始 text 来自 phase5_full_labeled.jsonl
  - Kimi 判定 prompt：给定 text + tag(中英 + 定义) → {accept: bool, reason: str}
  - 单条 1 tag 1 调用过细，改为 1 review × N tag 一次性判定
  - 阈值：source 的 precision ≥ 0.85 视为可信；< 0.85 标记为质量风险

成本估算：
  - 3 source × 50 samples × 1 call/sample = 150 Kimi 调用
  - tokens_in ~600/call × 150 = 90K
  - tokens_out ~200/call × 150 = 30K
  - ~$0.02 (Kimi pricing)
  - 耗时 concurrency=10, ~0.5s/call → ~10s

用法：
  python llm_output_spot_check.py \
    --base <vault>/04-输出结果/unified_labeling/phase5_full_labeled.jsonl \
    --relabel-trustpilot <vault>/04-输出结果/unified_labeling/phase6_multilingual_relabel.jsonl \
    --relabel-zendesk    <vault>/04-输出结果/unified_labeling/phase6_zendesk_relabel.jsonl \
    --relabel-amazon     <vault>/04-输出结果/unified_labeling/phase6_amazon_relabel.jsonl \
    --dict <vault>/04-输出结果/01-字典版本/tag_dictionary_v4.1.xlsx \
    --report <vault>/04-输出结果/03-审计报告/phase6_d7_spot_check.md \
    [--samples-per-source 50]
"""

from __future__ import annotations

import argparse
import asyncio
import json
import random
import sys
import time
from collections import Counter
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "07-LLM引擎"))


JUDGE_PROMPT = """You are an independent LLM judge for VOC tag classification quality.

For each given (review_text, list-of-predicted-tags) pair, evaluate each predicted tag:
- accept = true if the tag is clearly supported by evidence in review_text
- accept = false if NOT supported, contradicted, or too tenuous a connection
- Reason: 1 short sentence

Output JSON: {"verdicts": [{"tag_id": "TAG_X", "accept": true, "reason": "matches phrase..."}, ...]}

Be strict: prefer false negatives (rejecting weak matches) over false positives."""


@dataclass
class SourceStats:
    n_samples: int = 0
    n_total_tags: int = 0
    n_accepted: int = 0
    n_rejected: int = 0
    n_judge_failed: int = 0
    rejected_examples: list[dict[str, Any]] = field(default_factory=list)


def load_dict_meta(dict_path: Path) -> dict[str, dict[str, Any]]:
    wb = openpyxl.load_workbook(str(dict_path), read_only=True, data_only=True)
    ws = wb["01_通用标签主表"]
    hdr = [c.value for c in next(ws.iter_rows(max_row=1))]
    ix = {k: hdr.index(k) for k in [
        "标签ID", "VOC标签（中文）", "VOC标签（英文）", "标签定义",
    ]}
    out: dict[str, dict[str, Any]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        tid = row[ix["标签ID"]]
        if not tid:
            continue
        out[str(tid)] = {
            "tag_en": str(row[ix["VOC标签（英文）"]] or ""),
            "tag_cn": str(row[ix["VOC标签（中文）"]] or ""),
            "definition": str(row[ix["标签定义"]] or ""),
        }
    return out


def load_relabel_index(path: Path) -> dict[str, list[dict[str, Any]]]:
    """Map review_id → list of {tag_id, confidence} from F4/F3 output."""
    idx: dict[str, list[dict[str, Any]]] = {}
    with path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                r = json.loads(line)
            except json.JSONDecodeError:
                continue
            rid = r.get("review_id")
            labels = r.get("labels") or []
            if rid and labels:
                idx[rid] = [l for l in labels if isinstance(l, dict) and l.get("tag_id")]
    return idx


def collect_text_for_ids(base_path: Path, target_ids: set[str]) -> dict[str, dict[str, Any]]:
    """Stream base jsonl, return {rid: {text, data_source, rating, language}} for target ids."""
    out: dict[str, dict[str, Any]] = {}
    with base_path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                r = json.loads(line)
            except json.JSONDecodeError:
                continue
            rid = r.get("review_id")
            if rid in target_ids:
                out[rid] = {
                    "text": r.get("text") or "",
                    "data_source": r.get("data_source"),
                    "rating": r.get("rating"),
                    "language": r.get("language"),
                }
                if len(out) == len(target_ids):
                    break
    return out


async def judge_one(
    client,
    rid: str,
    text: str,
    pred_labels: list[dict[str, Any]],
    dict_meta: dict[str, dict[str, Any]],
    stats: SourceStats,
) -> list[dict[str, Any]]:
    text_clip = (text or "").replace("\n", " ").strip()
    if len(text_clip) > 800:
        text_clip = text_clip[:800] + "..."

    pred_lines = []
    for lbl in pred_labels:
        tid = lbl["tag_id"]
        meta = dict_meta.get(tid, {})
        defn = meta.get("definition", "")
        if len(defn) > 100:
            defn = defn[:100] + "..."
        pred_lines.append(
            f"- {tid}\t{meta.get('tag_en','')}\t{meta.get('tag_cn','')}\t({defn})"
        )

    user_prompt = f"""REVIEW (review_id={rid}):

{text_clip}

PREDICTED TAGS to verify:

{chr(10).join(pred_lines)}

For each tag_id above, output: {{"verdicts": [{{"tag_id": "...", "accept": true/false, "reason": "..."}}, ...]}}"""

    try:
        resp = await client.chat_async(
            vendor="kimi",
            messages=[
                {"role": "system", "content": JUDGE_PROMPT},
                {"role": "user", "content": user_prompt},
            ],
            response_format={"type": "json_object"},
            max_tokens=600,
            temperature=0.1,
        )
        if not resp.content:
            stats.n_judge_failed += 1
            return []
        data = json.loads(resp.content)
        verdicts = data.get("verdicts") or []
        result = []
        for v in verdicts:
            if not isinstance(v, dict):
                continue
            tid = v.get("tag_id")
            accept = v.get("accept")
            reason = v.get("reason", "")
            if tid is None or accept is None:
                continue
            stats.n_total_tags += 1
            if accept:
                stats.n_accepted += 1
            else:
                stats.n_rejected += 1
                if len(stats.rejected_examples) < 10:
                    stats.rejected_examples.append({
                        "review_id": rid,
                        "tag_id": tid,
                        "tag_en": dict_meta.get(tid, {}).get("tag_en", ""),
                        "tag_cn": dict_meta.get(tid, {}).get("tag_cn", ""),
                        "reason": reason,
                        "text_preview": text_clip[:200],
                    })
            result.append({"tag_id": tid, "accept": bool(accept), "reason": reason})
        return result
    except Exception:
        stats.n_judge_failed += 1
        return []


async def spot_check_source(
    client,
    source_name: str,
    relabel_idx: dict[str, list[dict[str, Any]]],
    base_text: dict[str, dict[str, Any]],
    dict_meta: dict[str, dict[str, Any]],
    n_samples: int,
    seed: int = 42,
) -> SourceStats:
    rng = random.Random(seed)
    candidate_ids = list(relabel_idx.keys())
    if len(candidate_ids) > n_samples:
        sampled_ids = rng.sample(candidate_ids, n_samples)
    else:
        sampled_ids = candidate_ids

    stats = SourceStats(n_samples=len(sampled_ids))
    sem = asyncio.Semaphore(10)

    async def process(rid: str) -> None:
        async with sem:
            ctx = base_text.get(rid)
            if not ctx:
                stats.n_judge_failed += 1
                return
            await judge_one(client, rid, ctx["text"], relabel_idx[rid], dict_meta, stats)

    print(f"  ⏳ {source_name}: judging {len(sampled_ids)} samples", file=sys.stderr)
    t0 = time.time()
    await asyncio.gather(*(process(rid) for rid in sampled_ids))
    elapsed = time.time() - t0
    precision = stats.n_accepted / max(stats.n_total_tags, 1)
    print(
        f"  ✅ {source_name}: {stats.n_total_tags} tags judged, "
        f"precision={precision:.3f} ({stats.n_accepted}/{stats.n_total_tags}), "
        f"failed={stats.n_judge_failed}, elapsed={elapsed:.1f}s",
        file=sys.stderr,
    )
    return stats


def render_report(
    results: dict[str, SourceStats],
    n_samples_per_source: int,
) -> str:
    md: list[str] = []
    p = md.append
    p("---")
    p("name: phase6-d7-llm-spot-check")
    p("description: Phase 6 D7 #3 LLM 输出抽样质量评估 — 用 Kimi 独立判定 F4/F3 LLM 重打的 tag 是否被原文支撑。"
      "当审计 trustpilot 多语言 / zendesk / amazon 三批 LLM 输出可信度时使用。")
    p(f"date: {datetime.now().strftime('%Y-%m-%d')}")
    p("phase: phase6")
    p("day: D7")
    p("doc_type: audit-report")
    p("module: voc-nlp")
    p("---")
    p("")
    p("# Phase 6 D7 #3 LLM 输出抽样质量评估")
    p("")
    p(f"- 每源抽样：{n_samples_per_source} records")
    p(f"- 判官：Kimi (独立第二意见，避免与 DeepSeek 同模型偏见)")
    p(f"- 时间：{datetime.now().isoformat(timespec='seconds')}")
    p("")
    p("## 一、Per-source Precision")
    p("")
    p("| 数据源 | samples | tags | accepted | rejected | judge_failed | **precision** | 评级 |")
    p("|---|---:|---:|---:|---:|---:|---:|:---:|")
    for name, s in results.items():
        prec = s.n_accepted / max(s.n_total_tags, 1)
        if prec >= 0.90:
            verdict = "🟢 优"
        elif prec >= 0.85:
            verdict = "🟢 良"
        elif prec >= 0.75:
            verdict = "🟡 中"
        else:
            verdict = "🔴 差"
        p(f"| {name} | {s.n_samples} | {s.n_total_tags} | {s.n_accepted} | {s.n_rejected} | "
          f"{s.n_judge_failed} | **{prec:.3f}** | {verdict} |")
    p("")
    p("> 阈值：precision ≥ 0.85 视为可信；< 0.85 标记质量风险")
    p("")
    p("## 二、Rejected 样本（每源最多 10 条）")
    p("")
    for name, s in results.items():
        p(f"### {name}")
        p("")
        if not s.rejected_examples:
            p("（无 rejected 样本）")
            p("")
            continue
        p("| review_id | tag_id (cn/en) | Kimi 判官 reason | text 截取 |")
        p("|---|---|---|---|")
        for ex in s.rejected_examples:
            tag_label = f"{ex['tag_id']} ({ex['tag_cn']}/{ex['tag_en']})"
            reason = (ex.get("reason") or "").replace("|", "\\|").replace("\n", " ")[:120]
            text = (ex.get("text_preview") or "").replace("|", "\\|").replace("\n", " ")[:120]
            p(f"| {ex['review_id']} | {tag_label} | {reason} | {text} |")
        p("")
    p("## 三、综合评估")
    p("")
    overall_acc = sum(s.n_accepted for s in results.values())
    overall_total = sum(s.n_total_tags for s in results.values())
    overall_prec = overall_acc / max(overall_total, 1)
    p(f"- **整体 precision** = {overall_acc}/{overall_total} = **{overall_prec:.3f}**")
    p(f"- 阈值：≥ 0.85 视为通过")
    if overall_prec >= 0.85:
        p(f"- 判定：✅ **通过** — F4/F3 LLM 输出整体质量可信")
    elif overall_prec >= 0.75:
        p(f"- 判定：🟡 **边缘** — 整体可用但建议重点检查弱源")
    else:
        p(f"- 判定：🔴 **不达标** — 需重打或调整 prompt")
    p("")
    return "\n".join(md)


async def run(args: argparse.Namespace) -> dict[str, SourceStats]:
    from llm_client import LLMClient
    client = LLMClient()

    print(f"⏳ Loading dict meta", file=sys.stderr)
    dict_meta = load_dict_meta(args.dict)
    print(f"   {len(dict_meta)} tags loaded", file=sys.stderr)

    sources = {
        "trustpilot (D4 F4)": args.relabel_trustpilot,
        "zendesk (D5 F3)": args.relabel_zendesk,
        "amazon_competitor (D5 F3)": args.relabel_amazon,
    }

    print(f"⏳ Loading relabel indexes", file=sys.stderr)
    indexes = {name: load_relabel_index(p) for name, p in sources.items()}
    for name, idx in indexes.items():
        print(f"   {name}: {len(idx)} records with labels", file=sys.stderr)

    rng_global = random.Random(42)
    sample_ids_per_source = {}
    all_target_ids: set[str] = set()
    for name, idx in indexes.items():
        ids = list(idx.keys())
        if len(ids) > args.samples_per_source:
            sampled = rng_global.sample(ids, args.samples_per_source)
        else:
            sampled = ids
        sample_ids_per_source[name] = sampled
        all_target_ids.update(sampled)

    print(f"⏳ Resolving text for {len(all_target_ids)} sample ids from base", file=sys.stderr)
    base_text = collect_text_for_ids(args.base, all_target_ids)
    print(f"   resolved text for {len(base_text)} ids", file=sys.stderr)

    results: dict[str, SourceStats] = {}
    for name in indexes:
        sampled_ids = sample_ids_per_source[name]
        partial_idx = {rid: indexes[name][rid] for rid in sampled_ids}
        stats = await spot_check_source(
            client, name, partial_idx, base_text, dict_meta, len(sampled_ids),
        )
        results[name] = stats
    return results


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(description="Phase 6 D7 #3 LLM 输出抽样质量评估")
    ap.add_argument("--base", required=True, type=Path)
    ap.add_argument("--relabel-trustpilot", required=True, type=Path)
    ap.add_argument("--relabel-zendesk", required=True, type=Path)
    ap.add_argument("--relabel-amazon", required=True, type=Path)
    ap.add_argument("--dict", required=True, type=Path)
    ap.add_argument("--report", required=True, type=Path)
    ap.add_argument("--samples-per-source", type=int, default=50)
    args = ap.parse_args(argv)

    for arg_name in ("base", "relabel_trustpilot", "relabel_zendesk", "relabel_amazon", "dict"):
        path = getattr(args, arg_name)
        if not path.is_file():
            print(f"❌ {arg_name} not found: {path}", file=sys.stderr); return 2

    results = asyncio.run(run(args))

    args.report.parent.mkdir(parents=True, exist_ok=True)
    args.report.write_text(render_report(results, args.samples_per_source), encoding="utf-8")
    print(f"\n📄 Report: {args.report}", file=sys.stderr)

    overall_acc = sum(s.n_accepted for s in results.values())
    overall_total = sum(s.n_total_tags for s in results.values())
    overall_prec = overall_acc / max(overall_total, 1)
    print(f"📊 Overall precision: {overall_prec:.3f} ({overall_acc}/{overall_total})", file=sys.stderr)

    return 0 if overall_prec >= 0.85 else 1


if __name__ == "__main__":
    sys.exit(main())
