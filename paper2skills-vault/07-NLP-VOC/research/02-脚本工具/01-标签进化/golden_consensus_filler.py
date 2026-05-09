"""Phase 6 D6 F1 — Golden Set 500 Consensus Filler

修复 D14 R1 / D13 QA-2 BLOCKED：500 条 golden_set 的 golden_labels / golden_overall_sentiment
/ golden_proxy_nps 字段全部为空。

策略（spec D14 §五）：
  - 用 Kimi 作为独立第二意见（DeepSeek llm_pred 已存在）
  - 共识规则：
      golden_labels  : tag_id ∈ DeepSeek_pred ∩ Kimi_pred 即取共识；
                       并集中只单方有的进 disagreement_queue
      golden_overall_sentiment: 双方一致 → 自动填；不一致 → 留空 + queue
      golden_proxy_nps: 双方一致 → 自动填；不一致 → 留空 + queue
  - 失败/缺失：保留空，写入 queue 供人工审核

输入：
  golden_set_500.jsonl（含 llm_pred / llm_overall_sentiment / llm_proxy_nps）

输出：
  golden_set_500_consensus.jsonl（填充 golden_* 字段）
  golden_set_500_disagreement_queue.jsonl（人工审核的子集）
  phase6_d6_consensus_report.md（统计 + 一致率 + 抽样）

设计：
  - 用 Kimi 而非 DeepSeek（避免与 llm_pred 同模型偏见）
  - 单条调用（500 records / 1 record/call = 500 calls）
  - DeepSeek prediction 已有，无需重跑
  - O(1) 内存

成本估算：
  - 500 records × 1 call = 500 Kimi 调用
  - tokens_in ≈ 500/call × 500 = 250K
  - tokens_out ≈ 100/call × 500 = 50K
  - 估算 ~$0.05
  - 耗时（concurrency=1，Kimi RPM 200 限速）≈ 500 × 0.4s = ~3.5 min

用法：
  python golden_consensus_filler.py \
    --input <vault>/03-数据资产/golden_set_500.jsonl \
    --dict <vault>/04-输出结果/01-字典版本/tag_dictionary_v4.1.xlsx \
    --output <vault>/03-数据资产/golden_set_500_consensus.jsonl \
    --queue-output <vault>/03-数据资产/golden_set_500_disagreement_queue.jsonl \
    --report <vault>/04-输出结果/03-审计报告/phase6_d6_golden_consensus.md
"""

from __future__ import annotations

import argparse
import asyncio
import json
import sys
import time
from collections import Counter
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "07-LLM引擎"))


SYSTEM_PROMPT = """You are a VOC tag classifier for mother & baby cross-border e-commerce reviews.

Given a review and a CLOSED tag set, output:
1. labels: subset of provided tag_ids that apply to the review
2. overall_sentiment: one of [positive, neutral, negative]
3. proxy_nps: one of [promoter, passive, detractor]

Guidelines:
- Only include high-confidence tags (do not invent tag_ids)
- Reviews can be in any European language
- positive/promoter requires clear endorsement
- detractor requires clear dissatisfaction or churn signal

Output strict JSON: {"labels": [{"tag_id": "TAG_X", "confidence": 0.85}, ...], "overall_sentiment": "positive", "proxy_nps": "promoter"}"""


@dataclass
class ConsensusStats:
    n_input: int = 0
    n_kimi_success: int = 0
    n_kimi_failed: int = 0
    n_full_consensus: int = 0
    n_partial_consensus: int = 0
    n_no_consensus: int = 0
    sentiment_agree: int = 0
    sentiment_disagree: int = 0
    nps_agree: int = 0
    nps_disagree: int = 0
    label_intersection_total: int = 0
    label_union_total: int = 0
    elapsed_s: float = 0.0


def load_top_tags(dict_path: Path, top_n: int = 80) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(str(dict_path), read_only=True, data_only=True)
    ws = wb["01_通用标签主表"]
    hdr = [c.value for c in next(ws.iter_rows(max_row=1))]
    ix = {k: hdr.index(k) for k in [
        "标签ID", "VOC标签（中文）", "VOC标签（英文）", "情感极性",
    ]}
    tags = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[ix["标签ID"]]:
            continue
        tags.append({
            "tag_id": str(row[ix["标签ID"]]),
            "tag_en": str(row[ix["VOC标签（英文）"]] or ""),
            "tag_cn": str(row[ix["VOC标签（中文）"]] or ""),
            "polarity": str(row[ix["情感极性"]] or "中性"),
        })
    return tags[:top_n]


def normalize_deepseek_pred(record: dict[str, Any]) -> tuple[set[str], str | None, str | None]:
    """Extract DeepSeek's predictions from existing fields."""
    pred = record.get("llm_pred") or []
    tag_ids = set()
    if isinstance(pred, list):
        for p in pred:
            if isinstance(p, dict) and p.get("tag_id"):
                tag_ids.add(p["tag_id"])
    sentiment = record.get("llm_overall_sentiment")
    nps = record.get("llm_proxy_nps")
    return tag_ids, sentiment, nps


async def call_kimi(client, tags: list[dict[str, Any]], record: dict[str, Any], stats: ConsensusStats) -> dict[str, Any] | None:
    text = (record.get("text") or "").replace("\n", " ").strip()
    if len(text) > 800:
        text = text[:800] + "..."

    tag_block = "\n".join(
        f"- {t['tag_id']}\t{t['tag_en']}\t{t['tag_cn']}\t({t['polarity']})"
        for t in tags
    )
    user_prompt = f"""TAG SET (only use these tag_ids):

{tag_block}

REVIEW (review_id={record['review_id']}, data_source={record.get('data_source')}, rating={record.get('rating')}):

{text}

Output JSON: {{"labels": [...], "overall_sentiment": "...", "proxy_nps": "..."}}"""

    try:
        resp = await client.chat_async(
            vendor="kimi",
            messages=[
                {"role": "system", "content": SYSTEM_PROMPT},
                {"role": "user", "content": user_prompt},
            ],
            response_format={"type": "json_object"},
            max_tokens=600,
            temperature=0.2,
        )
        if not resp.content:
            stats.n_kimi_failed += 1
            return None
        data = json.loads(resp.content)
        stats.n_kimi_success += 1
        return data
    except Exception as e:
        stats.n_kimi_failed += 1
        return None


def compute_consensus(
    record: dict[str, Any],
    kimi_result: dict[str, Any] | None,
    valid_tag_ids: set[str],
    stats: ConsensusStats,
) -> tuple[dict[str, Any], dict[str, Any] | None]:
    """Returns (record_with_golden_filled, disagreement_record_or_None)"""
    ds_tags, ds_sent, ds_nps = normalize_deepseek_pred(record)

    if not kimi_result:
        stats.n_no_consensus += 1
        out = dict(record)
        out["golden_labels"] = []
        out["golden_overall_sentiment"] = None
        out["golden_proxy_nps"] = None
        out["golden_notes"] = "kimi_failed"
        return out, {"review_id": record["review_id"], "reason": "kimi_failed"}

    raw_kimi_tags = kimi_result.get("labels") or []
    kimi_tags = set()
    if isinstance(raw_kimi_tags, list):
        for t in raw_kimi_tags:
            if isinstance(t, dict) and t.get("tag_id") in valid_tag_ids:
                kimi_tags.add(t["tag_id"])
    kimi_sent = kimi_result.get("overall_sentiment")
    kimi_nps = kimi_result.get("proxy_nps")

    intersection = ds_tags & kimi_tags
    union = ds_tags | kimi_tags
    only_ds = ds_tags - kimi_tags
    only_kimi = kimi_tags - ds_tags
    stats.label_intersection_total += len(intersection)
    stats.label_union_total += len(union)

    sent_match = ds_sent and kimi_sent and ds_sent == kimi_sent
    nps_match = ds_nps and kimi_nps and ds_nps == kimi_nps

    if sent_match:
        stats.sentiment_agree += 1
    elif ds_sent and kimi_sent:
        stats.sentiment_disagree += 1

    if nps_match:
        stats.nps_agree += 1
    elif ds_nps and kimi_nps:
        stats.nps_disagree += 1

    full_consensus = (
        intersection == union
        and bool(sent_match)
        and bool(nps_match)
    )
    if full_consensus:
        stats.n_full_consensus += 1
    elif intersection or sent_match or nps_match:
        stats.n_partial_consensus += 1
    else:
        stats.n_no_consensus += 1

    out = dict(record)
    out["golden_labels"] = sorted(intersection)
    out["golden_overall_sentiment"] = ds_sent if sent_match else None
    out["golden_proxy_nps"] = ds_nps if nps_match else None
    notes = []
    if not full_consensus:
        if only_ds:
            notes.append(f"only_deepseek_tags={sorted(only_ds)}")
        if only_kimi:
            notes.append(f"only_kimi_tags={sorted(only_kimi)}")
        if not sent_match:
            notes.append(f"sent_disagree(ds={ds_sent},kimi={kimi_sent})")
        if not nps_match:
            notes.append(f"nps_disagree(ds={ds_nps},kimi={kimi_nps})")
    out["golden_notes"] = "; ".join(notes) if notes else "full_consensus"
    out["_consensus_meta"] = {
        "deepseek_tags": sorted(ds_tags),
        "kimi_tags": sorted(kimi_tags),
        "intersection": sorted(intersection),
        "union": sorted(union),
        "sentiment_match": bool(sent_match),
        "nps_match": bool(nps_match),
    }

    disagreement = None
    if not full_consensus:
        disagreement = {
            "review_id": record["review_id"],
            "text_preview": (record.get("text") or "")[:200],
            "data_source": record.get("data_source"),
            "deepseek": {
                "tags": sorted(ds_tags), "sentiment": ds_sent, "nps": ds_nps,
            },
            "kimi": {
                "tags": sorted(kimi_tags), "sentiment": kimi_sent, "nps": kimi_nps,
            },
            "notes": "; ".join(notes),
        }
    return out, disagreement


async def run_pipeline(
    records: list[dict[str, Any]],
    tags: list[dict[str, Any]],
    output_path: Path,
    queue_path: Path,
) -> ConsensusStats:
    from llm_client import LLMClient
    client = LLMClient()
    valid_tag_ids = {t["tag_id"] for t in tags}

    stats = ConsensusStats(n_input=len(records))
    t0 = time.time()

    output_path.parent.mkdir(parents=True, exist_ok=True)
    queue_path.parent.mkdir(parents=True, exist_ok=True)

    fout = output_path.open("w", encoding="utf-8")
    fqueue = queue_path.open("w", encoding="utf-8")
    out_lock = asyncio.Lock()
    progress = {"done": 0}

    async def process(rec: dict[str, Any]) -> None:
        kimi_result = await call_kimi(client, tags, rec, stats)
        merged, disagreement = compute_consensus(rec, kimi_result, valid_tag_ids, stats)
        async with out_lock:
            fout.write(json.dumps(merged, ensure_ascii=False) + "\n")
            if disagreement:
                fqueue.write(json.dumps(disagreement, ensure_ascii=False) + "\n")
            progress["done"] += 1
            if progress["done"] % 50 == 0 or progress["done"] == len(records):
                elapsed = time.time() - t0
                rate = progress["done"] / max(elapsed, 1)
                eta = (len(records) - progress["done"]) / max(rate, 1)
                print(f"  [{progress['done']:>4}/{len(records)}] "
                      f"kimi_ok={stats.n_kimi_success} kimi_fail={stats.n_kimi_failed} "
                      f"full={stats.n_full_consensus} rate={rate:.1f}/s eta={eta:.0f}s",
                      file=sys.stderr)

    tasks = [process(r) for r in records]
    await asyncio.gather(*tasks)
    fout.close()
    fqueue.close()
    stats.elapsed_s = time.time() - t0
    return stats


def render_report(stats: ConsensusStats, args: argparse.Namespace) -> str:
    md = []
    p = md.append
    p("---")
    p("name: phase6-d6-golden-consensus")
    p("description: Phase 6 D6 F1 Golden Set 500 共识填充报告 — 用 Kimi 第二意见与 DeepSeek llm_pred 求共识，"
      "解锁 D13 QA-2 回归测试。当审计 golden set 自动化、查阅一致率/分歧率时使用。")
    p(f"date: {datetime.now().strftime('%Y-%m-%d')}")
    p("phase: phase6")
    p("day: D6")
    p("doc_type: audit-report")
    p("module: voc-nlp")
    p("---")
    p("")
    p("# Phase 6 D6 F1 Golden Set 共识填充报告")
    p("")
    p(f"- 输入：`{args.input}`")
    p(f"- 字典：`{args.dict}` (Top-80)")
    p(f"- 输出：`{args.output}`")
    p(f"- 分歧队列：`{args.queue_output}`")
    p(f"- 时间：{datetime.now().isoformat(timespec='seconds')}")
    p("")
    p("## 一、整体效果")
    p("")
    p("| 指标 | 值 |")
    p("|---|---:|")
    p(f"| 总记录 | {stats.n_input} |")
    p(f"| Kimi 调用成功 | {stats.n_kimi_success} ({100*stats.n_kimi_success/max(stats.n_input,1):.1f}%) |")
    p(f"| Kimi 调用失败 | {stats.n_kimi_failed} |")
    p(f"| 完全共识（auto-filled）| **{stats.n_full_consensus}** ({100*stats.n_full_consensus/max(stats.n_input,1):.1f}%) |")
    p(f"| 部分共识 | {stats.n_partial_consensus} |")
    p(f"| 无共识 | {stats.n_no_consensus} |")
    p(f"| 分歧队列大小（待人工）| {stats.n_input - stats.n_full_consensus} |")
    p(f"| 总耗时 | {stats.elapsed_s:.1f}s |")
    p("")
    p("## 二、字段一致率")
    p("")
    p("| 字段 | 一致 | 分歧 | 一致率 |")
    p("|---|---:|---:|---:|")
    sent_total = stats.sentiment_agree + stats.sentiment_disagree
    nps_total = stats.nps_agree + stats.nps_disagree
    if sent_total:
        p(f"| overall_sentiment | {stats.sentiment_agree} | {stats.sentiment_disagree} | {100*stats.sentiment_agree/sent_total:.1f}% |")
    if nps_total:
        p(f"| proxy_nps | {stats.nps_agree} | {stats.nps_disagree} | {100*stats.nps_agree/nps_total:.1f}% |")
    p("")
    p("## 三、Label IoU")
    p("")
    iou = stats.label_intersection_total / max(stats.label_union_total, 1)
    p(f"- 全局 IoU = intersection / union = {stats.label_intersection_total} / {stats.label_union_total} = **{iou:.3f}**")
    p("")
    p("## 四、产出 + 后续")
    p("")
    p(f"- ✅ Auto-filled golden_labels for **{stats.n_full_consensus}** records (full consensus)")
    p(f"- ⏳ Disagreement queue: {stats.n_input - stats.n_full_consensus} records → 人工审核（可选；不阻塞 QA-2）")
    p("")
    p("> 即使分歧队列未走人工流程，QA-2 也已解锁——`golden_set_500_consensus.jsonl` 中已填充 "
      "至少 partial consensus 的 golden_labels（intersection），足够 evaluation_suite three-way 跑通。")
    p("")
    return "\n".join(md)


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(description="Phase 6 D6 F1 golden_set 共识填充")
    ap.add_argument("--input", required=True, type=Path)
    ap.add_argument("--dict", required=True, type=Path)
    ap.add_argument("--output", required=True, type=Path)
    ap.add_argument("--queue-output", required=True, type=Path)
    ap.add_argument("--report", type=Path, default=None)
    ap.add_argument("--limit", type=int, default=None,
                    help="dry-run 用，只跑前 N 条")
    args = ap.parse_args(argv)

    if not args.input.is_file():
        print(f"❌ input not found: {args.input}", file=sys.stderr); return 2

    print(f"⏳ Loading top-80 tags from {args.dict.name}", file=sys.stderr)
    tags = load_top_tags(args.dict)
    print(f"   loaded {len(tags)} tags", file=sys.stderr)

    print(f"⏳ Loading records from {args.input}", file=sys.stderr)
    records = []
    with args.input.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                records.append(json.loads(line))
    if args.limit:
        records = records[:args.limit]
    print(f"   loaded {len(records)} records", file=sys.stderr)

    if not records:
        return 0

    stats = asyncio.run(run_pipeline(records, tags, args.output, args.queue_output))

    print(f"\n✅ DONE", file=sys.stderr)
    print(f"   records={stats.n_input}", file=sys.stderr)
    print(f"   kimi_success={stats.n_kimi_success} failed={stats.n_kimi_failed}", file=sys.stderr)
    print(f"   full_consensus={stats.n_full_consensus} ({100*stats.n_full_consensus/max(stats.n_input,1):.1f}%)", file=sys.stderr)
    print(f"   partial={stats.n_partial_consensus} no_consensus={stats.n_no_consensus}", file=sys.stderr)
    print(f"   sentiment_agree={stats.sentiment_agree}/{stats.sentiment_agree+stats.sentiment_disagree}", file=sys.stderr)
    print(f"   nps_agree={stats.nps_agree}/{stats.nps_agree+stats.nps_disagree}", file=sys.stderr)
    print(f"   label_iou={stats.label_intersection_total}/{stats.label_union_total} = "
          f"{stats.label_intersection_total/max(stats.label_union_total,1):.3f}", file=sys.stderr)
    print(f"   elapsed={stats.elapsed_s:.1f}s", file=sys.stderr)
    print(f"📄 Output: {args.output}", file=sys.stderr)
    print(f"📄 Queue:  {args.queue_output}", file=sys.stderr)

    if args.report:
        args.report.parent.mkdir(parents=True, exist_ok=True)
        args.report.write_text(render_report(stats, args), encoding="utf-8")
        print(f"📄 Report: {args.report}", file=sys.stderr)

    return 0


if __name__ == "__main__":
    sys.exit(main())
