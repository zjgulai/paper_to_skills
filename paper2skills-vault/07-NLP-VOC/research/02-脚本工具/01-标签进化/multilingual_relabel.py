"""Phase 6 D4 F4 — Multilingual Trustpilot Relabeler

修复 Week 2 Gate #10/#11 (raw 76.11% < 88% / eff 89.48% < 94%)。

根因：trustpilot 28K zero-tag 中 ~20K 是非英文（fr/de/es）
  - fr: 11,181 (39.7%)
  - de: 8,534 (30.3%)
  - en: 8,212 (29.2%) — 也尝试，可能字典未覆盖
  - es / ru / 其他: < 250

策略：
  - 用 v4.1 字典 Top-100 通用标签作为 LLM 闭集
  - DeepSeek json_object mode，支持多语言理解（已验证）
  - 提示词：明确告知 LLM 输入可能是 法/德/西/俄 等非英文，输出标签 ID
  - 异步并发 = 40（DeepSeek max_concurrent）

实现要点：
  - 流式输入，分批 50 条/请求（节省 token）
  - 输出 JSONL：{review_id, labels: [{tag_id, confidence}], _llm_meta}
  - 失败队列单独写文件，方便复跑
  - 进度条 + tps stats

用法：
  python multilingual_relabel.py \
    --input <vault>/04-输出结果/unified_labeling/phase5_full_labeled.jsonl \
    --dict <vault>/04-输出结果/01-字典版本/tag_dictionary_v4.1.xlsx \
    --output <vault>/04-输出结果/unified_labeling/phase6_multilingual_relabel.jsonl \
    --report <vault>/04-输出结果/03-审计报告/phase6_d4_multilingual.md \
    [--limit N] [--data-source trustpilot] [--zero-only]

成本估算（28K records, batch=50 → 560 calls）：
  - tokens_in ~ 4K/call × 560 = ~2.2M
  - tokens_out ~ 0.5K/call × 560 = ~280K
  - DeepSeek pricing 约 $0.5
  - 耗时（concurrency=40, ~5s/call）≈ 70s × 1.4 = ~70 min（保守）
"""

from __future__ import annotations

import argparse
import asyncio
import json
import sys
import time
from collections import Counter
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "07-LLM引擎"))


BATCH_SIZE = 10
TOP_TAGS_FOR_PROMPT = 80
SYSTEM_PROMPT = """You are a multilingual VOC tag classifier for a mother & baby cross-border e-commerce platform.

Reviews can be in English, French, German, Spanish, Italian, Russian, or other European languages.
You must classify each review against a CLOSED set of tags (provided in user message).

For each review:
1. Read the review text (regardless of language)
2. Identify ALL relevant tags from the closed set
3. Return tag_id list with confidence (0.5-0.95) per tag
4. Prefer 0-3 tags per review; only include high-confidence matches

Output strict JSON format:
{"results": [
  {"review_id": "<rid>", "labels": [{"tag_id": "TAG_GEN_E002", "confidence": 0.85}, ...]},
  ...
]}

If no tag matches, return empty labels array. Do not invent tag_ids not in the closed set."""


@dataclass
class RelabelStats:
    n_input: int = 0
    n_skipped_already_labeled: int = 0
    n_processed: int = 0
    n_with_new_labels: int = 0
    n_zero_labels_returned: int = 0
    n_llm_calls: int = 0
    n_llm_failed: int = 0
    n_total_new_labels: int = 0
    tag_counter: Counter = field(default_factory=Counter)
    elapsed_s: float = 0.0


def load_top_tags(dict_path: Path, top_n: int = TOP_TAGS_FOR_PROMPT) -> list[dict[str, Any]]:
    """从 v4.1 字典加载 Top-N 通用标签，用于 LLM closed-set prompt"""
    wb = openpyxl.load_workbook(str(dict_path), read_only=True, data_only=True)
    ws = wb["01_通用标签主表"]
    hdr = [c.value for c in next(ws.iter_rows(max_row=1))]
    ix = {k: hdr.index(k) for k in [
        "标签ID", "VOC标签（中文）", "VOC标签（英文）", "情感极性", "AIPL节点",
    ]}
    tags = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[ix["标签ID"]]:
            continue
        tags.append({
            "tag_id": str(row[ix["标签ID"]]),
            "tag_en": str(row[ix["VOC标签（英文）"]] or ""),
            "tag_cn": str(row[ix["VOC标签（中文）"]] or ""),
            "polarity": str(row[ix["情感极性"]] or "中性"),
            "aipl": str(row[ix["AIPL节点"]] or ""),
        })
    return tags[:top_n]


def filter_records(
    input_path: Path,
    data_source: str | None,
    zero_only: bool,
    limit: int | None,
) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    with input_path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                r = json.loads(line)
            except json.JSONDecodeError:
                continue
            if data_source and r.get("data_source") != data_source:
                continue
            if zero_only and (r.get("n_tags") or 0) > 0:
                continue
            out.append(r)
            if limit and len(out) >= limit:
                break
    return out


def build_user_prompt(tags: list[dict[str, Any]], batch: list[dict[str, Any]]) -> str:
    tag_lines = []
    for t in tags:
        tag_lines.append(f"- {t['tag_id']}\t{t['tag_en']}\t{t['tag_cn']}\t({t['polarity']}, AIPL={t['aipl']})")
    tag_block = "\n".join(tag_lines)

    review_lines = []
    for r in batch:
        text = (r.get("text") or "").replace("\n", " ").strip()
        if len(text) > 600:
            text = text[:600] + "..."
        review_lines.append(f"<review id=\"{r['review_id']}\">\n{text}\n</review>")
    review_block = "\n\n".join(review_lines)

    return f"""CLOSED TAG SET (Top-{len(tags)}; tag_id\\ttag_en\\ttag_cn\\t(polarity, AIPL)):

{tag_block}

CLASSIFY EACH REVIEW BELOW (multilingual; map to the tag_ids above):

{review_block}

Return JSON with key \"results\" array of {{review_id, labels:[{{tag_id, confidence}}]}}."""


async def relabel_batch(
    client,
    vendor: str,
    tags: list[dict[str, Any]],
    batch: list[dict[str, Any]],
    valid_tag_ids: set[str],
    stats: RelabelStats,
) -> list[dict[str, Any]]:
    """Returns list of {review_id, labels[], _llm_meta} for the batch."""
    user_prompt = build_user_prompt(tags, batch)
    try:
        resp = await client.chat_async(
            vendor=vendor,
            messages=[
                {"role": "system", "content": SYSTEM_PROMPT},
                {"role": "user", "content": user_prompt},
            ],
            response_format={"type": "json_object"},
            max_tokens=4000,
            temperature=0.2,
        )
        stats.n_llm_calls += 1
        data = json.loads(resp.content)
        results = data.get("results") or []
        results_by_id = {r.get("review_id"): r for r in results if isinstance(r, dict)}

        out = []
        for record in batch:
            rid = record["review_id"]
            llm_record = results_by_id.get(rid)
            if not llm_record:
                out.append({
                    "review_id": rid,
                    "labels": [],
                    "_llm_meta": {"vendor": vendor, "no_match": True},
                })
                stats.n_zero_labels_returned += 1
                continue
            raw_labels = llm_record.get("labels") or []
            clean_labels = []
            for lbl in raw_labels:
                if not isinstance(lbl, dict):
                    continue
                tid = lbl.get("tag_id")
                conf = lbl.get("confidence")
                if tid not in valid_tag_ids:
                    continue
                if not isinstance(conf, (int, float)):
                    continue
                clean_labels.append({
                    "tag_id": tid,
                    "confidence": round(min(0.95, max(0.5, float(conf))), 4),
                })
                stats.tag_counter[tid] += 1
            if clean_labels:
                stats.n_with_new_labels += 1
                stats.n_total_new_labels += len(clean_labels)
            else:
                stats.n_zero_labels_returned += 1
            out.append({
                "review_id": rid,
                "labels": clean_labels,
                "_llm_meta": {
                    "vendor": vendor,
                    "model": resp.model_used,
                    "latency_ms": round(resp.latency_ms),
                },
            })
        return out
    except Exception as e:
        stats.n_llm_failed += 1
        return [{
            "review_id": r["review_id"],
            "labels": [],
            "_llm_meta": {"vendor": vendor, "error": str(e)[:200]},
        } for r in batch]


async def run_pipeline(
    records: list[dict[str, Any]],
    tags: list[dict[str, Any]],
    output_path: Path,
    vendor: str = "deepseek",
) -> RelabelStats:
    from llm_client import LLMClient
    client = LLMClient()
    valid_tag_ids = {t["tag_id"] for t in tags}

    stats = RelabelStats(n_input=len(records))
    t0 = time.time()

    output_path.parent.mkdir(parents=True, exist_ok=True)

    batches = [records[i:i + BATCH_SIZE] for i in range(0, len(records), BATCH_SIZE)]
    print(f"⏳ {len(records):,} records → {len(batches)} batches × {BATCH_SIZE}/batch", file=sys.stderr)
    print(f"   vendor={vendor} (chat_async owns its semaphore internally)", file=sys.stderr)

    out_lock = asyncio.Lock()
    fout = output_path.open("w", encoding="utf-8")
    progress = {"done": 0}

    async def process(batch_idx: int, batch: list[dict[str, Any]]) -> None:
        results = await relabel_batch(client, vendor, tags, batch, valid_tag_ids, stats)
        async with out_lock:
            for r in results:
                fout.write(json.dumps(r, ensure_ascii=False) + "\n")
            progress["done"] += 1
            stats.n_processed += len(batch)
            if progress["done"] % 10 == 0 or progress["done"] == len(batches):
                elapsed = time.time() - t0
                rate = stats.n_processed / max(elapsed, 1)
                eta = (stats.n_input - stats.n_processed) / max(rate, 1)
                print(
                    f"  [{progress['done']:>4}/{len(batches)}] "
                    f"records={stats.n_processed:,} "
                    f"with_labels={stats.n_with_new_labels:,} "
                    f"failed={stats.n_llm_failed} "
                    f"rate={rate:.1f}rec/s eta={eta:.0f}s",
                    file=sys.stderr,
                )

    tasks = [process(i, b) for i, b in enumerate(batches)]
    await asyncio.gather(*tasks)
    fout.close()
    stats.elapsed_s = time.time() - t0
    return stats


def render_report(stats: RelabelStats, args: argparse.Namespace) -> str:
    md = []
    p = md.append
    p("---")
    p("name: phase6-d4-multilingual-relabel")
    p("description: Phase 6 D4 F4 多语言重打报告 — trustpilot 多语言 zero-tag 通过 LLM 闭集分类补齐标签。"
      "当审计 #10/#11 改善效果、查看 LLM 调用 + 标签分布时使用。")
    p(f"date: {datetime.now().strftime('%Y-%m-%d')}")
    p("phase: phase6")
    p("day: D4")
    p("doc_type: audit-report")
    p("module: voc-nlp")
    p("---")
    p("")
    p("# Phase 6 D4 F4 多语言重打报告")
    p("")
    p(f"- 输入：`{args.input}`")
    p(f"- 字典：`{args.dict}` (Top-{TOP_TAGS_FOR_PROMPT})")
    p(f"- 输出：`{args.output}`")
    p(f"- 过滤：data_source={args.data_source}, zero_only={args.zero_only}")
    if args.limit:
        p(f"- 限制：limit={args.limit}")
    p(f"- 运行时间：{datetime.now().isoformat(timespec='seconds')}")
    p("")
    p("## 一、整体效果")
    p("")
    p("| 指标 | 值 |")
    p("|---|---:|")
    p(f"| 候选 records | {stats.n_input:,} |")
    p(f"| 已 LLM 处理 | {stats.n_processed:,} |")
    p(f"| 产出新标签的 records | **{stats.n_with_new_labels:,}** ({100*stats.n_with_new_labels/max(stats.n_processed,1):.1f}%) |")
    p(f"| LLM 返回零标签 records | {stats.n_zero_labels_returned:,} |")
    p(f"| 新增 labels 总数 | {stats.n_total_new_labels:,} |")
    p(f"| LLM 调用次数 | {stats.n_llm_calls:,} |")
    p(f"| LLM 失败次数 | {stats.n_llm_failed:,} |")
    p(f"| 总耗时 | {stats.elapsed_s:.1f}s |")
    p("")
    p("## 二、Top-20 标签分布（新增 labels）")
    p("")
    p("| Tag ID | 命中次数 |")
    p("|---|---:|")
    for tid, n in stats.tag_counter.most_common(20):
        p(f"| {tid} | {n:,} |")
    p("")
    return "\n".join(md)


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(description="Phase 6 D4 F4 多语言重打")
    ap.add_argument("--input", required=True, type=Path)
    ap.add_argument("--dict", required=True, type=Path)
    ap.add_argument("--output", required=True, type=Path)
    ap.add_argument("--report", type=Path, default=None)
    ap.add_argument("--limit", type=int, default=None,
                    help="限制处理记录数（dry-run 用）")
    ap.add_argument("--data-source", default=None,
                    help="过滤 data_source（如 trustpilot）")
    ap.add_argument("--zero-only", action="store_true",
                    help="仅处理 n_tags=0 的记录")
    ap.add_argument("--vendor", default="deepseek", choices=["deepseek", "kimi"])
    args = ap.parse_args(argv)

    if not args.input.is_file():
        print(f"❌ input not found: {args.input}", file=sys.stderr); return 2
    if not args.dict.is_file():
        print(f"❌ dict not found: {args.dict}", file=sys.stderr); return 2

    print(f"⏳ Loading top-{TOP_TAGS_FOR_PROMPT} tags from {args.dict.name}", file=sys.stderr)
    tags = load_top_tags(args.dict)
    print(f"   loaded {len(tags)} tags", file=sys.stderr)

    print(f"⏳ Filtering records: data_source={args.data_source}, zero_only={args.zero_only}", file=sys.stderr)
    records = filter_records(args.input, args.data_source, args.zero_only, args.limit)
    print(f"   filtered {len(records):,} candidate records", file=sys.stderr)

    if not records:
        print(f"⚠️ no records to process", file=sys.stderr)
        return 0

    stats = asyncio.run(run_pipeline(records, tags, args.output, vendor=args.vendor))

    print(f"\n✅ DONE", file=sys.stderr)
    print(f"   records={stats.n_input:,} processed={stats.n_processed:,}", file=sys.stderr)
    print(f"   with_labels={stats.n_with_new_labels:,} ({100*stats.n_with_new_labels/max(stats.n_processed,1):.1f}%)", file=sys.stderr)
    print(f"   total_new_labels={stats.n_total_new_labels:,}", file=sys.stderr)
    print(f"   llm_calls={stats.n_llm_calls} failed={stats.n_llm_failed}", file=sys.stderr)
    print(f"   elapsed={stats.elapsed_s:.1f}s ({stats.n_processed/max(stats.elapsed_s,1):.1f} rec/s)", file=sys.stderr)
    print(f"📄 Output: {args.output}", file=sys.stderr)

    if args.report:
        args.report.parent.mkdir(parents=True, exist_ok=True)
        args.report.write_text(render_report(stats, args), encoding="utf-8")
        print(f"📄 Report: {args.report}", file=sys.stderr)

    return 1 if stats.n_llm_failed > stats.n_llm_calls * 0.05 else 0


if __name__ == "__main__":
    sys.exit(main())
