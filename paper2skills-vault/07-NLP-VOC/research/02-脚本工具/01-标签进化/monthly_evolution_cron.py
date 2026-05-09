"""Phase 5 D12 — Monthly Evolution Cron (8-step pipeline)

每月 1 日 02:00 触发的字典进化闭环（决策 3，§9）：

  Step 1 zero_label_extractor   零标签 + 低置信度（<0.6）样本提取
  Step 2 closed_relabel         用最新字典对 Step 1 输出重跑
  Step 3 open_set_sampling      Step 2 仍零标签的 5% 采样供 LLM 自由生成
  Step 4 candidate_filter       三过滤（频率 ≥ 10 / Jaccard < 0.3 / LLM 相关性 ≥ 0.6）
  Step 5 alchemist_lf           通过候选 → Label Function（弱监督）
  Step 6 active_learning        不确定候选写入审核队列
  Step 7 dict_update            v_n → v_{n+1} 增量更新
  Step 8 bi_recompute           触发 dual_coverage / MAA / AGRS 重算

设计原则：
  - 完全可以 --dry-run（不调 LLM、不写正式字典；仅产出 *_dryrun.* 文件）
  - 8 步独立 exit code 可追踪；任意步失败即跳出
  - --tenant 注入路径根（默认 momcozy；v6.0 多租户前瞻）
  - 完成 / 失败均推飞书 webhook（如 ~/.paper2skills/feishu_webhook 存在）
"""

from __future__ import annotations

import argparse
import json
import os
import shlex
import shutil
import subprocess
import sys
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any


REPO_ROOT = Path("/Users/pray/project/paper_to_skills")
RESEARCH_ROOT_DEFAULT = REPO_ROOT / "paper2skills-vault/07-NLP-VOC/research"
FEISHU_WEBHOOK_FILE = Path("~/.paper2skills/feishu_webhook").expanduser()
DRY_RUN_LOG_DIR = Path("logs")
MAX_NEW_TAGS_PER_RUN = 10


@dataclass
class StepResult:
    name: str
    ok: bool
    duration_s: float
    detail: str = ""
    artifacts: list[str] = field(default_factory=list)
    skipped: bool = False


def _post_feishu(text: str) -> None:
    if not FEISHU_WEBHOOK_FILE.is_file():
        return
    try:
        webhook = FEISHU_WEBHOOK_FILE.read_text(encoding="utf-8").strip()
        if not webhook:
            return
        import urllib.request
        body = json.dumps({"msg_type": "text", "content": {"text": text}}).encode("utf-8")
        req = urllib.request.Request(
            webhook, data=body,
            headers={"Content-Type": "application/json"},
        )
        urllib.request.urlopen(req, timeout=10).read()
    except Exception as e:
        print(f"⚠️ feishu push failed: {e}", file=sys.stderr)


def _now() -> str:
    return datetime.now().isoformat(timespec="seconds")


def _safe_run(
    args: list[str],
    *,
    capture_log: Path | None = None,
    timeout: int = 1800,
) -> tuple[int, str, str]:
    print(f"  $ {' '.join(shlex.quote(a) for a in args)}", file=sys.stderr)
    proc = subprocess.run(args, capture_output=True, text=True, timeout=timeout)
    out_tail = proc.stdout[-2000:] if proc.stdout else ""
    err_tail = proc.stderr[-2000:] if proc.stderr else ""
    if capture_log:
        capture_log.parent.mkdir(parents=True, exist_ok=True)
        capture_log.write_text(
            f"$ {' '.join(args)}\n\n=== STDOUT ===\n{proc.stdout}\n=== STDERR ===\n{proc.stderr}",
            encoding="utf-8",
        )
    return proc.returncode, out_tail, err_tail


def step1_zero_label_extractor(
    input_jsonl: Path,
    out_jsonl: Path,
    *,
    dry_run: bool,
) -> StepResult:
    t0 = datetime.now()
    out_jsonl.parent.mkdir(parents=True, exist_ok=True)
    n_total = 0
    n_zero = 0
    n_low_conf = 0
    n_kept = 0
    sample_limit = 5000 if dry_run else 200_000
    with input_jsonl.open("r", encoding="utf-8") as fin, out_jsonl.open("w", encoding="utf-8") as fout:
        for line in fin:
            if not line.strip():
                continue
            n_total += 1
            try:
                r = json.loads(line)
            except json.JSONDecodeError:
                continue
            n_tags = r.get("n_tags") or 0
            labels = r.get("labels") or []
            avg_conf = (
                sum(float(l.get("confidence") or 0.0) for l in labels) / max(len(labels), 1)
                if labels else 0.0
            )
            is_zero = n_tags == 0
            is_low = labels and avg_conf < 0.6
            if is_zero:
                n_zero += 1
            if is_low:
                n_low_conf += 1
            if (is_zero or is_low) and n_kept < sample_limit:
                fout.write(line)
                n_kept += 1
    dur = (datetime.now() - t0).total_seconds()
    return StepResult(
        name="step1_zero_label_extractor",
        ok=True,
        duration_s=dur,
        detail=f"total={n_total:,} zero={n_zero:,} low_conf={n_low_conf:,} kept={n_kept:,}",
        artifacts=[str(out_jsonl)],
    )


def step2_closed_relabel(
    input_jsonl: Path,
    out_jsonl: Path,
    *,
    dry_run: bool,
) -> StepResult:
    """dry-run 不真跑 LLM。模拟：复制 input → output，记录"应跑"行数"""
    t0 = datetime.now()
    out_jsonl.parent.mkdir(parents=True, exist_ok=True)
    if dry_run:
        shutil.copyfile(input_jsonl, out_jsonl)
        n = sum(1 for _ in input_jsonl.open("r", encoding="utf-8"))
        return StepResult(
            name="step2_closed_relabel",
            ok=True, duration_s=(datetime.now() - t0).total_seconds(),
            detail=f"[dry-run] would relabel {n:,} records with latest dict",
            artifacts=[str(out_jsonl)],
            skipped=True,
        )
    return StepResult(
        name="step2_closed_relabel", ok=False,
        duration_s=(datetime.now() - t0).total_seconds(),
        detail="非 dry-run 模式需调 phase5_unified_labeler.py（生产环境实施）",
    )


def step3_open_set_sampling(
    input_jsonl: Path,
    out_jsonl: Path,
    *,
    sample_rate: float,
    dry_run: bool,
) -> StepResult:
    t0 = datetime.now()
    out_jsonl.parent.mkdir(parents=True, exist_ok=True)
    import random
    rng = random.Random(42)
    n_in = 0
    n_out = 0
    with input_jsonl.open("r", encoding="utf-8") as fin, out_jsonl.open("w", encoding="utf-8") as fout:
        for line in fin:
            if not line.strip():
                continue
            n_in += 1
            try:
                r = json.loads(line)
            except json.JSONDecodeError:
                continue
            if (r.get("n_tags") or 0) != 0:
                continue
            if rng.random() < sample_rate:
                fout.write(line)
                n_out += 1
    dur = (datetime.now() - t0).total_seconds()
    return StepResult(
        name="step3_open_set_sampling",
        ok=True, duration_s=dur,
        detail=f"in={n_in:,} sampled_zero_only={n_out:,} ({sample_rate*100:.1f}%)",
        artifacts=[str(out_jsonl)],
        skipped=dry_run,
    )


def step4_candidate_filter(
    sampled_jsonl: Path,
    dict_xlsx: Path,
    out_dir: Path,
    *,
    dry_run: bool,
) -> StepResult:
    t0 = datetime.now()
    out_dir.mkdir(parents=True, exist_ok=True)
    out_json = out_dir / ("candidate_filtered_dryrun.json" if dry_run else "candidate_filtered.json")
    filter_script = (
        Path(__file__).resolve().parent / "phase5_d9_filter.py"
    )
    if not filter_script.is_file():
        return StepResult(
            name="step4_candidate_filter", ok=False,
            duration_s=(datetime.now() - t0).total_seconds(),
            detail=f"phase5_d9_filter.py not found at {filter_script}",
        )
    if dry_run:
        out_json.write_text(json.dumps({
            "dry_run": True,
            "approved_candidates": [
                {"tag_en": "fake_candidate_1", "tag_cn": "演练候选1",
                 "support": 12, "jaccard_max": 0.18, "llm_relevance": 0.72},
            ],
            "rejected": [],
            "audit": {"input": str(sampled_jsonl), "dict": str(dict_xlsx)},
        }, ensure_ascii=False, indent=2), encoding="utf-8")
        return StepResult(
            name="step4_candidate_filter", ok=True,
            duration_s=(datetime.now() - t0).total_seconds(),
            detail="[dry-run] synthesized 1 fake candidate to exercise pipeline",
            artifacts=[str(out_json)],
            skipped=True,
        )
    return StepResult(
        name="step4_candidate_filter", ok=False,
        duration_s=(datetime.now() - t0).total_seconds(),
        detail="非 dry-run 模式需调用 phase5_d9_filter.py 完整三过滤（生产实施）",
    )


def step5_alchemist_lf(
    candidate_json: Path,
    out_dir: Path,
    *,
    dry_run: bool,
) -> StepResult:
    t0 = datetime.now()
    out_dir.mkdir(parents=True, exist_ok=True)
    out_py = out_dir / ("generated_lfs_dryrun.py" if dry_run else "generated_lfs.py")
    if dry_run:
        out_py.write_text(
            '"""ALCHEmist 弱监督 LF stub (dry-run)"""\n'
            '\n'
            'def lf_演练候选1(text: str) -> int:\n'
            '    """命中 → 1，未命中 → 0；非负向且置信度 > 0.6"""\n'
            '    return 1 if "fake_candidate_keyword" in text.lower() else 0\n',
            encoding="utf-8",
        )
        return StepResult(
            name="step5_alchemist_lf", ok=True,
            duration_s=(datetime.now() - t0).total_seconds(),
            detail="[dry-run] generated stub LF for 1 candidate",
            artifacts=[str(out_py)],
            skipped=True,
        )
    return StepResult(
        name="step5_alchemist_lf", ok=False,
        duration_s=(datetime.now() - t0).total_seconds(),
        detail="非 dry-run 模式需 alchemist_label_generator 生产 LF",
    )


def step6_active_learning(
    candidate_json: Path,
    out_jsonl: Path,
    *,
    dry_run: bool,
) -> StepResult:
    t0 = datetime.now()
    out_jsonl.parent.mkdir(parents=True, exist_ok=True)
    if dry_run:
        out_jsonl.write_text(json.dumps({
            "review_id": "synthetic_dryrun_001",
            "tag_en": "uncertain_candidate_demo",
            "tag_cn": "示例不确定候选",
            "uncertainty": 0.55,
            "reason": "LLM 相关性 0.55 - 0.65 边缘区间，转人工",
        }, ensure_ascii=False) + "\n", encoding="utf-8")
        return StepResult(
            name="step6_active_learning", ok=True,
            duration_s=(datetime.now() - t0).total_seconds(),
            detail="[dry-run] queued 1 synthetic uncertain candidate",
            artifacts=[str(out_jsonl)],
            skipped=True,
        )
    return StepResult(
        name="step6_active_learning", ok=False,
        duration_s=(datetime.now() - t0).total_seconds(),
        detail="非 dry-run 模式需 active_learning_queue.py 处理（生产实施）",
    )


def step7_dict_update(
    candidate_json: Path,
    base_dict_xlsx: Path,
    out_xlsx: Path,
    *,
    dry_run: bool,
) -> StepResult:
    """dry-run 时：复制 v4.0 → v4.1_dryrun，对 candidate 做最多 MAX_NEW_TAGS_PER_RUN 个新增（追加到 01_通用标签主表）"""
    t0 = datetime.now()
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    try:
        candidates = json.loads(candidate_json.read_text(encoding="utf-8"))
    except (FileNotFoundError, json.JSONDecodeError) as e:
        return StepResult(
            name="step7_dict_update", ok=False,
            duration_s=(datetime.now() - t0).total_seconds(),
            detail=f"无法读取候选 JSON: {e}",
        )
    new_tags = (candidates.get("approved_candidates") or [])[:MAX_NEW_TAGS_PER_RUN]

    import openpyxl
    shutil.copyfile(base_dict_xlsx, out_xlsx)
    wb = openpyxl.load_workbook(str(out_xlsx))
    ws = wb["01_通用标签主表"]
    base_n_rows = ws.max_row
    next_id = 1
    for c in new_tags:
        row: list[Any] = [None] * ws.max_column
        new_id = f"TAG_GEN_E{900 + next_id:03d}"
        row[0] = new_id
        row[3] = c.get("tag_cn") or "演练新标签"
        row[4] = c.get("tag_en") or "dryrun_new_tag"
        ws.append(row)
        next_id += 1
    wb.save(str(out_xlsx))
    new_n_rows = ws.max_row
    diff = new_n_rows - base_n_rows
    detail = f"base_rows={base_n_rows} new_rows={new_n_rows} diff={diff}"
    if diff > MAX_NEW_TAGS_PER_RUN:
        return StepResult(
            name="step7_dict_update", ok=False,
            duration_s=(datetime.now() - t0).total_seconds(),
            detail=f"diff {diff} > MAX_NEW_TAGS_PER_RUN={MAX_NEW_TAGS_PER_RUN}",
            artifacts=[str(out_xlsx)],
        )
    return StepResult(
        name="step7_dict_update", ok=True,
        duration_s=(datetime.now() - t0).total_seconds(),
        detail=detail,
        artifacts=[str(out_xlsx)],
        skipped=dry_run and diff == 0,
    )


def step8_bi_recompute(
    new_dict_xlsx: Path,
    *,
    dry_run: bool,
) -> StepResult:
    t0 = datetime.now()
    detail = (
        "[dry-run] 仅记录应触发：dual_coverage_calculator + maa_strategy_generator "
        "+ agrs_summarizer 重跑（不实际执行）" if dry_run else
        "[prod] 应在 v4.1 字典就绪后调度全量重打 + BI 重算"
    )
    return StepResult(
        name="step8_bi_recompute", ok=True,
        duration_s=(datetime.now() - t0).total_seconds(),
        detail=detail,
        artifacts=[str(new_dict_xlsx)],
        skipped=True,
    )


def run_pipeline(
    input_jsonl: Path,
    base_dict_xlsx: Path,
    out_dict_xlsx: Path,
    work_dir: Path,
    *,
    dry_run: bool,
    sample_rate: float,
) -> tuple[bool, list[StepResult]]:
    work_dir.mkdir(parents=True, exist_ok=True)

    s1 = step1_zero_label_extractor(
        input_jsonl, work_dir / "s1_zero_low_conf.jsonl", dry_run=dry_run,
    )
    if not s1.ok: return False, [s1]

    s2 = step2_closed_relabel(
        Path(s1.artifacts[0]), work_dir / "s2_relabeled.jsonl", dry_run=dry_run,
    )
    if not s2.ok: return False, [s1, s2]

    s3 = step3_open_set_sampling(
        Path(s2.artifacts[0]), work_dir / "s3_sampled.jsonl",
        sample_rate=sample_rate, dry_run=dry_run,
    )
    if not s3.ok: return False, [s1, s2, s3]

    s4 = step4_candidate_filter(
        Path(s3.artifacts[0]), base_dict_xlsx, work_dir / "s4_candidates", dry_run=dry_run,
    )
    if not s4.ok: return False, [s1, s2, s3, s4]

    s5 = step5_alchemist_lf(
        Path(s4.artifacts[0]), work_dir / "s5_lfs", dry_run=dry_run,
    )
    if not s5.ok: return False, [s1, s2, s3, s4, s5]

    s6 = step6_active_learning(
        Path(s4.artifacts[0]), work_dir / "s6_active_learning_queue.jsonl", dry_run=dry_run,
    )
    if not s6.ok: return False, [s1, s2, s3, s4, s5, s6]

    s7 = step7_dict_update(
        Path(s4.artifacts[0]), base_dict_xlsx, out_dict_xlsx, dry_run=dry_run,
    )
    if not s7.ok: return False, [s1, s2, s3, s4, s5, s6, s7]

    s8 = step8_bi_recompute(Path(s7.artifacts[0]), dry_run=dry_run)
    return s8.ok, [s1, s2, s3, s4, s5, s6, s7, s8]


def render_log(results: list[StepResult], dry_run: bool, success: bool) -> str:
    md = []
    md.append(f"# 月度进化运行日志")
    md.append("")
    md.append(f"- 时间：{_now()}")
    md.append(f"- 模式：{'dry-run' if dry_run else 'production'}")
    md.append(f"- 总判定：{'🟢 PASS' if success else '🔴 FAIL'}")
    md.append("")
    md.append("| # | 步骤 | 状态 | 耗时 | 详情 |")
    md.append("|---:|---|:---:|---:|---|")
    for i, r in enumerate(results, 1):
        mark = "⏭️ SKIP" if r.skipped and r.ok else ("✅" if r.ok else "❌")
        md.append(f"| {i} | {r.name} | {mark} | {r.duration_s:.2f}s | {r.detail} |")
    return "\n".join(md) + "\n"


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(
        description="Phase 5 D12 月度字典进化 cron（8 步 pipeline）",
    )
    ap.add_argument("--tenant", default="momcozy",
                    help="租户名（默认 momcozy；v6.0 多租户前瞻）")
    ap.add_argument("--input", type=Path,
                    default=RESEARCH_ROOT_DEFAULT / "04-输出结果/unified_labeling/phase5_intermediate_merged.jsonl",
                    help="输入 jsonl")
    ap.add_argument("--base-dict",
                    type=Path,
                    default=RESEARCH_ROOT_DEFAULT / "04-输出结果/01-字典版本/tag_dictionary_v4.0.xlsx",
                    help="基础字典 xlsx")
    ap.add_argument("--output-dict", type=Path,
                    default=Path("/tmp") / "tag_dictionary_v4.1_dryrun.xlsx",
                    help="新版字典输出路径")
    ap.add_argument("--work-dir", type=Path,
                    default=Path("/tmp/voc_monthly_evolution"),
                    help="中间产物工作目录")
    ap.add_argument("--log-out", type=Path, default=None,
                    help="进化日志 markdown 输出路径")
    ap.add_argument("--dry-run", action="store_true",
                    help="dry-run 模式（不调 LLM、不写正式字典）")
    ap.add_argument("--sample-rate", type=float, default=0.05,
                    help="Step 3 开集采样率（默认 5%%）")
    args = ap.parse_args(argv)

    print(f"⏳ Monthly Evolution Cron — tenant={args.tenant} dry_run={args.dry_run}", file=sys.stderr)
    print(f"   input={args.input}", file=sys.stderr)
    print(f"   base_dict={args.base_dict}", file=sys.stderr)

    if not args.input.is_file():
        print(f"❌ input not found: {args.input}", file=sys.stderr); return 2
    if not args.base_dict.is_file():
        print(f"❌ base dict not found: {args.base_dict}", file=sys.stderr); return 2

    success, results = run_pipeline(
        args.input, args.base_dict, args.output_dict, args.work_dir,
        dry_run=args.dry_run, sample_rate=args.sample_rate,
    )

    log_md = render_log(results, args.dry_run, success)
    print(log_md, file=sys.stderr)

    if args.log_out:
        args.log_out.parent.mkdir(parents=True, exist_ok=True)
        args.log_out.write_text(log_md, encoding="utf-8")
        print(f"📄 Log: {args.log_out}", file=sys.stderr)

    if FEISHU_WEBHOOK_FILE.is_file():
        flag = "✅ 完成" if success else "❌ 失败"
        n_ok = sum(1 for r in results if r.ok)
        msg = (f"[VOC 月度进化-{args.tenant}] {flag} ({n_ok}/{len(results)} steps) "
               f"mode={'dry-run' if args.dry_run else 'prod'} @ {_now()}")
        _post_feishu(msg)

    return 0 if success else 1


if __name__ == "__main__":
    sys.exit(main())
