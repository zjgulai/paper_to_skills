"""Phase 6 D1 F7 — v4.0 → v4.1 Dictionary Fixer

修复 v4.0 字典三类字段缺失（Sisyphus D14 诊断发现）：

  F7.1  aspect → tag_id 映射（10_Aspect库.关联tag_ids 100% 空 → 填）
        规则：对每个 aspect，用 aspect_en 在 tag 的 en/cn/kw 字段做 token 交集
              打分；保留 Top-5 最强匹配；空匹配记 <NO_MATCH>

  F7.2  aspect_cn 中文名补齐（55/56 是 "【待填写】" → LLM 单轮批量翻译）
        DeepSeek 一次性输入所有 aspect_en + category，JSON 模式返回翻译

  F7.3  业务动作/策略包/对应原子指标 LLM 补齐（品线表空值 25-75%）
        只处理通用标签主表（Sheet 1）的 58 条 "优化建议" 空值；
        DeepSeek 分批输入 tag_id + tag_cn + sentiment_polarity，
        JSON 模式返回 optimization_suggestion / priority

产出：
  - v4.1 xlsx（复制 v4.0 → 改 3 处 → 写 v4.1）
  - v4.1_dict_fix_report.md（diff 报告：per-sheet cells changed）

用法：
  python v41_dict_fixer.py \
    --base-dict <vault>/04-输出结果/01-字典版本/tag_dictionary_v4.0.xlsx \
    --output-dict <vault>/04-输出结果/01-字典版本/tag_dictionary_v4.1.xlsx \
    --report <vault>/04-输出结果/03-审计报告/phase6_d1_v41_dict_fix.md \
    [--skip-llm]   # 只跑 F7.1 规则匹配，不调 LLM

成本估算（完整跑）：
  - F7.2: 1 次 LLM 调用（55 aspect 一次性翻译）
  - F7.3: ~3 次 LLM 调用（按 20 tag/批 分批）
  - 总计 ~4 次 DeepSeek 调用，< 0.05 USD
"""

from __future__ import annotations

import argparse
import json
import re
import shutil
import sys
from collections import Counter
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "07-LLM引擎"))


TOP_K_ASPECT_TAGS = 5
LLM_BATCH_TAGS = 20
STOPWORDS = {
    "the", "a", "an", "and", "or", "of", "to", "for", "in", "on", "at",
    "is", "are", "was", "were", "be", "with", "by",
}


@dataclass
class FixReport:
    aspect_mapped: int = 0
    aspect_no_match: int = 0
    aspect_cn_filled: int = 0
    tag_optimization_filled: int = 0
    llm_calls: int = 0
    llm_failed: int = 0
    notes: list[str] = field(default_factory=list)


def _tokenize(s: str | None) -> set[str]:
    if not s:
        return set()
    return {w for w in re.findall(r"[a-z]{3,}", str(s).lower()) if w not in STOPWORDS}


def _score_tag_for_aspect(aspect_en: str, aspect_cat: str, tag: dict[str, Any]) -> float:
    """规则打分：token 交集 × 字段权重。en > kw > cn"""
    a_tokens = _tokenize(aspect_en)
    if not a_tokens:
        return 0.0
    en_tokens = _tokenize(tag.get("en"))
    kw_tokens = _tokenize(tag.get("kw"))
    cn_tokens = _tokenize(tag.get("cn"))

    en_overlap = len(a_tokens & en_tokens)
    kw_overlap = len(a_tokens & kw_tokens)
    cn_overlap = len(a_tokens & cn_tokens)

    if en_overlap == 0 and kw_overlap == 0 and cn_overlap == 0:
        return 0.0

    a_en_lower = aspect_en.lower()
    substring_boost = 0.0
    tag_en_lower = str(tag.get("en") or "").lower()
    if a_en_lower and a_en_lower in tag_en_lower:
        substring_boost = 1.0
    elif tag_en_lower and tag_en_lower in a_en_lower:
        substring_boost = 0.5

    score = 1.0 * en_overlap + 0.7 * kw_overlap + 0.4 * cn_overlap + substring_boost

    norm = score / len(a_tokens)
    return round(norm, 3)


def load_tags_from_sheet1(ws: Worksheet) -> dict[str, dict[str, Any]]:
    hdr = [c.value for c in next(ws.iter_rows(max_row=1))]
    ix = {k: hdr.index(k) for k in [
        "标签ID", "VOC标签（中文）", "VOC标签（英文）",
        "英文关键词/典型表达", "消费者习惯关键词/原话短语", "标签定义",
        "情感极性", "业务动作/责任部门",
    ]}
    tags: dict[str, dict[str, Any]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        tid = row[ix["标签ID"]]
        if not tid:
            continue
        tags[str(tid)] = {
            "cn": row[ix["VOC标签（中文）"]],
            "en": row[ix["VOC标签（英文）"]],
            "kw": row[ix["英文关键词/典型表达"]],
            "habit_kw": row[ix["消费者习惯关键词/原话短语"]],
            "defn": row[ix["标签定义"]],
            "polarity": row[ix["情感极性"]],
            "biz_action": row[ix["业务动作/责任部门"]],
        }
    return tags


def f7_1_map_aspect_to_tags(
    ws_aspect: Worksheet,
    tags: dict[str, dict[str, Any]],
    report: FixReport,
) -> None:
    hdr = [c.value for c in next(ws_aspect.iter_rows(max_row=1))]
    ix_aen = hdr.index("aspect_en")
    ix_cat = hdr.index("category")
    ix_link = hdr.index("关联tag_ids")

    for row_idx, row in enumerate(ws_aspect.iter_rows(min_row=2, max_row=ws_aspect.max_row), start=2):
        aen = row[ix_aen].value
        cat = row[ix_cat].value
        if not aen:
            continue
        current = row[ix_link].value
        if current and str(current).strip() and str(current).strip() != "<NO_MATCH>":
            continue
        scored: list[tuple[str, float]] = []
        for tid, t in tags.items():
            s = _score_tag_for_aspect(str(aen), str(cat or ""), t)
            if s > 0:
                scored.append((tid, s))
        scored.sort(key=lambda x: -x[1])
        top = scored[:TOP_K_ASPECT_TAGS]
        if not top:
            ws_aspect.cell(row=row_idx, column=ix_link + 1, value="<NO_MATCH>")
            report.aspect_no_match += 1
        else:
            ws_aspect.cell(
                row=row_idx, column=ix_link + 1,
                value="|".join(f"{tid}:{score}" for tid, score in top),
            )
            report.aspect_mapped += 1


def f7_2_fill_aspect_cn(
    ws_aspect: Worksheet,
    report: FixReport,
    skip_llm: bool,
) -> None:
    hdr = [c.value for c in next(ws_aspect.iter_rows(max_row=1))]
    ix_aid = hdr.index("aspect_id")
    ix_aen = hdr.index("aspect_en")
    ix_acn = hdr.index("aspect_cn")

    todo: list[tuple[int, str, str]] = []
    for row_idx, row in enumerate(ws_aspect.iter_rows(min_row=2, max_row=ws_aspect.max_row), start=2):
        aid = row[ix_aid].value
        aen = row[ix_aen].value
        acn = row[ix_acn].value
        if aid and aen and (not acn or "【待填写】" in str(acn)):
            todo.append((row_idx, str(aid), str(aen)))

    if not todo:
        report.notes.append("F7.2: no aspect_cn needs filling")
        return

    if skip_llm:
        for row_idx, aid, aen in todo:
            ws_aspect.cell(row=row_idx, column=ix_acn + 1, value=f"<PENDING_LLM:{aen}>")
        report.notes.append(f"F7.2: skip-llm mode, {len(todo)} aspects marked <PENDING_LLM>")
        return

    try:
        from llm_client import LLMClient
    except ImportError as e:
        report.notes.append(f"F7.2: llm_client import failed: {e}")
        return

    client = LLMClient()
    prompt_items = "\n".join(f"{aid}\t{aen}" for _, aid, aen in todo)
    system = (
        "You are a bilingual e-commerce domain expert (Chinese + English). "
        "Translate each English aspect (noun phrase) into concise Chinese (2-6 chars). "
        "Context: mother & baby cross-border e-commerce. "
        "Return JSON: {\"aspect_id_1\": \"中文\", \"aspect_id_2\": \"中文\", ...}"
    )
    user = f"Translate each aspect_en to aspect_cn:\n\n{prompt_items}"
    try:
        resp = client.chat_sync(
            vendor="deepseek",
            messages=[{"role": "system", "content": system},
                      {"role": "user", "content": user}],
            response_format={"type": "json_object"},
            max_tokens=2000,
            temperature=0.2,
        )
        report.llm_calls += 1
        data = json.loads(resp.content)
        for row_idx, aid, aen in todo:
            cn = data.get(aid) or data.get(aen)
            if cn and isinstance(cn, str):
                ws_aspect.cell(row=row_idx, column=ix_acn + 1, value=cn.strip())
                report.aspect_cn_filled += 1
    except Exception as e:
        report.llm_failed += 1
        report.notes.append(f"F7.2 LLM call failed: {e!r}")


def f7_3_fill_optimization(
    ws_main: Worksheet,
    report: FixReport,
    skip_llm: bool,
) -> None:
    hdr = [c.value for c in next(ws_main.iter_rows(max_row=1))]
    ix_id = hdr.index("标签ID")
    ix_cn = hdr.index("VOC标签（中文）")
    ix_en = hdr.index("VOC标签（英文）")
    ix_pol = hdr.index("情感极性")
    ix_opt = hdr.index("优化建议")
    ix_pri = hdr.index("优化优先级")

    todo: list[tuple[int, str, str, str, str]] = []
    for row_idx, row in enumerate(ws_main.iter_rows(min_row=2, max_row=ws_main.max_row), start=2):
        tid = row[ix_id].value
        opt = row[ix_opt].value
        if not tid:
            continue
        if opt and str(opt).strip():
            continue
        cn = str(row[ix_cn].value or "")
        en = str(row[ix_en].value or "")
        pol = str(row[ix_pol].value or "中性")
        todo.append((row_idx, str(tid), cn, en, pol))

    if not todo:
        report.notes.append("F7.3: all rows already have 优化建议")
        return
    if skip_llm:
        for row_idx, tid, cn, en, pol in todo:
            ws_main.cell(row=row_idx, column=ix_opt + 1, value=f"<PENDING_LLM:{pol}>")
            ws_main.cell(row=row_idx, column=ix_pri + 1, value="P2")
        report.notes.append(f"F7.3: skip-llm mode, {len(todo)} tags marked <PENDING_LLM> + priority P2 default")
        return

    try:
        from llm_client import LLMClient
    except ImportError as e:
        report.notes.append(f"F7.3: llm_client import failed: {e}")
        return
    client = LLMClient()

    for batch_start in range(0, len(todo), LLM_BATCH_TAGS):
        batch = todo[batch_start:batch_start + LLM_BATCH_TAGS]
        items = "\n".join(
            f"{tid}\t{cn}\t{en}\t{pol}" for _, tid, cn, en, pol in batch
        )
        system = (
            "You are a product strategy expert for mother & baby cross-border e-commerce. "
            "For each VOC tag, provide: "
            "(1) optimization_suggestion: concise Chinese action (≤ 30 chars), "
            "(2) priority: P0/P1/P2 (P0=urgent for 负向 high-reach, P2=routine for 正向/中性). "
            "Return JSON: {\"tag_id\": {\"opt\": \"建议\", \"pri\": \"P0/P1/P2\"}, ...}"
        )
        user = f"Tag rows (tag_id\\ttag_cn\\ttag_en\\tpolarity):\n\n{items}"
        try:
            resp = client.chat_sync(
                vendor="deepseek",
                messages=[{"role": "system", "content": system},
                          {"role": "user", "content": user}],
                response_format={"type": "json_object"},
                max_tokens=2500,
                temperature=0.2,
            )
            report.llm_calls += 1
            data = json.loads(resp.content)
            for row_idx, tid, cn, en, pol in batch:
                rec = data.get(tid)
                if isinstance(rec, dict):
                    opt = rec.get("opt")
                    pri = rec.get("pri")
                    if opt:
                        ws_main.cell(row=row_idx, column=ix_opt + 1, value=str(opt).strip())
                    if pri in ("P0", "P1", "P2"):
                        ws_main.cell(row=row_idx, column=ix_pri + 1, value=pri)
                    if opt:
                        report.tag_optimization_filled += 1
        except Exception as e:
            report.llm_failed += 1
            report.notes.append(
                f"F7.3 batch {batch_start}-{batch_start+len(batch)} LLM call failed: {e!r}"
            )


def render_report(report: FixReport, base: Path, out: Path) -> str:
    lines = []
    p = lines.append
    p("---")
    p("name: phase6-d1-v41-dict-fix")
    p("description: Phase 6 D1 F7 字典字段修复报告 — aspect→tag 映射 + aspect_cn 中文 + 优化建议补齐。"
      "当验证 v4.0 → v4.1 字段质量修复效果、查看 LLM 调用成本时使用。")
    p(f"date: {datetime.now().strftime('%Y-%m-%d')}")
    p("phase: phase6")
    p("day: D1")
    p("doc_type: audit-report")
    p("module: voc-nlp")
    p("---")
    p("")
    p("# Phase 6 D1 F7 字典字段修复报告（v4.0 → v4.1）")
    p("")
    p(f"- **基础字典**：`{base}`")
    p(f"- **产出字典**：`{out}`")
    p(f"- **运行时间**：{datetime.now().isoformat(timespec='seconds')}")
    p("")
    p("## 一、修复统计")
    p("")
    p("| 子任务 | 实施 | 数量 |")
    p("|---|---|---:|")
    p(f"| F7.1 aspect→tag_id 映射（规则打分）| 填入 | {report.aspect_mapped} |")
    p(f"| F7.1 aspect→tag_id 映射 | 无匹配（记 <NO_MATCH>）| {report.aspect_no_match} |")
    p(f"| F7.2 aspect_cn 中文补齐（LLM）| 成功填入 | {report.aspect_cn_filled} |")
    p(f"| F7.3 优化建议补齐（LLM）| 成功填入 | {report.tag_optimization_filled} |")
    p(f"| LLM 调用次数 | — | {report.llm_calls} |")
    p(f"| LLM 失败次数 | — | {report.llm_failed} |")
    p("")
    if report.notes:
        p("## 二、备注")
        p("")
        for n in report.notes:
            p(f"- {n}")
        p("")
    return "\n".join(lines)


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(description="Phase 6 D1 F7 v4.0 → v4.1 字典修复")
    ap.add_argument("--base-dict", required=True, type=Path)
    ap.add_argument("--output-dict", required=True, type=Path)
    ap.add_argument("--report", type=Path, default=None)
    ap.add_argument("--skip-llm", action="store_true",
                    help="仅跑 F7.1 规则匹配，F7.2/F7.3 标记 <PENDING_LLM>")
    args = ap.parse_args(argv)

    if not args.base_dict.is_file():
        print(f"❌ base dict not found: {args.base_dict}", file=sys.stderr); return 2

    print(f"⏳ 复制 base → out", file=sys.stderr)
    args.output_dict.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(args.base_dict, args.output_dict)

    print(f"⏳ 打开 xlsx", file=sys.stderr)
    wb = openpyxl.load_workbook(str(args.output_dict))
    ws_main = wb["01_通用标签主表"]
    ws_aspect = wb["10_Aspect库"]

    report = FixReport()

    print(f"⏳ F7.1 aspect → tag_id 映射", file=sys.stderr)
    tags = load_tags_from_sheet1(ws_main)
    print(f"   loaded {len(tags)} tags", file=sys.stderr)
    f7_1_map_aspect_to_tags(ws_aspect, tags, report)
    print(f"   mapped={report.aspect_mapped} no_match={report.aspect_no_match}", file=sys.stderr)

    print(f"⏳ F7.2 aspect_cn 中文补齐 (skip_llm={args.skip_llm})", file=sys.stderr)
    f7_2_fill_aspect_cn(ws_aspect, report, args.skip_llm)
    print(f"   filled={report.aspect_cn_filled}", file=sys.stderr)

    print(f"⏳ F7.3 优化建议补齐 (skip_llm={args.skip_llm})", file=sys.stderr)
    f7_3_fill_optimization(ws_main, report, args.skip_llm)
    print(f"   filled={report.tag_optimization_filled}", file=sys.stderr)

    print(f"⏳ 写回 xlsx", file=sys.stderr)
    wb.save(str(args.output_dict))
    print(f"📄 {args.output_dict}", file=sys.stderr)

    if args.report:
        args.report.parent.mkdir(parents=True, exist_ok=True)
        md = render_report(report, args.base_dict, args.output_dict)
        args.report.write_text(md, encoding="utf-8")
        print(f"📄 Report: {args.report}", file=sys.stderr)

    if report.llm_failed > 0:
        print(f"⚠️  LLM failures: {report.llm_failed}", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
