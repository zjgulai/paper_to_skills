"""Phase 7 D1 — ETL: phase6_d9_filtered.jsonl + tag_dictionary_v4.1.xlsx → voc_bi (postgres)

Loads the final D9 filtered dataset (Gate 7/7 + precision 0.896) plus the v4.1
tag dictionary into a 4-table star schema in postgres, ready for Superset/Metabase.

Tables (see sql/voc_bi_schema.sql for DDL):
  - dim_tag           (267 rows from v4.1 dict 01_通用标签主表)
  - voc_review        (364,569 rows, 1 per review)
  - voc_label         (~689K rows, 1 per (review, label))
  - voc_brand_mention (~varies, 1 per (review, brand))

Design:
  - Streaming jsonl read (O(1) memory)
  - Batch INSERT with execute_values (1000 rows/batch)
  - Idempotent: TRUNCATE before insert (re-runnable)
  - Connection from ~/.paper2skills/voc_bi_pg.json (chmod 600)
  - Sentiment string→float normalize (handles legacy 'positive'/'negative')

Usage:
  python etl_to_postgres.py \
    --input <vault>/04-输出结果/unified_labeling/phase6_d9_filtered.jsonl \
    --dict <vault>/04-输出结果/01-字典版本/tag_dictionary_v4.1.xlsx
"""

from __future__ import annotations

import argparse
import json
import sys
import time
from pathlib import Path
from typing import Any, Iterator

import openpyxl
import psycopg2
from psycopg2.extras import execute_values


DB_CONFIG_PATH = Path("~/.paper2skills/voc_bi_pg.json").expanduser()
BATCH_SIZE = 1000

_SENTIMENT_STR_TO_FLOAT = {
    "positive": 1.0, "negative": -1.0, "neutral": 0.0,
    "pos": 1.0, "neg": -1.0, "neu": 0.0,
    "正向": 1.0, "负向": -1.0, "中性": 0.0,
}


def _to_sentiment_float(raw: Any) -> float | None:
    """Normalize sentiment_calibrated which may be float or string in legacy data."""
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return float(raw)
    if isinstance(raw, str):
        key = raw.strip().lower()
        if key in _SENTIMENT_STR_TO_FLOAT:
            return _SENTIMENT_STR_TO_FLOAT[key]
        try:
            return float(raw)
        except ValueError:
            return None
    return None


def load_pg_config(path: Path = DB_CONFIG_PATH) -> dict[str, Any]:
    if not path.is_file():
        raise FileNotFoundError(f"voc_bi pg config not found: {path}")
    return json.loads(path.read_text(encoding="utf-8"))


def connect(cfg: dict[str, Any]) -> Any:
    return psycopg2.connect(
        host=cfg["host"], port=cfg["port"],
        database=cfg["database"], user=cfg["user"], password=cfg["password"],
    )


def stream_jsonl(path: Path) -> Iterator[dict[str, Any]]:
    with path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                yield json.loads(line)
            except json.JSONDecodeError:
                continue


def load_dim_tag_rows(dict_path: Path) -> list[tuple]:
    """Read v4.1 dict 01_通用标签主表 → list of dim_tag rows."""
    wb = openpyxl.load_workbook(str(dict_path), read_only=True, data_only=True)
    ws = wb["01_通用标签主表"]
    hdr = [c.value for c in next(ws.iter_rows(max_row=1))]
    ix = {k: hdr.index(k) for k in [
        "标签ID", "VOC标签（中文）", "VOC标签（英文）", "AIPL节点",
        "情感极性", "主责部门", "业务动作/责任部门", "策略包",
        "是否通用标签", "审核状态",
    ]}
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        tid = row[ix["标签ID"]]
        if not tid:
            continue
        is_general_raw = row[ix["是否通用标签"]]
        is_general = (
            None if is_general_raw is None or str(is_general_raw).strip() == ""
            else (str(is_general_raw).strip() == "是")
        )
        rows.append((
            str(tid),
            str(row[ix["VOC标签（中文）"]] or "") or None,
            str(row[ix["VOC标签（英文）"]] or "") or None,
            str(row[ix["AIPL节点"]] or "") or None,
            str(row[ix["情感极性"]] or "") or None,
            str(row[ix["主责部门"]] or "") or None,
            str(row[ix["业务动作/责任部门"]] or "") or None,
            str(row[ix["策略包"]] or "") or None,
            is_general,
            str(row[ix["审核状态"]] or "") or None,
        ))
    return rows


def insert_dim_tag(conn: Any, rows: list[tuple]) -> int:
    with conn.cursor() as cur:
        cur.execute("TRUNCATE dim_tag CASCADE;")
        execute_values(
            cur,
            """INSERT INTO dim_tag (tag_id, tag_cn, tag_en, aipl_node, polarity,
                                    dept_owner, biz_action, strategy_pkg, is_general, audit_status)
               VALUES %s""",
            rows, page_size=BATCH_SIZE,
        )
    conn.commit()
    return len(rows)


def review_to_row(r: dict[str, Any]) -> tuple:
    brand_mentions = r.get("brand_mentions") or []
    label_sources = r.get("label_sources") or []
    return (
        r.get("review_id"),
        r.get("text"),
        r.get("data_source"),
        r.get("platform"),
        r.get("source_type"),
        r.get("asin"),
        r.get("spu_code"),
        r.get("product_line"),
        r.get("category"),
        float(r["rating"]) if isinstance(r.get("rating"), (int, float)) else None,
        r.get("language"),
        float(r["sentiment_polarity"]) if isinstance(r.get("sentiment_polarity"), (int, float)) else None,
        r.get("sentiment_calibration"),
        r.get("proxy_nps"),
        r.get("aipl_stage"),
        r.get("persona_derived"),
        len(brand_mentions),
        bool(r.get("brand_comparison")),
        float(r["_quality_score"]) if isinstance(r.get("_quality_score"), (int, float)) else None,
        int(r["n_tags"]) if isinstance(r.get("n_tags"), int) else None,
        r.get("label_source"),
        list(label_sources) if isinstance(label_sources, list) else None,
        bool(r.get("_phase6_d4_meta")),
    )


def label_to_rows(rid: str, labels: list[Any]) -> Iterator[tuple]:
    for lbl in labels:
        if not isinstance(lbl, dict) or not lbl.get("tag_id"):
            continue
        yield (
            rid,
            lbl.get("tag_id"),
            lbl.get("tag_cn"),
            lbl.get("tag_en"),
            lbl.get("aipl_node"),
            lbl.get("sentiment_preset"),
            _to_sentiment_float(lbl.get("sentiment_calibrated")),
            float(lbl["confidence"]) if isinstance(lbl.get("confidence"), (int, float)) else None,
            float(lbl["_confidence_original"]) if isinstance(lbl.get("_confidence_original"), (int, float)) else None,
            float(lbl["_confidence_lift"]) if isinstance(lbl.get("_confidence_lift"), (int, float)) else None,
            lbl.get("_source"),
        )


def brand_to_rows(rid: str, brands: list[Any]) -> Iterator[tuple]:
    for b in brands:
        if isinstance(b, str) and b.strip():
            yield (rid, b.strip())


def etl_reviews(
    conn: Any,
    input_path: Path,
) -> dict[str, int]:
    review_buffer: list[tuple] = []
    label_buffer: list[tuple] = []
    brand_buffer: list[tuple] = []
    n_reviews = 0
    n_labels = 0
    n_brands = 0
    t0 = time.time()

    with conn.cursor() as cur:
        cur.execute("TRUNCATE voc_review CASCADE;")
    conn.commit()

    review_sql = """INSERT INTO voc_review (
        review_id, text, data_source, platform, source_type, asin, spu_code,
        product_line, category, rating, language, sentiment_polarity,
        sentiment_calibration, proxy_nps, aipl_stage, persona_derived,
        brand_count, brand_comparison, quality_score, n_tags,
        label_source, label_sources, has_phase6_d4_meta
    ) VALUES %s"""

    label_sql = """INSERT INTO voc_label (
        review_id, tag_id, tag_cn, tag_en, aipl_node, sentiment_preset,
        sentiment_calibrated, confidence, confidence_original, confidence_lift, source
    ) VALUES %s"""

    brand_sql = """INSERT INTO voc_brand_mention (review_id, brand_name) VALUES %s"""

    def flush() -> None:
        nonlocal n_reviews, n_labels, n_brands
        with conn.cursor() as cur:
            if review_buffer:
                execute_values(cur, review_sql, review_buffer, page_size=BATCH_SIZE)
                n_reviews += len(review_buffer)
            if label_buffer:
                execute_values(cur, label_sql, label_buffer, page_size=BATCH_SIZE)
                n_labels += len(label_buffer)
            if brand_buffer:
                execute_values(cur, brand_sql, brand_buffer, page_size=BATCH_SIZE)
                n_brands += len(brand_buffer)
        conn.commit()
        review_buffer.clear()
        label_buffer.clear()
        brand_buffer.clear()

    for r in stream_jsonl(input_path):
        rid = r.get("review_id")
        if not rid:
            continue
        review_buffer.append(review_to_row(r))
        for label_row in label_to_rows(rid, r.get("labels") or []):
            label_buffer.append(label_row)
        for brand_row in brand_to_rows(rid, r.get("brand_mentions") or []):
            brand_buffer.append(brand_row)
        if len(review_buffer) >= BATCH_SIZE:
            flush()
            if n_reviews % 50000 == 0:
                elapsed = time.time() - t0
                rate = n_reviews / max(elapsed, 1)
                print(f"  [{n_reviews:,} reviews / {n_labels:,} labels / {n_brands:,} brands] "
                      f"{rate:.0f} rec/s", file=sys.stderr)
    flush()
    return {"reviews": n_reviews, "labels": n_labels, "brands": n_brands,
            "elapsed_s": int(time.time() - t0)}


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(description="Phase 7 D1 ETL: jsonl + xlsx → postgres")
    ap.add_argument("--input", required=True, type=Path)
    ap.add_argument("--dict", required=True, type=Path)
    args = ap.parse_args(argv)

    if not args.input.is_file():
        print(f"❌ input not found: {args.input}", file=sys.stderr); return 2
    if not args.dict.is_file():
        print(f"❌ dict not found: {args.dict}", file=sys.stderr); return 2

    cfg = load_pg_config()
    print(f"⏳ Connecting postgres @ {cfg['host']}:{cfg['port']}/{cfg['database']}", file=sys.stderr)
    conn = connect(cfg)

    try:
        print(f"⏳ Loading dim_tag from {args.dict.name}", file=sys.stderr)
        dim_rows = load_dim_tag_rows(args.dict)
        n_dim = insert_dim_tag(conn, dim_rows)
        print(f"   {n_dim} dim_tag rows inserted", file=sys.stderr)

        print(f"⏳ Streaming reviews from {args.input.name} ({args.input.stat().st_size//1024//1024} MB)",
              file=sys.stderr)
        stats = etl_reviews(conn, args.input)
        print(f"\n✅ DONE", file=sys.stderr)
        print(f"   reviews:  {stats['reviews']:,}", file=sys.stderr)
        print(f"   labels:   {stats['labels']:,}", file=sys.stderr)
        print(f"   brands:   {stats['brands']:,}", file=sys.stderr)
        print(f"   elapsed:  {stats['elapsed_s']}s "
              f"({stats['reviews']/max(stats['elapsed_s'],1):.0f} rec/s)", file=sys.stderr)
        return 0
    finally:
        conn.close()


if __name__ == "__main__":
    sys.exit(main())
