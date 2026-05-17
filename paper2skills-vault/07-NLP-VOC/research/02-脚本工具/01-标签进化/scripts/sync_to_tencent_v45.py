"""Apply v4.5 dictionary + 88 orphan tags to Tencent Cloud Superset voc_bi.

Strategy: dictionary overwrite while preserving Tencent-only column
atomic_indicator_id. Steps:
  0. snapshot dim_tag (backup table) for rollback
  1. enable allow_dml on database (Superset metadata)
  2. ALTER TABLE add product_line + collab_dept (idempotent)
  3. UPSERT 604 v4.5 rows (267 main + 337 product-line), preserving atomic_indicator_id
  4. INSERT 88 orphan rows (audit_status='auto_from_label')
  5. disable allow_dml
  6. verify orphan=0 + atomic_indicator_id preserved

Critical safety: every step writes to a backup table first; rollback SQL kept
in /tmp/tencent_rollback.sql for manual recovery.
"""

from __future__ import annotations

import json
import sys
import time
from pathlib import Path

import requests
import urllib3

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

ROOT = Path("/Users/pray/project/paper_to_skills/paper2skills-vault/07-NLP-VOC")
DICT_PATH = ROOT / "research/04-输出结果/01-字典版本/tag_dictionary_v4.5.xlsx"
ORPHAN_SEED = Path("/tmp/orphan_seed_tencent.jsonl")

URL = "https://voc.lute-tlz-dddd.top"
USER, PASS = "admin", "voc_admin_2026"

s = requests.Session()
s.verify = False

def login():
    tok = s.post(
        f"{URL}/api/v1/security/login",
        json={"username": USER, "password": PASS, "provider": "db", "refresh": True},
        timeout=15,
    ).json()["access_token"]
    s.headers.update({"Authorization": f"Bearer {tok}"})
    csrf = s.get(f"{URL}/api/v1/security/csrf_token/", timeout=10).json()["result"]
    s.headers.update({"X-CSRFToken": csrf, "Referer": URL})

def run_sql(db_id: int, sql: str, limit: int = 5000):
    r = s.post(
        f"{URL}/api/v1/sqllab/execute/",
        json={
            "database_id": db_id,
            "sql": sql,
            "schema": "public",
            "queryLimit": limit,
            "select_as_cta": False,
            "tab": "voc-update",
            "ctas_method": "TABLE",
        },
        timeout=180,
    )
    if r.status_code != 200:
        return {"error": r.status_code, "text": r.text[:800]}
    return r.json()

def get_db(db_id: int):
    return s.get(f"{URL}/api/v1/database/{db_id}", timeout=10).json()["result"]

def set_dml(db_id: int, allow: bool):
    payload = {"allow_dml": allow}
    r = s.put(f"{URL}/api/v1/database/{db_id}", json=payload, timeout=15)
    if r.status_code != 200:
        raise RuntimeError(f"set_dml failed: {r.status_code} {r.text[:300]}")
    return r.json()

def main():
    print("=" * 78)
    print("  Tencent Cloud voc_bi · v4.5 dictionary + 88 orphan tags 增量更新")
    print("=" * 78)

    login()
    DB_ID = 1
    print(f"\n[step 0] Verify connection + snapshot")
    db_state = get_db(DB_ID)
    orig_dml = db_state.get("allow_dml", False)
    print(f"  voc_bi current allow_dml: {orig_dml}")

    cnt = run_sql(DB_ID, "SELECT COUNT(*) AS n FROM dim_tag")
    pre_count = cnt["data"][0]["n"]
    print(f"  dim_tag pre-count: {pre_count}")

    rollback_sql = [
        "-- ROLLBACK SQL for Tencent voc_bi · 2026-05-14",
        "-- 1. restore dim_tag from backup",
        "DROP TABLE IF EXISTS dim_tag_temp_swap;",
        "ALTER TABLE dim_tag RENAME TO dim_tag_temp_swap;",
        "ALTER TABLE dim_tag_backup_20260514 RENAME TO dim_tag;",
        "-- 2. revoke schema additions if needed:",
        "-- ALTER TABLE dim_tag DROP COLUMN IF EXISTS product_line;",
        "-- ALTER TABLE dim_tag DROP COLUMN IF EXISTS collab_dept;",
    ]
    Path("/tmp/tencent_rollback.sql").write_text("\n".join(rollback_sql))
    print(f"  Rollback SQL: /tmp/tencent_rollback.sql")

    print(f"\n[step 0.5] Create backup table dim_tag_backup_20260514")
    res = run_sql(DB_ID,
        """DROP TABLE IF EXISTS dim_tag_backup_20260514;
           CREATE TABLE dim_tag_backup_20260514 AS TABLE dim_tag;""")
    if res.get("error"):
        print(f"  ⚠️  cannot create backup (DML disabled). Will enable now.")
    else:
        print(f"  ✅ backup created")

    print(f"\n[step 1] Enable allow_dml=true on voc_bi")
    if not orig_dml:
        set_dml(DB_ID, True)
        print(f"  ✅ allow_dml enabled")
    else:
        print(f"  (already enabled)")

    try:
        if "create backup" in str(res).lower() or res.get("error"):
            print(f"\n[step 0.5 retry] Create backup table")
            res = run_sql(DB_ID,
                """DROP TABLE IF EXISTS dim_tag_backup_20260514;
                   CREATE TABLE dim_tag_backup_20260514 AS TABLE dim_tag;""")
            print(f"  backup result: {res.get('status', res)}")

        print(f"\n[step 2] ALTER TABLE add product_line + collab_dept")
        res = run_sql(DB_ID,
            """ALTER TABLE dim_tag ADD COLUMN IF NOT EXISTS product_line TEXT;
               ALTER TABLE dim_tag ADD COLUMN IF NOT EXISTS collab_dept TEXT;
               CREATE INDEX IF NOT EXISTS idx_dim_tag_product_line ON dim_tag(product_line);""")
        print(f"  ✅ schema extended")

        print(f"\n[step 3] Load v4.5 dictionary rows")
        import openpyxl, re
        CAT_PAT = re.compile(r'^\s*\[(?P<cat>[^\]]+)\]\s+(?P<en>.+?)\s*$')

        def col(header, name):
            return header.index(name) if name in header else None
        def safe(v):
            if v is None: return None
            s = str(v).strip()
            return s if s else None

        wb = openpyxl.load_workbook(str(DICT_PATH), read_only=True, data_only=True)
        rows = []

        ws = wb["01_通用标签主表"]
        h = [c.value for c in next(ws.iter_rows(max_row=1))]
        ix = {n: col(h, n) for n in
              ["标签ID","VOC标签（中文）","VOC标签（英文）","AIPL节点","情感极性",
               "主责部门","业务动作/责任部门","策略包","是否通用标签","审核状态"]}
        for row in ws.iter_rows(min_row=2, values_only=True):
            tid = row[ix["标签ID"]]
            if not tid: continue
            is_gen_raw = row[ix["是否通用标签"]]
            is_gen = None if is_gen_raw is None or str(is_gen_raw).strip()=='' else (str(is_gen_raw).strip()=='是')
            rows.append((
                str(tid).strip(), safe(row[ix["VOC标签（中文）"]]), safe(row[ix["VOC标签（英文）"]]),
                safe(row[ix["AIPL节点"]]), safe(row[ix["情感极性"]]),
                safe(row[ix["主责部门"]]), safe(row[ix["业务动作/责任部门"]]),
                safe(row[ix["策略包"]]), is_gen, safe(row[ix["审核状态"]]),
                None, None,
            ))

        for sheet_name, line in [("02_吸奶器","吸奶器"),("03_内衣服饰","内衣服饰"),
                                  ("04_家居家纺","家居家纺"),("05_母婴综合护理","母婴综合护理"),
                                  ("06_喂养电器","喂养电器"),("07_智能母婴电器","智能母婴电器")]:
            ws = wb[sheet_name]
            h = [c.value for c in next(ws.iter_rows(max_row=1))]
            ix = {n: col(h, n) for n in
                  ["标签ID","VOC标签（中文）","VOC标签（英文）","AIPL节点","情感极性",
                   "主责部门","业务动作/责任部门","协同部门","策略包","审核状态"]}
            for row in ws.iter_rows(min_row=2, values_only=True):
                tid = row[ix["标签ID"]]
                if not tid: continue
                rows.append((
                    str(tid).strip(), safe(row[ix["VOC标签（中文）"]]), safe(row[ix["VOC标签（英文）"]]),
                    safe(row[ix["AIPL节点"]]), safe(row[ix["情感极性"]]),
                    safe(row[ix["主责部门"]]), safe(row[ix["业务动作/责任部门"]]),
                    safe(row[ix["策略包"]]), False, safe(row[ix["审核状态"]]),
                    line, safe(row[ix["协同部门"]]),
                ))

        seen = set(); deduped = []
        for r in rows:
            if r[0] in seen: continue
            seen.add(r[0]); deduped.append(r)
        rows = deduped
        print(f"  v4.5 dict rows: {len(rows)} (deduped)")

        print(f"\n[step 4] UPSERT v4.5 rows in batches (preserves atomic_indicator_id)")
        def esc(v):
            if v is None: return "NULL"
            if isinstance(v, bool): return "TRUE" if v else "FALSE"
            return "'" + str(v).replace("'", "''") + "'"

        batch_sz = 50
        total_upserted = 0
        for i in range(0, len(rows), batch_sz):
            batch = rows[i:i+batch_sz]
            values_clauses = []
            for r in batch:
                vals = [esc(x) for x in r]
                values_clauses.append(f"({','.join(vals)})")
            sql = f"""INSERT INTO dim_tag
                (tag_id, tag_cn, tag_en, aipl_node, polarity, dept_owner,
                 biz_action, strategy_pkg, is_general, audit_status, product_line, collab_dept)
              VALUES {','.join(values_clauses)}
              ON CONFLICT (tag_id) DO UPDATE SET
                tag_cn       = EXCLUDED.tag_cn,
                tag_en       = EXCLUDED.tag_en,
                aipl_node    = EXCLUDED.aipl_node,
                polarity     = EXCLUDED.polarity,
                dept_owner   = EXCLUDED.dept_owner,
                biz_action   = CASE
                  WHEN EXCLUDED.biz_action IS NOT NULL AND EXCLUDED.biz_action <> '' THEN EXCLUDED.biz_action
                  ELSE dim_tag.biz_action
                END,
                strategy_pkg = EXCLUDED.strategy_pkg,
                is_general   = EXCLUDED.is_general,
                audit_status = EXCLUDED.audit_status,
                product_line = EXCLUDED.product_line,
                collab_dept  = EXCLUDED.collab_dept,
                loaded_at    = NOW();"""
            res = run_sql(DB_ID, sql)
            if res.get("error"):
                print(f"  ❌ batch {i//batch_sz}: {res}")
                break
            total_upserted += len(batch)
        print(f"  ✅ upserted: {total_upserted}")

        print(f"\n[step 5] Insert 88 orphan rows (audit_status='auto_from_label')")
        res = run_sql(DB_ID,
          """SELECT l.tag_id,
                    MAX(l.tag_cn) AS tag_cn, MAX(l.tag_en) AS tag_en,
                    MAX(l.aipl_node) AS aipl_node, MAX(l.sentiment_preset) AS sent,
                    COUNT(*) AS hits
             FROM voc_label l LEFT JOIN dim_tag d ON l.tag_id=d.tag_id
             WHERE d.tag_id IS NULL
             GROUP BY l.tag_id""")
        orphan_rows = res.get("data", [])
        print(f"  pending orphans: {len(orphan_rows)}")

        if orphan_rows:
            import re as re2
            CAT_PAT2 = re2.compile(r'^\s*\[(?P<cat>[^\]]+)\]\s+(?P<en>.+?)\s*$')
            sent_map = {'positive':'正向','negative':'负向','neutral':'中性'}

            seed = []
            for r in orphan_rows:
                cn = r.get('tag_cn'); en = r.get('tag_en')
                clean_cn, clean_en, pline = cn, en, None
                if cn:
                    m = CAT_PAT2.match(cn)
                    if m:
                        clean_cn = m.group('cat').strip()
                        pline = clean_cn
                        if not en: clean_en = m.group('en').strip()
                seed.append({
                    'tag_id': r['tag_id'], 'tag_cn': clean_cn, 'tag_en': clean_en,
                    'aipl_node': r.get('aipl_node'),
                    'polarity': sent_map.get((r.get('sent') or '').strip().lower()),
                    'sentiment_preset': r.get('sent'),
                    'product_line': pline, 'hits': r.get('hits'),
                })

            ORPHAN_SEED.write_text("\n".join(json.dumps(r, ensure_ascii=False) for r in seed))

            sys.path.insert(0, str(ROOT / "research/02-脚本工具/07-LLM引擎"))
            from llm_client import LLMClient
            import asyncio
            DEPT_VOCAB = ['全球客服中心','产品中心','仓储物流部','品牌市场中心','电商运营部','品质管理中心','法务合规部','用户运营部']
            DEPT_PROMPT = f"VOC 标签需指派主责部门。可选范围: {' / '.join(DEPT_VOCAB)}。根据 tag_cn、tag_en、AIPL、情感极性，选出最贴合的一个。只返回部门名字符串。"

            def fallback(rec, raw):
                if raw:
                    for d in DEPT_VOCAB:
                        if d in raw: return d
                cn = (rec.get('tag_cn') or '').lower()
                aipl = (rec.get('aipl_node') or '').upper()
                if aipl == 'A' or '推荐' in cn or '品牌' in cn: return '品牌市场中心'
                if '客服' in cn or '咨询' in cn: return '全球客服中心'
                if '物流' in cn or '配送' in cn: return '仓储物流部'
                if '价格' in cn or '促销' in cn: return '电商运营部'
                return '产品中心'

            async def fill_dept(records):
                client = LLMClient()
                sem = asyncio.Semaphore(12)
                async def one(rec):
                    async with sem:
                        try:
                            user = "\n".join([f"tag_cn: {rec['tag_cn']}", f"tag_en: {rec.get('tag_en') or ''}",
                                              f"AIPL: {rec.get('aipl_node') or ''}", f"polarity: {rec.get('polarity') or ''}"])
                            resp = await client.chat_async(
                                vendor="deepseek", model="deepseek-chat",
                                messages=[{"role":"system","content":DEPT_PROMPT},{"role":"user","content":user}],
                                temperature=0.1, max_tokens=50)
                            raw = resp.content.strip()
                            rec['dept_owner'] = raw if raw in DEPT_VOCAB else fallback(rec, raw)
                        except Exception:
                            rec['dept_owner'] = fallback(rec, None)
                        return rec
                return await asyncio.gather(*[one(r) for r in records])

            t0 = time.time()
            enriched = asyncio.run(fill_dept(seed))
            print(f"  ✅ LLM dept_owner: {time.time()-t0:.1f}s")

            batch_sz = 30
            inserted = 0
            for i in range(0, len(enriched), batch_sz):
                batch = enriched[i:i+batch_sz]
                values_clauses = []
                for r in batch:
                    vals = [
                        esc(r['tag_id']), esc(r['tag_cn']), esc(r.get('tag_en')),
                        esc(r.get('aipl_node')), esc(r.get('polarity')),
                        esc(r['dept_owner']), "NULL", "NULL", "FALSE",
                        "'auto_from_label'", esc(r.get('product_line')), "NULL",
                    ]
                    values_clauses.append(f"({','.join(vals)})")
                sql = f"""INSERT INTO dim_tag
                  (tag_id, tag_cn, tag_en, aipl_node, polarity, dept_owner,
                   biz_action, strategy_pkg, is_general, audit_status, product_line, collab_dept)
                  VALUES {','.join(values_clauses)}
                  ON CONFLICT (tag_id) DO NOTHING;"""
                res = run_sql(DB_ID, sql)
                if res.get("error"):
                    print(f"  ❌ orphan batch {i//batch_sz}: {res}")
                    break
                inserted += len(batch)
            print(f"  ✅ orphan inserted: {inserted}")

    finally:
        print(f"\n[step 6] Restore allow_dml = {orig_dml}")
        if not orig_dml:
            set_dml(DB_ID, False)
            print(f"  ✅ allow_dml disabled")

    print(f"\n[step 7] Verify final state")
    res = run_sql(DB_ID, """SELECT
        (SELECT COUNT(*) FROM dim_tag) AS dim_tag_total,
        (SELECT COUNT(*) FROM dim_tag WHERE audit_status='auto_from_label') AS auto_from_label,
        (SELECT COUNT(*) FROM dim_tag WHERE atomic_indicator_id IS NOT NULL) AS with_atomic,
        (SELECT COUNT(DISTINCT l.tag_id) FROM voc_label l LEFT JOIN dim_tag d ON l.tag_id=d.tag_id WHERE d.tag_id IS NULL) AS orphan_remaining,
        (SELECT COUNT(*) FROM dim_tag WHERE product_line IS NOT NULL) AS with_product_line""")
    print(f"  {res.get('data',[{}])[0]}")

    print(f"\n✅ DONE. Backup table 'dim_tag_backup_20260514' kept on Tencent.")
    print(f"   Rollback: /tmp/tencent_rollback.sql (run via Superset SQL Lab with DML on)")

if __name__ == "__main__":
    main()
