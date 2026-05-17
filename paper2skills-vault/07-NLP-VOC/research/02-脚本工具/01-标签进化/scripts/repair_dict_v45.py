"""v4.4 -> v4.5 dictionary repair.

Pipeline:
1. mechanical cleanup of '[品类] english' tag_cn (62 rows): split into pure tag_cn
2. LLM batch fill (230 cells):
   - main sheet 6 placeholders/empties (tag_cn / 主责部门 / 故事线 / 策略包)
   - product sheets 140 业务动作/责任部门
   - product sheets + main 84 策略包 (closed-set vocab)
3. write v4.5 xlsx + diff report

Modes:
  --sample        sample 5 of each LLM category, print proposals only (no write)
  --spot N        run LLM for first N cells of each category, print proposals
  --batch         run LLM for ALL cells, write v4.5
"""

from __future__ import annotations

import argparse
import asyncio
import copy
import json
import re
import shutil
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[2] / "07-LLM引擎"))

import openpyxl
from llm_client import LLMClient

ROOT = Path("/Users/pray/project/paper_to_skills/paper2skills-vault/07-NLP-VOC")
SRC = ROOT / "research/04-输出结果/01-字典版本/tag_dictionary_v4.4.xlsx"
DST = ROOT / "research/04-输出结果/01-字典版本/tag_dictionary_v4.5.xlsx"
WORKSPACE = ROOT / "research/04-输出结果/01-字典版本/_v45_workspace"
WORKSPACE.mkdir(parents=True, exist_ok=True)

MAIN_SHEET = "01_通用标签主表"
PRODUCT_SHEETS = ["02_吸奶器","03_内衣服饰","04_家居家纺","05_母婴综合护理","06_喂养电器","07_智能母婴电器"]

PLACEHOLDER_PAT = re.compile(r'^(【.*?】|未定义|待填|tbd|TBD|未填写|-|null|None|N/A|n/a)$', re.IGNORECASE)
CATEGORY_PREFIX_PAT = re.compile(r'^\s*\[(?P<cat>[^\]]+)\]\s+(?P<en>.+?)\s*$')

STRATEGY_VOCABULARY = [
    "核心体验改良包","履约提速包","品牌声量监测包","信息架构优化包","一次解决提升包",
    "售后服务提升包","合规背书强化包","物流体验优化包","产品体验优化包","质量缺陷闭环包",
    "竞品对比内容包","价值表达优化包","质量回溯包","会员权益优化包","私域内容优化包",
    "个性化运营增强包","品牌认知校准包","人群内容匹配包","种草真实度优化包","达人合作治理包",
    "评价参考价值优化包","到手价透明包","证据内容补强包","跨平台口径统一包","核心卖点校准包",
    "复购激活包","舆情应急修复包","需求缺口闭环包","承诺一致性校对包","产品安全紧急处理包",
    "退货体验优化包","订单体验优化包","广告证据化治理包","频控降扰包","首响提速包",
    "信任恢复运营包","社群活跃提升包","宣传兑现校正包","库存管理优化包","订单流程优化包",
    "支付体验优化包","物流时效提升包","末端配送治理包","退货流程优化包","退款时效优化包",
    "质保服务优化包","客服响应优化包","产品信息完善包","用户满意度提升包",
]

def is_placeholder(v) -> bool:
    if v is None:
        return False
    s = str(v).strip()
    return s == "" or bool(PLACEHOLDER_PAT.match(s))

def is_empty(v) -> bool:
    return v is None or str(v).strip() == ""

def header_index(ws, name):
    h = [c.value for c in ws[1]]
    return h.index(name) if name in h else None

def scan_repairs(wb):
    """Return list of repair tasks. Each task = dict(kind, sheet, row, col, ctx, current).

    Two-pass: first collect cleanup_tag_cn tasks, then collect LLM tasks using
    cleaned tag_cn so LLM never sees '[品类] english' dirty values.
    """
    cleanup_map = {}
    tasks = []
    for sn in PRODUCT_SHEETS:
        ws = wb[sn]
        cn_idx = header_index(ws, "VOC标签（中文）")
        for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            cn_val = row[cn_idx]
            mobj = CATEGORY_PREFIX_PAT.match(str(cn_val)) if cn_val else None
            if mobj:
                cleaned_cn = mobj.group("cat").strip()
                cleanup_map[(sn, r_idx)] = (cleaned_cn, mobj.group("en").strip())

    ws_main = wb[MAIN_SHEET]
    m = {n: header_index(ws_main, n) for n in
         ["标签ID","VOC标签（中文）","VOC标签（英文）","AIPL节点","情感极性",
          "故事线关联","策略包","主责部门","标签定义"]}

    for r_idx, row in enumerate(ws_main.iter_rows(min_row=2, values_only=True), 2):
        if not row[m["标签ID"]]:
            continue
        ctx = {
            "tag_id": row[m["标签ID"]],
            "tag_cn": row[m["VOC标签（中文)"]] if "VOC标签（中文)" in m else row[m["VOC标签（中文）"]],
            "tag_en": row[m["VOC标签（英文）"]],
            "aipl":   row[m["AIPL节点"]],
            "polarity": row[m["情感极性"]],
            "definition": row[m["标签定义"]],
            "dept_owner": row[m["主责部门"]],
        }
        if is_placeholder(row[m["VOC标签（中文）"]]):
            tasks.append({"kind":"tag_cn", "sheet":MAIN_SHEET, "row":r_idx,
                          "col":m["VOC标签（中文）"]+1, "ctx":ctx,
                          "current":row[m["VOC标签（中文）"]]})
        for f in ["故事线关联","策略包","主责部门"]:
            if is_empty(row[m[f]]):
                tasks.append({"kind":f, "sheet":MAIN_SHEET, "row":r_idx,
                              "col":m[f]+1, "ctx":ctx, "current":row[m[f]]})

    for sn in PRODUCT_SHEETS:
        ws = wb[sn]
        p = {n: header_index(ws, n) for n in
             ["标签ID","VOC标签（中文）","VOC标签（英文）","AIPL节点","情感极性",
              "故事线关联","策略包","业务动作/责任部门","主责部门","协同部门","标签定义"]}
        for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if not row[p["标签ID"]]:
                continue
            cleaned = cleanup_map.get((sn, r_idx))
            tag_cn_for_ctx = cleaned[0] if cleaned else row[p["VOC标签（中文）"]]
            tag_en_for_ctx = row[p["VOC标签（英文）"]] or (cleaned[1] if cleaned else None)
            ctx = {
                "tag_id": row[p["标签ID"]],
                "tag_cn": tag_cn_for_ctx,
                "tag_en": tag_en_for_ctx,
                "aipl":   row[p["AIPL节点"]],
                "polarity": row[p["情感极性"]],
                "definition": row[p["标签定义"]],
                "dept_owner": row[p["主责部门"]],
                "co_dept":    row[p["协同部门"]],
                "product_line": sn[3:],
            }
            if cleaned:
                tasks.append({"kind":"cleanup_tag_cn", "sheet":sn, "row":r_idx,
                              "col":p["VOC标签（中文）"]+1, "ctx":ctx,
                              "current":row[p["VOC标签（中文）"]],
                              "proposed_cn": cleaned[0],
                              "proposed_en_fallback": cleaned[1]})
            for f in ["业务动作/责任部门","策略包"]:
                if is_empty(row[p[f]]):
                    tasks.append({"kind":f, "sheet":sn, "row":r_idx,
                                  "col":p[f]+1, "ctx":ctx, "current":row[p[f]]})
    return tasks

PROMPTS = {
    "tag_cn": (
        "你是 VOC 标签词典编辑。下面这个标签的中文名缺失/占位。根据其 tag_en、AIPL、情感极性、定义，"
        "给出 4-12 字的标准中文标签名。只返回中文标签字符串，不要解释，不要引号。\n"
    ),
    "主责部门": (
        "VOC 标签需指派主责部门。可选范围（机器可读）: 全球客服中心 / 产品中心 / 仓储物流部 / "
        "品牌市场中心 / 电商运营部 / 品质管理中心 / 法务合规部。"
        "根据 tag_cn、tag_en、AIPL、情感极性、定义，选出最贴合的一个。只返回部门名字符串。\n"
    ),
    "故事线关联": (
        "VOC 标签需关联故事线。可选范围: 承诺偏差 / Gap聚类 / 信任恢复 / 国家×SKU优先级 / "
        "履约链路 / 复购裂变 / 内容种草 / 缺陷闭环 / 满意度。可多选用分号分隔，最多 2 个。"
        "根据 tag_cn、tag_en、AIPL、情感极性，给出最贴合的 1-2 个故事线。只返回结果字符串。\n"
    ),
    "策略包": (
        "VOC 标签需关联策略包。必须从下列闭集中选 1 个：\n" +
        "  ".join(STRATEGY_VOCABULARY) + "\n"
        "根据 tag_cn、tag_en、AIPL、情感极性、主责部门，选出最贴合的一个。只返回策略包名字符串，必须严格在上述闭集中。\n"
    ),
    "业务动作/责任部门": (
        "VOC 标签需要一行『业务动作描述』，格式『{主责部门}：{动词开头的具体动作}』，20-50 字之间。"
        "动作必须可执行、贴合 tag_cn 语义、与情感极性方向一致（负向→修复改进，正向→放大复用）。"
        "只返回完整的业务动作字符串，不要引号，不要解释。\n"
    ),
}

def build_user_prompt(task):
    ctx = task["ctx"]
    parts = [
        f"tag_id: {ctx['tag_id']}",
        f"tag_cn: {ctx['tag_cn']}",
        f"tag_en: {ctx['tag_en']}",
        f"AIPL: {ctx['aipl']}",
        f"polarity: {ctx['polarity']}",
    ]
    if ctx.get("definition"):
        parts.append(f"definition: {ctx['definition']}")
    if ctx.get("dept_owner"):
        parts.append(f"主责部门: {ctx['dept_owner']}")
    if ctx.get("co_dept"):
        parts.append(f"协同部门: {ctx['co_dept']}")
    if ctx.get("product_line"):
        parts.append(f"产品品线: {ctx['product_line']}")
    return "\n".join(parts)

async def llm_call(client: LLMClient, task):
    sys_prompt = PROMPTS[task["kind"]]
    user_prompt = build_user_prompt(task)
    messages = [
        {"role": "system", "content": sys_prompt},
        {"role": "user", "content": user_prompt},
    ]
    resp = await client.chat_async(
        vendor="deepseek",
        messages=messages,
        model="deepseek-chat",
        temperature=0.1,
        max_tokens=200,
    )
    return resp.content.strip()

async def run_llm_batch(tasks, concurrency=10):
    client = LLMClient()
    sem = asyncio.Semaphore(concurrency)

    async def _one(t):
        async with sem:
            try:
                t["proposed"] = await llm_call(client, t)
                t["status"] = "ok"
            except Exception as e:
                t["status"] = f"error: {e}"
            return t

    results = await asyncio.gather(*[_one(t) for t in tasks])
    return results

def apply_to_workbook(wb, tasks):
    """Apply proposed values back to workbook in-memory."""
    applied = 0
    skipped = 0
    for t in tasks:
        if t["kind"] == "cleanup_tag_cn":
            ws = wb[t["sheet"]]
            cell_cn = ws.cell(row=t["row"], column=t["col"])
            cell_cn.value = t["proposed_cn"]
            en_col_idx = header_index(ws, "VOC标签（英文）") + 1
            cur_en = ws.cell(row=t["row"], column=en_col_idx).value
            if not cur_en or str(cur_en).strip() == "":
                ws.cell(row=t["row"], column=en_col_idx).value = t["proposed_en_fallback"]
            applied += 1
            continue
        if t.get("status") != "ok":
            skipped += 1
            continue
        ws = wb[t["sheet"]]
        ws.cell(row=t["row"], column=t["col"]).value = t["proposed"]
        applied += 1
    return applied, skipped

def write_diff_report(tasks, out_path):
    lines = [
        f"# tag_dictionary v4.4 → v4.5 修复 diff 报告",
        f"",
        f"- 生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f"- 源文件：tag_dictionary_v4.4.xlsx",
        f"- 目标文件：tag_dictionary_v4.5.xlsx",
        f"",
        f"## 修复汇总",
        f"",
    ]
    by_kind = Counter(t["kind"] for t in tasks)
    by_sheet_kind = Counter((t["sheet"], t["kind"]) for t in tasks)
    lines.append("| 修复类型 | 总数 |")
    lines.append("|---|---:|")
    for k, v in by_kind.most_common():
        lines.append(f"| {k} | {v} |")
    lines.append("")
    lines.append("## 分 Sheet 修复明细")
    lines.append("")
    lines.append("| Sheet | 修复类型 | 行数 |")
    lines.append("|---|---|---:|")
    for (sn, k), v in sorted(by_sheet_kind.items()):
        lines.append(f"| {sn} | {k} | {v} |")
    lines.append("")
    lines.append("## 修复样例（前 30 条）")
    lines.append("")
    for i, t in enumerate(tasks[:30], 1):
        if t["kind"] == "cleanup_tag_cn":
            new_v = t["proposed_cn"]
        else:
            new_v = t.get("proposed", "(skipped)")
        lines.append(f"### {i}. {t['sheet']} row {t['row']} · {t['kind']}")
        lines.append(f"- tag_id: `{t['ctx']['tag_id']}`")
        lines.append(f"- tag_cn: `{t['ctx']['tag_cn']}`")
        lines.append(f"- 原值: `{t.get('current')!r}`")
        lines.append(f"- 新值: `{new_v}`")
        lines.append("")
    out_path.write_text("\n".join(lines), encoding="utf-8")

async def main_async(mode, n):
    wb = openpyxl.load_workbook(SRC)
    tasks = scan_repairs(wb)
    print(f"Total repair tasks: {len(tasks)}")
    print(f"  by kind: {Counter(t['kind'] for t in tasks)}")

    cleanup_tasks = [t for t in tasks if t["kind"] == "cleanup_tag_cn"]
    llm_tasks = [t for t in tasks if t["kind"] != "cleanup_tag_cn"]
    print(f"  cleanup (mechanical): {len(cleanup_tasks)}")
    print(f"  llm needed: {len(llm_tasks)}")

    if mode == "sample":
        per_kind = {}
        for t in llm_tasks:
            per_kind.setdefault(t["kind"], []).append(t)
        sample = []
        for k, lst in per_kind.items():
            sample.extend(lst[:2])
        print(f"\n--- SAMPLE MODE: running {len(sample)} LLM cells (2 per kind) + previewing {min(5, len(cleanup_tasks))} cleanups ---")
        if sample:
            sample = await run_llm_batch(sample, concurrency=5)
        for t in cleanup_tasks[:5]:
            print(f"\n[cleanup] {t['sheet']} r{t['row']} | '{t['current']}' → tag_cn='{t['proposed_cn']}'  tag_en_fallback='{t['proposed_en_fallback']}'")
        for t in sample:
            print(f"\n[{t['kind']}] {t['sheet']} r{t['row']} | tag={t['ctx']['tag_cn']!r}")
            print(f"  proposed: {t.get('proposed', t.get('status'))!r}")
        return

    if mode == "spot":
        spot_tasks = []
        per_kind = {}
        for t in llm_tasks:
            per_kind.setdefault(t["kind"], []).append(t)
        for k, lst in per_kind.items():
            spot_tasks.extend(lst[:n])
        print(f"\n--- SPOT MODE: running {len(spot_tasks)} LLM cells ({n} per kind) ---")
        spot_tasks = await run_llm_batch(spot_tasks, concurrency=10)
        out = WORKSPACE / f"spot_check_{n}.jsonl"
        with open(out, "w") as f:
            for t in spot_tasks:
                ent = {k:v for k,v in t.items() if k != "ctx"}
                ent["ctx"] = {k:v for k,v in t["ctx"].items() if v is not None}
                f.write(json.dumps(ent, ensure_ascii=False, default=str) + "\n")
        print(f"Wrote spot check: {out}")
        for t in spot_tasks[:10]:
            print(f"  [{t['kind']}] {t['sheet']} | tag={t['ctx']['tag_cn']!r} → {t.get('proposed', t.get('status'))!r}")
        return

    if mode == "batch":
        print(f"\n--- BATCH MODE: applying {len(cleanup_tasks)} cleanups + {len(llm_tasks)} LLM calls ---")
        llm_tasks = await run_llm_batch(llm_tasks, concurrency=10)
        all_tasks = cleanup_tasks + llm_tasks
        applied, skipped = apply_to_workbook(wb, all_tasks)
        print(f"Applied: {applied}  Skipped: {skipped}")
        wb.save(DST)
        print(f"Wrote: {DST}")
        diff_path = ROOT / "research/04-输出结果/01-字典版本/dept_repair_v45_diff.md"
        write_diff_report(all_tasks, diff_path)
        print(f"Diff report: {diff_path}")
        return

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--mode", choices=["sample","spot","batch"], required=True)
    parser.add_argument("-n", type=int, default=5)
    args = parser.parse_args()
    asyncio.run(main_async(args.mode, args.n))
