"""Repair v4.2 -> v4.3 dictionary: enum normalization + schema alignment + dept misplacement cleanup.

Default mappings (option A · 默认):

1. AIPL 节点 旧 → 新（按业务语义对齐）:
     A   → L1 Awareness
     I   → L2 Interest
     P   → L3 Purchase   (P/P1/P2/P3 的「购买阶段」)
     P1  → L3
     P2  → L3
     P3  → L3
     B   → L4 Loyalty   (B 表示 Brand/Loyalty 阶段)
     B1  → L4
     A1  → L1
     A3  → L1
     I1  → L2
     L1/L2/L3/L4 keep as-is

2. Proxy NPS 贡献 旧 → 新:
     Detractor驱动      → detractor
     Strong_Detractor   → detractor
     Promoter驱动       → promoter
     Strong_Promoter    → promoter
     Passive驱动        → passive
     Passive影响        → passive
     中性               → passive
     中性-Passive       → passive
     promoter/passive/detractor 保持
     空 / None 保持空

3. Schema 对齐: 所有品类 sheet (02-07) 补齐主表 5 列:
     合理性评分 / 风险等级 / 问题诊断 / 优化建议 / 优化优先级
     新增列保持空值

4. 部门字段错位: 把 3 类业务动作字符串从 "主责部门 / 协同部门" 列移到 "业务动作/责任部门" 列:
     分析竞品对比趋势 (18)
     降低负面反馈率   (8)
     推动供应商改进   (8)
"""

from __future__ import annotations
import sys
from collections import Counter
from copy import copy
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[3]
SRC = ROOT / "04-输出结果/01-字典版本/tag_dictionary_v4.2.xlsx"
DST = ROOT / "04-输出结果/01-字典版本/tag_dictionary_v4.3.xlsx"
DIFF = ROOT / "04-输出结果/01-字典版本/dept_repair_v43_diff.md"

PRIMARY = ["01_通用标签主表", "02_吸奶器", "03_内衣服饰", "04_家居家纺",
           "05_母婴综合护理", "06_喂养电器", "07_智能母婴电器"]

AIPL_MAP = {
    "A": "L1", "A1": "L1", "A3": "L1",
    "I": "L2", "I1": "L2",
    "P": "L3", "P1": "L3", "P2": "L3", "P3": "L3",
    "B": "L4", "B1": "L4",
}

NPS_MAP = {
    "Detractor驱动": "detractor",
    "Strong_Detractor": "detractor",
    "strong_detractor": "detractor",
    "Promoter驱动": "promoter",
    "Strong_Promoter": "promoter",
    "strong_promoter": "promoter",
    "Passive驱动": "passive",
    "Passive影响": "passive",
    "中性": "passive",
    "中性-Passive": "passive",
    "weak_promoter": "passive",
    "weak_detractor": "passive",
}

DEPT_TO_BIZACTION = {"分析竞品对比趋势", "降低负面反馈率", "推动供应商改进"}

EXTRA_COLS = ["合理性评分", "风险等级", "问题诊断", "优化建议", "优化优先级"]


def normalize_aipl(v):
    if v is None: return v
    s = str(v).strip()
    return AIPL_MAP.get(s, s)


def normalize_nps(v):
    if v is None: return v
    s = str(v).strip()
    if not s: return v
    return NPS_MAP.get(s, s)


def main():
    if not SRC.exists():
        print(f"ERROR: source not found: {SRC}", file=sys.stderr)
        sys.exit(2)

    print(f"Reading {SRC}")
    wb = openpyxl.load_workbook(SRC)

    log_aipl = Counter()
    log_nps = Counter()
    log_schema = []
    log_dept_moved = Counter()
    log_rows = {}

    for sname in PRIMARY:
        ws = wb[sname]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        n_rows_before = ws.max_row - 1
        log_rows[sname] = {"before": n_rows_before}

        def col_idx(name):
            return headers.index(name) if name in headers else None

        aipl_i = col_idx("AIPL节点")
        nps_i = col_idx("Proxy NPS贡献")
        zhuze_i = col_idx("主责部门")
        xietong_i = col_idx("协同部门")
        bizact_i = col_idx("业务动作/责任部门")

        for row in ws.iter_rows(min_row=2):
            if aipl_i is not None:
                cell = row[aipl_i]
                old = cell.value
                new = normalize_aipl(old)
                if new != old and new is not None:
                    log_aipl[(str(old), str(new))] += 1
                    cell.value = new
            if nps_i is not None:
                cell = row[nps_i]
                old = cell.value
                new = normalize_nps(old)
                if new != old and new is not None:
                    log_nps[(str(old), str(new))] += 1
                    cell.value = new

        if (zhuze_i is not None or xietong_i is not None) and bizact_i is not None:
            for row in ws.iter_rows(min_row=2):
                for col_i in (zhuze_i, xietong_i):
                    if col_i is None: continue
                    cell = row[col_i]
                    if cell.value is None: continue
                    s = str(cell.value).strip()
                    if s in DEPT_TO_BIZACTION:
                        ba_cell = row[bizact_i]
                        existing = str(ba_cell.value).strip() if ba_cell.value else ""
                        if existing and s not in existing:
                            ba_cell.value = f"{existing}; {s}"
                        elif not existing:
                            ba_cell.value = s
                        cell.value = None
                        log_dept_moved[s] += 1

        if sname != "01_通用标签主表":
            current_max_col = ws.max_column
            for extra in EXTRA_COLS:
                if extra not in headers:
                    current_max_col += 1
                    ws.cell(row=1, column=current_max_col, value=extra)
                    log_schema.append((sname, extra))

        log_rows[sname]["after"] = ws.max_row - 1

    print(f"Writing {DST}")
    wb.save(DST)

    lines = [
        "# v4.2 → v4.3 修复 diff 报告",
        "",
        f"- 源: `{SRC.name}`",
        f"- 目标: `{DST.name}`",
        "",
        "## 1. AIPL 节点枚举重命名",
        "",
        "| 旧值 | 新值 | 替换次数 |",
        "|---|---|---|",
    ]
    for (old, new), cnt in sorted(log_aipl.items(), key=lambda x: -x[1]):
        lines.append(f"| {old} | {new} | {cnt} |")
    if not log_aipl:
        lines.append("| (无) | | |")

    lines += [
        "",
        "## 2. Proxy NPS 贡献枚举重命名",
        "",
        "| 旧值 | 新值 | 替换次数 |",
        "|---|---|---|",
    ]
    for (old, new), cnt in sorted(log_nps.items(), key=lambda x: -x[1]):
        lines.append(f"| {old} | {new} | {cnt} |")
    if not log_nps:
        lines.append("| (无) | | |")

    lines += [
        "",
        "## 3. 部门字段错位：业务动作字符串移到 业务动作/责任部门 列",
        "",
        "| 字符串 | 移动次数 |",
        "|---|---|",
    ]
    for s, cnt in sorted(log_dept_moved.items(), key=lambda x: -x[1]):
        lines.append(f"| {s} | {cnt} |")
    if not log_dept_moved:
        lines.append("| (无) | |")

    lines += [
        "",
        "## 4. Schema 对齐：品类 sheet 补齐主表 5 列",
        "",
        "| sheet | 新增列 |",
        "|---|---|",
    ]
    for s, col in log_schema:
        lines.append(f"| {s} | {col} |")
    if not log_schema:
        lines.append("| (无) | |")

    lines += [
        "",
        "## 5. 行数完整性",
        "",
        "| sheet | 修复前 | 修复后 |",
        "|---|---|---|",
    ]
    for s, d in log_rows.items():
        lines.append(f"| {s} | {d['before']} | {d['after']} |")

    DIFF.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"Diff -> {DIFF}")
    print(f"\nAIPL changes: {sum(log_aipl.values())}")
    print(f"NPS changes:  {sum(log_nps.values())}")
    print(f"Dept moves:   {sum(log_dept_moved.values())}")
    print(f"Schema cols added: {len(log_schema)}")


if __name__ == "__main__":
    main()
