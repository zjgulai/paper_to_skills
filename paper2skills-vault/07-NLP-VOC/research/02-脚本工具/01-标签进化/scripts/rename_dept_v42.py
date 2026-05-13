"""Generate v4.2 tag dictionary from v4.1 by renaming dept names.

Per `09-部门重命名/01-rename-canonical-mapping.md`:

  19 mapping rules + Q3=A rule (compound `部门：描述` strings -> only the
  prefix before `：` is renamed; the trailing business description is kept).

Outputs:
  - tag_dictionary_v4.2.xlsx
  - dept_rename_v42_diff.md  (every cell that changed, grouped by sheet)

This file ONLY touches the new v4.2 workbook. v4.1 is read-only.
"""

from __future__ import annotations

import re
import sys
from collections import Counter
from pathlib import Path

import openpyxl
from openpyxl import load_workbook

DEPT_COLUMNS = {"主责部门", "协同部门", "业务动作/责任部门"}

MAPPING: dict[str, str] = {
    "全球客服与体验中心": "全球客服中心",
    "产品中心/品线":      "产品中心",
    "供应链中心":         "仓储物流部",
    "品控部":             "品质管理中心",
    "质量与法规部":       "法务合规部",
    "市场部":             "品牌市场中心",
    "产品市场部":         "产品中心",
    "客服部":             "全球客服中心",
    "会员运营部":         "用户运营部",
    "品牌运营部":         "品牌市场中心",
    "私域运营部":         "用户运营部",
    "产品研发部":         "产品中心",
    "声学实验室":         "产品中心",
    "国际物流部":         "仓储物流部",
    "产品规划部":         "产品中心",
    "物流运营部":         "仓储物流部",
    "产品运营部":         "产品中心",
    "运营部":             "电商运营部",
    "产品部":             "产品中心",
    "全球营销服务中心":   "品牌市场中心",
    "营运部":             "电商运营部",
    "产品设计部":         "产品中心",
    "供应链部":           "仓储物流部",
    "研发部":             "产品中心",
    "设计部":             "产品中心",
    "公关":               "品牌市场中心",
    "法务":               "法务合规部",
    "用户体验部":         "全球客服中心",
}

SUFFIX_PROTECT = (
    "电商运营部", "物流运营部", "品牌运营部", "私域运营部",
    "用户运营部", "会员运营部", "产品运营部",
)

ROOT = Path(__file__).resolve().parents[3]
SRC = ROOT / "04-输出结果/01-字典版本/tag_dictionary_v4.1.xlsx"
DST = ROOT / "04-输出结果/01-字典版本/tag_dictionary_v4.2.xlsx"
DIFF = ROOT / "04-输出结果/01-字典版本/dept_rename_v42_diff.md"


def _split_delims(text: str) -> list[tuple[str, str]]:
    """Split a cell value on ; ；, , whitespace -- preserving the delimiter.

    Returns list of (token, delim_after). Last token has empty delim.
    """
    parts: list[tuple[str, str]] = []
    buf: list[str] = []
    for ch in text:
        if ch in "；;,，":
            parts.append(("".join(buf), ch))
            buf = []
        else:
            buf.append(ch)
    parts.append(("".join(buf), ""))
    return parts


def _rename_single_token(token: str) -> tuple[str, bool]:
    """Rename one bare token. Returns (new_token, changed?)."""
    stripped = token.strip()
    if not stripped:
        return token, False

    # 1. Compound "部门：描述" — replace only the prefix.
    m = re.match(r"^([^：:]+)([：:])(.*)$", stripped)
    if m:
        prefix, sep, rest = m.group(1), m.group(2), m.group(3)
        for old, new in MAPPING.items():
            if prefix == old:
                new_full = new + sep + rest
                lead = token[: len(token) - len(token.lstrip())]
                tail = token[len(token.rstrip()):]
                return f"{lead}{new_full}{tail}", True
        return token, False

    # 2. Plain dept name (exact match against MAPPING keys).
    for old, new in MAPPING.items():
        if stripped == old:
            lead = token[: len(token) - len(token.lstrip())]
            tail = token[len(token.rstrip()):]
            return f"{lead}{new}{tail}", True

    # 3. Standalone '运营部' inside a longer token like 'KOL运营部'/'内容运营部' —
    #    suffix-protected, do NOT rewrite. Same for '部门' inside other tokens.
    if "运营部" in stripped and not any(stripped == s or stripped.endswith(s) for s in SUFFIX_PROTECT):
        # Treat as named department only if EXACTLY '运营部'; otherwise keep
        # (KOL运营部 / 内容运营部 / 培训部 are out of scope per Q2=A).
        pass

    return token, False


def rename_cell(value: object) -> tuple[object, list[tuple[str, str]]]:
    """Apply all rename rules to a cell value.

    Returns (new_value, [(old_token, new_token), ...]).
    """
    if value is None:
        return value, []
    text = str(value)
    if not text.strip():
        return value, []

    changes: list[tuple[str, str]] = []
    tokens = _split_delims(text)
    new_tokens: list[tuple[str, str]] = []
    for tok, delim in tokens:
        new_tok, changed = _rename_single_token(tok)
        if changed:
            changes.append((tok.strip(), new_tok.strip()))
        new_tokens.append((new_tok, delim))
    new_text = "".join(t + d for t, d in new_tokens)
    if new_text == text:
        return value, []
    return new_text, changes


def process_workbook(src: Path, dst: Path, diff_md: Path) -> None:
    print(f"Reading {src}")
    wb = load_workbook(src)
    diff_lines: list[str] = []
    total_changes = 0
    sheet_change_counter: Counter[str] = Counter()
    dept_freq_before: Counter[str] = Counter()
    dept_freq_after: Counter[str] = Counter()

    for sname in wb.sheetnames:
        ws = wb[sname]
        headers: list[str | None] = []
        first = True
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=False):
            for cell in row:
                headers.append(str(cell.value).strip() if cell.value else None)
            first = False
            break

        dept_col_idx = [i for i, h in enumerate(headers) if h in DEPT_COLUMNS]
        if not dept_col_idx:
            continue

        per_sheet: list[str] = []
        for row in ws.iter_rows(min_row=2):
            for ci in dept_col_idx:
                cell = row[ci]
                if cell.value is None:
                    continue
                # before stats
                for old in MAPPING:
                    if old in str(cell.value):
                        dept_freq_before[old] += 1
                new_value, changes = rename_cell(cell.value)
                if changes:
                    old_v = cell.value
                    cell.value = new_value
                    per_sheet.append(
                        f"- row {cell.row} col {headers[ci]} (`{ci}`): "
                        f"`{old_v}` → `{new_value}`"
                    )
                    for o, n in changes:
                        diff_lines.append(f"  - `{o}` → `{n}`")
                # after stats
                if new_value:
                    for n in MAPPING.values():
                        if str(new_value) == n or (isinstance(new_value, str) and (
                            new_value.startswith(n + "：")
                            or new_value.startswith(n + ":")
                            or any(p == n for p in re.split(r"[；;,，]", new_value))
                        )):
                            dept_freq_after[n] += 1

        if per_sheet:
            sheet_change_counter[sname] += len(per_sheet)
            total_changes += len(per_sheet)

    print(f"Writing {dst}")
    wb.save(dst)

    diff_md.write_text(
        _build_diff_report(
            src, dst, total_changes, sheet_change_counter,
            dept_freq_before, dept_freq_after,
        ),
        encoding="utf-8",
    )
    print(f"Diff -> {diff_md}")
    print(f"Total dept cells changed: {total_changes}")
    print(f"Sheets touched: {len(sheet_change_counter)}")


def _build_diff_report(
    src: Path, dst: Path, total: int,
    sheet_counter: Counter[str],
    before: Counter[str], after: Counter[str],
) -> str:
    lines = [
        "# 字典 v4.1 → v4.2 部门重命名 diff 报告",
        "",
        f"- 源: `{src.name}`",
        f"- 目标: `{dst.name}`",
        f"- 总变更 cell 数: **{total}**",
        f"- 受影响 sheet 数: **{len(sheet_counter)}**",
        "",
        "## 按 sheet 变更分布",
        "",
        "| sheet | 变更 cell 数 |",
        "|---|---|",
    ]
    for s, n in sheet_counter.most_common():
        lines.append(f"| {s} | {n} |")
    lines += [
        "",
        "## 老部门名 → 命中次数 (执行前)",
        "",
        "| old | hits |",
        "|---|---|",
    ]
    for old, _ in sorted(before.items(), key=lambda kv: -kv[1]):
        if before[old] > 0:
            lines.append(f"| {old} | {before[old]} |")
    lines += [
        "",
        "## 新部门名 → 命中次数 (执行后)",
        "",
        "| new | hits |",
        "|---|---|",
    ]
    targets = sorted(set(MAPPING.values()))
    for n in targets:
        lines.append(f"| {n} | {after[n]} |")
    lines += [
        "",
        "## 映射规则参考",
        "",
        "见 `09-部门重命名/01-rename-canonical-mapping.md`",
    ]
    return "\n".join(lines) + "\n"


if __name__ == "__main__":
    if not SRC.exists():
        print(f"ERROR: source not found: {SRC}", file=sys.stderr)
        sys.exit(2)
    if DST.exists():
        print(f"INFO: overwriting existing {DST}")
    process_workbook(SRC, DST, DIFF)
    print("DONE")
