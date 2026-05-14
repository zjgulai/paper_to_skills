"""Map 267 tags → 50 SAT_xxx atomic indicators, produce v4.4 dict + diff report.

Mapping rules: 标签主题 → SAT_xxx (per 02-sat-indicators-draft.md).
Output:
- tag_dictionary_v4.4.xlsx (new column atomic_indicator_id added to all primary sheets)
- mvp_l2_sat_mapping_diff.md (orphan tag list + coverage stats)
"""

from __future__ import annotations
import sys
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[3]
SRC = ROOT / "04-输出结果/01-字典版本/tag_dictionary_v4.3.xlsx"
DST = ROOT / "04-输出结果/01-字典版本/tag_dictionary_v4.4.xlsx"
DIFF = ROOT / "04-输出结果/01-字典版本/mvp_l2_sat_mapping_diff.md"

PRIMARY = [
    "01_通用标签主表", "02_吸奶器", "03_内衣服饰", "04_家居家纺",
    "05_母婴综合护理", "06_喂养电器", "07_智能母婴电器",
]

THEME_TO_SAT = {
    "产品核心性能": "SAT_L1_01",
    "产品质量缺陷": "SAT_L1_02",
    "使用舒适度":   "SAT_L1_03",
    "产品体验负面": "SAT_L1_03",
    "操作便捷性":   "SAT_L1_04",
    "产品易用性":   "SAT_L1_04",
    "静音与噪音感知": "SAT_L1_05",
    "产品属性-静音": "SAT_L1_05",
    "配件问题":     "SAT_L1_06",
    "充电问题":     "SAT_L1_07",
    "产品属性-无线": "SAT_L1_07",
    "产品安全":     "SAT_L1_08",
    "产品耐用性":   "SAT_L1_09",
    "产品属性-便携": "SAT_L1_10",
    "产品属性-材质": "SAT_L1_11",
    "外观与设计":   "SAT_L1_12",
    "产品属性-颜色": "SAT_L1_12",
    "性能表现":     "SAT_L1_13",
    "质量感知":     "SAT_L1_13",
    "尺码合身":     "SAT_L1_14",
    "使用场景-夜间": "SAT_L1_15",
    "使用场景-工作": "SAT_L1_15",
    "使用场景-旅行": "SAT_L1_15",
    "客服售后":     "SAT_L1_16",
    "品牌初始印象": "SAT_L1_17",
    "种草内容可信度": "SAT_L1_18",
    "达人推荐可信度": "SAT_L1_18",
    "需求内容匹配":  "SAT_L1_18",
    "广告打扰感":    "SAT_L1_18",
    "广告真实性感知": "SAT_L1_18",
    "宣传兑现度":    "SAT_L1_18",

    "参数规格透明度": "SAT_L2_01",
    "信息获取便捷性": "SAT_L2_02",
    "价格透明度":     "SAT_L2_02",
    "跨平台信息一致性": "SAT_L2_03",
    "评价参考价值":   "SAT_L2_04",
    "证据内容充分度": "SAT_L2_05",
    "售后响应速度":   "SAT_L2_06",
    "客服专业度":     "SAT_L2_07",
    "问题解决效率":   "SAT_L2_08",
    "品牌信任恢复":   "SAT_L2_09",
    "复购意愿":       "SAT_L2_10",
    "质量投诉闭环":   "SAT_L2_11",
    "产品咨询":       "SAT_L2_12",

    "物流时效":     "SAT_L3_01",
    "运输时长":     "SAT_L3_01",
    "发货延迟":     "SAT_L3_01",
    "物流跟踪":     "SAT_L3_02",
    "配送体验":     "SAT_L3_03",
    "配送窗口":     "SAT_L3_03",
    "投递问题":     "SAT_L3_03",
    "包装完整性":   "SAT_L3_04",
    "包裹丢失":     "SAT_L3_05",
    "价格价值感":   "SAT_L3_06",
    "价格与价值":   "SAT_L3_06",
    "承诺一致性":   "SAT_L3_07",
    "安全/合规认证可信度": "SAT_L3_08",
    "核心功能认知准确度":   "SAT_L3_09",
    "竞品对比支持":   "SAT_L3_10",
    "第三方测评可信度": "SAT_L3_11",
    "订单取消":       "SAT_L3_12",
    "取消原因":       "SAT_L3_12",
    "退货原因":       "SAT_L3_13",
    "客服体验":       "SAT_L3_14",
    "个性化服务":     "SAT_L3_15",
    "会员权益感知":   "SAT_L3_16",
    "私域内容相关性": "SAT_L3_17",
    "社群活跃与参与度": "SAT_L3_17",
    "负面舆情修复":   "SAT_L3_17",
    "总体正面情感":   "SAT_L3_17",
    "推荐意愿":       "SAT_L3_17",
    "用户满意度":     "SAT_L3_17",
    "订单履约":       "SAT_L3_01",

    "品牌提及":     "SAT_L4_01",

    "自动发现-待审核": "SAT_L1_03",
    "产品功能":     "SAT_L1_01",
    "产品问题":     "SAT_L1_02",
    "清洁与维护便利度": "SAT_L1_04",
    "评估对比":     "SAT_L3_10",
    "品牌认知":     "SAT_L1_17",
    "需求缺口（GAP）": "SAT_L1_18",
    "信息查询":     "SAT_L2_02",
    "产品设计":     "SAT_L1_12",
    "持续使用":     "SAT_L2_10",
    "复购行为":     "SAT_L2_10",
    "材质":         "SAT_L1_11",
    "材质安全":     "SAT_L1_08",
    "【待填写】":   "SAT_L1_03",
}

THEME_BY_NODE_FALLBACK = {
    "L1": "SAT_L1_03",
    "L2": "SAT_L2_08",
    "L3": "SAT_L3_14",
    "L4": "SAT_L4_01",
}

VALID_SAT_IDS = set(THEME_TO_SAT.values()) | set(THEME_BY_NODE_FALLBACK.values())


def main():
    if not SRC.exists():
        print(f"ERROR: missing {SRC}", file=sys.stderr)
        sys.exit(2)

    print(f"Reading {SRC}")
    wb = openpyxl.load_workbook(SRC)

    stats = {
        "matched_by_theme": Counter(),
        "matched_by_node_fallback": Counter(),
        "orphan_no_node": [],
        "orphan_unmapped_theme": Counter(),
        "sat_coverage": Counter(),
        "per_sheet_total": {},
        "per_sheet_filled": {},
    }

    for sname in PRIMARY:
        ws = wb[sname]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]

        if "atomic_indicator_id" not in headers:
            new_col = ws.max_column + 1
            ws.cell(row=1, column=new_col, value="atomic_indicator_id")
            headers.append("atomic_indicator_id")

        sat_i = headers.index("atomic_indicator_id")
        tag_i = headers.index("标签ID") if "标签ID" in headers else 0
        theme_i = headers.index("标签主题") if "标签主题" in headers else None
        aipl_i = headers.index("AIPL节点") if "AIPL节点" in headers else None
        name_i = headers.index("VOC标签（中文）") if "VOC标签（中文）" in headers else 3

        total = 0
        filled = 0
        for row in ws.iter_rows(min_row=2):
            total += 1
            tag_id = row[tag_i].value if tag_i < len(row) else None
            theme = row[theme_i].value if theme_i is not None and theme_i < len(row) else None
            aipl = row[aipl_i].value if aipl_i is not None and aipl_i < len(row) else None
            name = row[name_i].value if name_i < len(row) else None

            sat = None
            if theme and str(theme).strip() in THEME_TO_SAT:
                sat = THEME_TO_SAT[str(theme).strip()]
                stats["matched_by_theme"][str(theme).strip()] += 1
            elif aipl and str(aipl).strip() in THEME_BY_NODE_FALLBACK:
                sat = THEME_BY_NODE_FALLBACK[str(aipl).strip()]
                stats["matched_by_node_fallback"][str(aipl).strip()] += 1
                if theme:
                    stats["orphan_unmapped_theme"][str(theme).strip()] += 1
            else:
                stats["orphan_no_node"].append((sname, tag_id, theme, aipl, name))

            if sat:
                row[sat_i].value = sat
                stats["sat_coverage"][sat] += 1
                filled += 1

        stats["per_sheet_total"][sname] = total
        stats["per_sheet_filled"][sname] = filled

    print(f"Writing {DST}")
    wb.save(DST)

    lines = [
        "# v4.3 → v4.4 原子指标映射 diff 报告",
        "",
        f"- 源: `{SRC.name}`",
        f"- 目标: `{DST.name}`",
        f"- 映射规则数: 标签主题 → SAT 共 {len(THEME_TO_SAT)} 条 + 4 个 AIPL 节点 fallback",
        "",
        "## 1. 各 sheet 填充率",
        "",
        "| sheet | 总标签 | 已填充 SAT | 覆盖率 |",
        "|---|---|---|---|",
    ]
    for s in PRIMARY:
        total = stats["per_sheet_total"].get(s, 0)
        filled = stats["per_sheet_filled"].get(s, 0)
        rate = filled / total * 100 if total else 0
        flag = "✅" if rate == 100 else ("⚠️" if rate >= 95 else "❌")
        lines.append(f"| {s} | {total} | {filled} | {flag} {rate:.1f}% |")

    lines += [
        "",
        "## 2. 50 SAT 覆盖统计（每个 SAT 关联多少标签）",
        "",
        "| SAT ID | 关联标签数 |",
        "|---|---|",
    ]
    for sat in sorted(VALID_SAT_IDS):
        cnt = stats["sat_coverage"].get(sat, 0)
        flag = "✅" if cnt > 0 else "❌"
        lines.append(f"| {sat} | {flag} {cnt} |")

    lines += [
        "",
        "## 3. 映射来源分布",
        "",
        f"- 通过 标签主题 直接匹配: {sum(stats['matched_by_theme'].values())}",
        f"- 通过 AIPL 节点 fallback: {sum(stats['matched_by_node_fallback'].values())}",
        f"- 完全无法映射 (无主题且无节点): {len(stats['orphan_no_node'])}",
        "",
        "## 4. AIPL fallback 命中的标签主题 (这些主题未在 THEME_TO_SAT 中显式定义)",
        "",
        "| 标签主题 | 次数 | 建议 |",
        "|---|---|---|",
    ]
    for t, c in sorted(stats["orphan_unmapped_theme"].items(), key=lambda x: -x[1]):
        lines.append(f"| {t!r} | {c} | 考虑在 THEME_TO_SAT 中显式映射 |")

    if stats["orphan_no_node"]:
        lines += [
            "",
            "## 5. ❌ 无法映射的孤立标签",
            "",
            "| sheet | tag_id | 标签主题 | AIPL | tag_cn |",
            "|---|---|---|---|---|",
        ]
        for s, tid, th, ap, nm in stats["orphan_no_node"]:
            lines.append(f"| {s} | {tid} | {th} | {ap} | {nm} |")

    DIFF.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"Diff -> {DIFF}")
    print()
    print(f"Total tags scanned:          {sum(stats['per_sheet_total'].values())}")
    print(f"Total tags with SAT filled:  {sum(stats['per_sheet_filled'].values())}")
    print(f"Coverage by theme:           {sum(stats['matched_by_theme'].values())}")
    print(f"Coverage by AIPL fallback:   {sum(stats['matched_by_node_fallback'].values())}")
    print(f"Orphan (no theme/node):      {len(stats['orphan_no_node'])}")


if __name__ == "__main__":
    main()
