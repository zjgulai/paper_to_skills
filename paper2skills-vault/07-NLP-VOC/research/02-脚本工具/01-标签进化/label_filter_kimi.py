"""Phase 6 D9 Method C — High-Risk Tag Filter (Post-processing with Kimi)

D7/D8 暴露：D5 lenient 数据 precision 0.64，9 个高风险 tag 系统性误判。
D8 重打 strict prompt 把 precision 拉到 0.89 但召回掉了 7pp 导致 Gate 5/7。

Method C 折中方案：
  - 保留 D5 lenient 数据的高召回（Gate 7/7 PASS）
  - 仅对 9 个高风险 tag 用 Kimi 单独验证，删掉判 reject 的
  - 其余 70+ tag 不动（D7 spot check 显示这些 tag 精度尚可）

效果预期：
  - Gate 维持 7/7 PASS（删除少量误判 tag，覆盖率波动 < 1pp）
  - 9 个高风险 tag 的 precision 跨 0.85
  - 整体 precision 从 0.64 → ~0.85+

输入：phase6_d5_final.jsonl (Phase 5 收官版数据)
输出：phase6_d9_filtered.jsonl

设计：
  - 流式遍历 D5 final
  - 仅对含 ≥1 高风险标签 的 record 调 Kimi
  - 把多个高风险 tag 一次性 batched 给 Kimi 判定
  - 接受 → 保留；拒绝 → 删除
  - 非高风险标签一律保留（避免触动 D5 既有正确标签）

成本：
  - 29,761 records / 10 records-per-call ≈ 2,976 calls
  - 总 high-risk labels = 33,953（Kimi 一次判定 1 review × N tags）
  - Kimi tokens_in ~500/call × 2976 = ~1.5M
  - 估算 ~$0.5
  - concurrency=10, ~0.6s/call → ~30 min

用法：
  python label_filter_kimi.py \
    --input <vault>/04-输出结果/unified_labeling/phase6_d5_final.jsonl \
    --dict <vault>/04-输出结果/01-字典版本/tag_dictionary_v4.1.xlsx \
    --output <vault>/04-输出结果/unified_labeling/phase6_d9_filtered.jsonl \
    --report <vault>/04-输出结果/03-审计报告/phase6_d9_filter.md
"""

from __future__ import annotations

import argparse
import asyncio
import json
import sys
import time
from collections import Counter
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "07-LLM引擎"))


HIGH_RISK_TAGS = {
    "TAG_P1_001",  # 核心卖点清晰
    "TAG_P1_009",  # 物超所值
    "TAG_P1_010",  # 性价比差
    "TAG_I_001",   # 信息难找
    "TAG_L1_074",  # 退货-不符预期
    "TAG_L2_002",  # 一次解决问题
    "TAG_L2_005",  # 会再次购买
    "TAG_L2_006",  # 不会再次购买
    "TAG_P2_001",  # 错件/漏件/多件
}

CONCURRENCY = 10
RECORDS_PER_BATCH = 10

JUDGE_PROMPT = """You are a strict VOC tag verifier. For each given (review_text, candidate_tag) pair, decide if the tag is precisely supported by the review.

STRICT TAG RULES:
- TAG_P1_001 核心卖点清晰: ONLY if review explicitly mentions the brand's CORE selling point or differentiating feature (e.g. "wearable design unique", "the only pump that..."). Generic praise like "great product" or "love it" does NOT qualify.
- TAG_P1_009 物超所值: REQUIRES explicit price/value mention (e.g. "$30 worth it", "性价比高", "vaut son prix"). Rating 5 alone does NOT qualify.
- TAG_P1_010 性价比差: REQUIRES explicit price-related complaint. Generic dissatisfaction does NOT qualify.
- TAG_I_001 信息难找: REQUIRES explicit complaint about finding info (specs/order details). Shipping/missing items do NOT qualify.
- TAG_L1_074 退货-不符预期: REQUIRES actual return action AND expectations not met. Pure dissatisfaction does NOT qualify.
- TAG_L2_002 一次解决问题: REQUIRES issue raised AND fully resolved. Quick CS reply alone does NOT qualify.
- TAG_L2_005 会再次购买: REQUIRES explicit repurchase intent (e.g. "will buy again"). Generic praise does NOT qualify.
- TAG_L2_006 不会再次购买: REQUIRES explicit non-repurchase statement. Negative review alone does NOT qualify.
- TAG_P2_001 错件/漏件/多件: REQUIRES explicit wrong/missing/extra physical components. Order delays do NOT qualify.

Be strict: prefer false negative (rejecting weak match) over false positive.

Output JSON: {"verdicts": [{"review_id": "...", "tag_id": "...", "keep": true/false}, ...]}"""


@dataclass
class FilterStats:
    n_input: int = 0
    n_records_no_hr: int = 0
    n_records_with_hr: int = 0
    n_high_risk_total: int = 0
    n_kept: int = 0
    n_dropped: int = 0
    n_judge_failed: int = 0
    drop_per_tag: Counter = field(default_factory=Counter)
    kept_per_tag: Counter = field(default_factory=Counter)
    elapsed_s: float = 0.0


def load_dict_meta(dict_path: Path) -> dict[str, dict[str, Any]]:
    wb = openpyxl.load_workbook(str(dict_path), read_only=True, data_only=True)
    ws = wb["01_通用标签主表"]
    hdr = [c.value for c in next(ws.iter_rows(max_row=1))]
    ix = {k: hdr.index(k) for k in [
        "标签ID", "VOC标签（中文）", "VOC标签（英文）", "标签定义",
    ]}
    out: dict[str, dict[str, Any]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        tid = row[ix["标签ID"]]
        if not tid:
            continue
        out[str(tid)] = {
            "tag_en": str(row[ix["VOC标签（英文）"]] or ""),
            "tag_cn": str(row[ix["VOC标签（中文）"]] or ""),
        }
    return out


def collect_high_risk_jobs(
    record: dict[str, Any],
) -> list[tuple[str, dict[str, Any]]]:
    """Return list of (tag_id, label_dict) where tag is high-risk."""
    jobs = []
    for lbl in record.get("labels") or []:
        if not isinstance(lbl, dict):
            continue
        tid = lbl.get("tag_id")
        if tid in HIGH_RISK_TAGS:
            jobs.append((tid, lbl))
    return jobs


async def judge_batch(
    client,
    records_with_hr: list[tuple[dict[str, Any], list[dict[str, Any]]]],
    dict_meta: dict[str, dict[str, Any]],
    stats: FilterStats,
) -> dict[str, set[str]]:
    """Batch-judge multiple records in one Kimi call.

    Returns dict[review_id → set of tag_ids to KEEP].
    """
    review_blocks = []
    for record, hr_labels in records_with_hr:
        text = (record.get("text") or "").replace("\n", " ").strip()
        if len(text) > 400:
            text = text[:400] + "..."
        rid = record["review_id"]
        tag_lines = []
        for lbl in hr_labels:
            tid = lbl["tag_id"]
            meta = dict_meta.get(tid, {})
            tag_lines.append(f"  - {tid}\t{meta.get('tag_en','')}\t{meta.get('tag_cn','')}")
        review_blocks.append(
            f"<review id=\"{rid}\">\n{text}\n  candidates:\n{chr(10).join(tag_lines)}\n</review>"
        )

    user_prompt = f"""For each review below, verify each candidate tag against the STRICT TAG RULES.

{chr(10).join(review_blocks)}

Output JSON: {{"verdicts": [{{"review_id": "...", "tag_id": "...", "keep": true/false}}, ...]}}
Include ONE entry per (review_id, tag_id) pair."""

    keep_map: dict[str, set[str]] = {r["review_id"]: set() for r, _ in records_with_hr}
    try:
        resp = await client.chat_async(
            vendor="kimi",
            messages=[
                {"role": "system", "content": JUDGE_PROMPT},
                {"role": "user", "content": user_prompt},
            ],
            response_format={"type": "json_object"},
            max_tokens=1200,
            temperature=0.1,
        )
        if not resp.content:
            stats.n_judge_failed += len(records_with_hr)
            for record, hr_labels in records_with_hr:
                keep_map[record["review_id"]] = {lbl["tag_id"] for lbl in hr_labels}
            return keep_map
        data = json.loads(resp.content)
        verdicts = data.get("verdicts") or []
        for v in verdicts:
            if not isinstance(v, dict):
                continue
            rid = v.get("review_id")
            tid = v.get("tag_id")
            keep = v.get("keep")
            if rid in keep_map and tid in HIGH_RISK_TAGS and isinstance(keep, bool) and keep:
                keep_map[rid].add(tid)
        return keep_map
    except Exception:
        stats.n_judge_failed += len(records_with_hr)
        for record, hr_labels in records_with_hr:
            keep_map[record["review_id"]] = {lbl["tag_id"] for lbl in hr_labels}
        return keep_map


async def judge_record(
    client,
    record: dict[str, Any],
    high_risk_labels: list[dict[str, Any]],
    dict_meta: dict[str, dict[str, Any]],
    stats: FilterStats,
) -> set[str]:
    """Returns set of tag_ids to KEEP (drop = high_risk - keep)."""
    keep_map = await judge_batch(client, [(record, high_risk_labels)], dict_meta, stats)
    return keep_map.get(record["review_id"], set())


def filter_record(
    record: dict[str, Any],
    keep_set: set[str],
    high_risk_label_ids: set[str],
    stats: FilterStats,
) -> dict[str, Any]:
    new_labels: list[dict[str, Any]] = []
    for lbl in record.get("labels") or []:
        if not isinstance(lbl, dict):
            new_labels.append(lbl)
            continue
        tid = lbl.get("tag_id")
        if tid not in HIGH_RISK_TAGS:
            new_labels.append(lbl)
            continue
        if tid in keep_set:
            stats.n_kept += 1
            stats.kept_per_tag[tid] += 1
            new_labels.append(lbl)
        else:
            stats.n_dropped += 1
            stats.drop_per_tag[tid] += 1
    out = dict(record)
    out["labels"] = new_labels
    out["n_tags"] = len(new_labels)
    if record.get("n_tags") != len(new_labels):
        out["_phase6_d9_filter"] = {
            "original_n_tags": record.get("n_tags"),
            "filtered_n_tags": len(new_labels),
            "high_risk_dropped": sorted(high_risk_label_ids - keep_set),
        }
    return out


async def run_pipeline(
    input_path: Path,
    output_path: Path,
    dict_meta: dict[str, dict[str, Any]],
) -> FilterStats:
    from llm_client import LLMClient
    client = LLMClient()

    stats = FilterStats()
    sem = asyncio.Semaphore(CONCURRENCY)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    fout = output_path.open("w", encoding="utf-8")
    out_lock = asyncio.Lock()
    progress = {"verified_records": 0, "passed_through": 0, "batches_done": 0}
    t0 = time.time()

    async def write_filtered(record: dict[str, Any], keep_set: set[str], hr_label_ids: set[str]) -> None:
        out_record = filter_record(record, keep_set, hr_label_ids, stats)
        async with out_lock:
            fout.write(json.dumps(out_record, ensure_ascii=False) + "\n")

    async def process_batch(batch: list[tuple[dict[str, Any], list[dict[str, Any]]]]) -> None:
        async with sem:
            keep_map = await judge_batch(client, batch, dict_meta, stats)
            progress["batches_done"] += 1
            progress["verified_records"] += len(batch)
        for record, hr_labels in batch:
            keep_set = keep_map.get(record["review_id"], set())
            hr_label_ids = {lbl["tag_id"] for lbl in hr_labels}
            await write_filtered(record, keep_set, hr_label_ids)

        if progress["batches_done"] % 50 == 0:
            elapsed = time.time() - t0
            rate = progress["verified_records"] / max(elapsed, 1)
            est_total = 29761
            eta = (est_total - progress["verified_records"]) / max(rate, 1)
            print(
                f"  batches={progress['batches_done']} verified={progress['verified_records']:,}/{est_total:,} "
                f"kept={stats.n_kept} dropped={stats.n_dropped} "
                f"rate={rate:.1f}rec/s eta={eta:.0f}s",
                file=sys.stderr,
            )

    print(f"⏳ Streaming D5 final, batching {RECORDS_PER_BATCH} records/Kimi call", file=sys.stderr)
    pending_batch: list[tuple[dict[str, Any], list[dict[str, Any]]]] = []
    pending_tasks: list[asyncio.Task] = []

    with input_path.open("r", encoding="utf-8") as fin:
        for line in fin:
            line = line.strip()
            if not line:
                continue
            try:
                record = json.loads(line)
            except json.JSONDecodeError:
                continue
            stats.n_input += 1
            hr_jobs = collect_high_risk_jobs(record)
            if not hr_jobs:
                stats.n_records_no_hr += 1
                progress["passed_through"] += 1
                async with out_lock:
                    fout.write(json.dumps(record, ensure_ascii=False) + "\n")
                continue
            stats.n_records_with_hr += 1
            stats.n_high_risk_total += len(hr_jobs)
            pending_batch.append((record, [lbl for _, lbl in hr_jobs]))
            if len(pending_batch) >= RECORDS_PER_BATCH:
                pending_tasks.append(asyncio.create_task(process_batch(pending_batch)))
                pending_batch = []

    if pending_batch:
        pending_tasks.append(asyncio.create_task(process_batch(pending_batch)))

    print(f"⏳ {len(pending_tasks)} batches dispatched, awaiting completion", file=sys.stderr)
    await asyncio.gather(*pending_tasks)
    fout.close()
    stats.elapsed_s = time.time() - t0
    return stats


def render_report(stats: FilterStats, args: argparse.Namespace) -> str:
    md = []
    p = md.append
    p("---")
    p("name: phase6-d9-filter")
    p("description: Phase 6 D9 Method C 高风险标签 Kimi 后处理过滤报告。"
      "保留 D5 lenient 数据的高召回，仅对 9 个高风险 tag 用 Kimi 验证删除误判，"
      "目标精度 + 召回双过 Gate。")
    p(f"date: {datetime.now().strftime('%Y-%m-%d')}")
    p("phase: phase6")
    p("day: D9")
    p("doc_type: audit-report")
    p("module: voc-nlp")
    p("---")
    p("")
    p("# Phase 6 D9 Method C 高风险标签 Kimi 过滤报告")
    p("")
    p(f"- 输入：`{args.input}`")
    p(f"- 输出：`{args.output}`")
    p(f"- 时间：{datetime.now().isoformat(timespec='seconds')}")
    p("")
    p("## 一、整体效果")
    p("")
    drop_rate = stats.n_dropped / max(stats.n_high_risk_total, 1)
    p("| 指标 | 值 |")
    p("|---|---:|")
    p(f"| 总记录 | {stats.n_input:,} |")
    p(f"| 含高风险 tag 的记录 | {stats.n_records_with_hr:,} ({100*stats.n_records_with_hr/max(stats.n_input,1):.2f}%) |")
    p(f"| 总高风险 tag 数 | {stats.n_high_risk_total:,} |")
    p(f"| 保留 (Kimi accept) | **{stats.n_kept:,}** ({100*stats.n_kept/max(stats.n_high_risk_total,1):.1f}%) |")
    p(f"| 删除 (Kimi reject) | **{stats.n_dropped:,}** ({100*drop_rate:.1f}%) |")
    p(f"| Kimi 判官失败 | {stats.n_judge_failed:,} |")
    p(f"| 总耗时 | {stats.elapsed_s:.1f}s |")
    p("")
    p("## 二、Per-tag 删除/保留")
    p("")
    p("| tag_id | tag_cn | 原数 | kept | dropped | drop% |")
    p("|---|---|---:|---:|---:|---:|")
    all_tags = sorted(set(stats.kept_per_tag) | set(stats.drop_per_tag))
    for tid in all_tags:
        kept = stats.kept_per_tag.get(tid, 0)
        dropped = stats.drop_per_tag.get(tid, 0)
        total = kept + dropped
        drop_pct = 100 * dropped / max(total, 1)
        p(f"| {tid} | (见字典) | {total} | {kept} | {dropped} | {drop_pct:.1f}% |")
    p("")
    return "\n".join(md)


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(description="Phase 6 D9 Method C 高风险标签 Kimi 过滤")
    ap.add_argument("--input", required=True, type=Path)
    ap.add_argument("--dict", required=True, type=Path)
    ap.add_argument("--output", required=True, type=Path)
    ap.add_argument("--report", type=Path, default=None)
    args = ap.parse_args(argv)

    if not args.input.is_file():
        print(f"❌ input not found: {args.input}", file=sys.stderr); return 2
    if not args.dict.is_file():
        print(f"❌ dict not found: {args.dict}", file=sys.stderr); return 2

    print(f"⏳ Loading dict meta", file=sys.stderr)
    dict_meta = load_dict_meta(args.dict)
    print(f"   {len(dict_meta)} tags loaded", file=sys.stderr)

    stats = asyncio.run(run_pipeline(args.input, args.output, dict_meta))

    print(f"\n✅ DONE", file=sys.stderr)
    print(f"   records={stats.n_input:,} (no_hr={stats.n_records_no_hr:,}, with_hr={stats.n_records_with_hr:,})", file=sys.stderr)
    print(f"   high_risk_tags={stats.n_high_risk_total:,} kept={stats.n_kept:,} dropped={stats.n_dropped:,}", file=sys.stderr)
    print(f"   judge_failed={stats.n_judge_failed:,}", file=sys.stderr)
    print(f"   elapsed={stats.elapsed_s:.1f}s", file=sys.stderr)
    print(f"📄 Output: {args.output}", file=sys.stderr)

    if args.report:
        args.report.parent.mkdir(parents=True, exist_ok=True)
        args.report.write_text(render_report(stats, args), encoding="utf-8")
        print(f"📄 Report: {args.report}", file=sys.stderr)

    return 0


if __name__ == "__main__":
    sys.exit(main())
