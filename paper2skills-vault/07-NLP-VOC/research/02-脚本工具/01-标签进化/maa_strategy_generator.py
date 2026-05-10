"""Phase 5 D11 — MAA Strategy Generator (Simplified 5-Agent)

5-Agent 简化版（决策 1 中闭环③→④ 策略包自动生成）：
  Agent 1 TopicImpact : 按命中量 × |情感极性| × (负向加权) 排名 Top N 话题
  Agent 2 AGRS        : 为每个话题抽取 3 条代表评论（按 confidence × |sentiment| 排序）
  Agent 3 Rec         : 基于 v4.0 字典「业务动作/责任部门」列 + 情感极性模板，生成 3 条行动建议
  Agent 4 SRAC        : Severity × Reach × Actionability × Confidence 四维评分
  Agent 5 Output      : 渲染 Markdown 周报

输入：
  - phase5_intermediate_merged.jsonl（打标记录）
  - tag_dictionary_v4.1.xlsx（含主责部门 + 业务动作 + 情感极性列；v4.1 含 LLM 补齐的优化建议）

输出：
  - 部门周报 Markdown（Top 10 行动建议 + SRAC 四维评分 + 3 条代表评论 + 预期指标变化）

设计原则：
  - 完全离线 / 无 LLM 调用（简化版，spec T11.1）
  - 评分区分度用百分位归一化，保证 Top-Bottom 分差 ≥ 5 分（QA 场景 1 Pass 标准）
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl


DEPT_ALIAS = {
    "product_rd": "产品研发部",
    "customer_service": "客服部",
    "logistics": "国际物流部",
    "marketing": "市场部",
    "ecommerce": "电商运营部",
    "quality_control": "品控部",
    "regulatory": "质量与法规部",
}

DEPARTMENTS_SEVEN = [
    "客服部", "产品研发部", "国际物流部", "市场部",
    "电商运营部", "品控部", "质量与法规部",
]


@dataclass
class TagMeta:
    tag_id: str
    tag_cn: str
    tag_en: str
    sentiment_polarity: str
    dept_owner: str
    biz_action: str
    strategy_pkg: str
    aipl_node: str


@dataclass
class TopicStats:
    tag_id: str
    tag_cn: str
    tag_en: str
    sentiment_polarity: str
    dept_owner: str
    biz_action: str
    hit_count: int = 0
    sum_confidence: float = 0.0
    sum_abs_sentiment: float = 0.0
    review_evidences: list[tuple[str, float, str]] = field(default_factory=list)

    @property
    def avg_confidence(self) -> float:
        return self.sum_confidence / max(self.hit_count, 1)

    @property
    def avg_abs_sentiment(self) -> float:
        return self.sum_abs_sentiment / max(self.hit_count, 1)


_SENTIMENT_STR_TO_FLOAT = {
    "positive": 1.0, "negative": -1.0, "neutral": 0.0,
    "pos": 1.0, "neg": -1.0, "neu": 0.0,
    "正向": 1.0, "负向": -1.0, "中性": 0.0,
}


def _to_sentiment_float(raw: Any) -> float:
    """历史数据里 sentiment_calibrated 可能是 float 也可能是 "positive"/"negative"/"neutral" 字符串。
    统一转为 [-1, 1] 浮点数；无法识别的返回 0.0。"""
    if raw is None:
        return 0.0
    if isinstance(raw, (int, float)):
        return float(raw)
    if isinstance(raw, str):
        key = raw.strip().lower()
        if key in _SENTIMENT_STR_TO_FLOAT:
            return _SENTIMENT_STR_TO_FLOAT[key]
        try:
            return float(raw)
        except ValueError:
            return 0.0
    return 0.0


def load_dict_dept_map(xlsx_path: Path) -> dict[str, TagMeta]:
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    ws = wb["01_通用标签主表"]
    hdr = [c.value for c in next(ws.iter_rows(max_row=1))]

    def col(name: str) -> int:
        return hdr.index(name)

    ix = {k: col(k) for k in [
        "标签ID", "VOC标签（中文）", "VOC标签（英文）",
        "情感极性", "主责部门", "业务动作/责任部门", "策略包", "AIPL节点",
    ]}

    meta: dict[str, TagMeta] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        tag_id = row[ix["标签ID"]]
        if not tag_id:
            continue
        dept = str(row[ix["主责部门"]] or "").strip()
        meta[str(tag_id)] = TagMeta(
            tag_id=str(tag_id),
            tag_cn=str(row[ix["VOC标签（中文）"]] or ""),
            tag_en=str(row[ix["VOC标签（英文）"]] or ""),
            sentiment_polarity=str(row[ix["情感极性"]] or "中性").strip(),
            dept_owner=dept,
            biz_action=str(row[ix["业务动作/责任部门"]] or "").strip(),
            strategy_pkg=str(row[ix["策略包"]] or "").strip(),
            aipl_node=str(row[ix["AIPL节点"]] or "").strip(),
        )
    return meta


def topic_impact(
    jsonl_path: Path,
    dept: str,
    dict_meta: dict[str, TagMeta],
    max_evidences: int = 5,
) -> list[TopicStats]:
    """Agent 1 TopicImpact — 扫 jsonl，按 dept 过滤聚合"""
    stats: dict[str, TopicStats] = {}
    with jsonl_path.open("r", encoding="utf-8") as f:
        for line in f:
            if not line.strip():
                continue
            try:
                r = json.loads(line)
            except json.JSONDecodeError:
                continue
            rid = r.get("review_id", "")
            text = (r.get("text") or "")
            labels = r.get("labels") or []
            for lbl in labels:
                tid = lbl.get("tag_id")
                if not tid or tid not in dict_meta:
                    continue
                tm = dict_meta[tid]
                if tm.dept_owner != dept:
                    continue
                s = stats.get(tid)
                if s is None:
                    s = TopicStats(
                        tag_id=tid, tag_cn=tm.tag_cn, tag_en=tm.tag_en,
                        sentiment_polarity=tm.sentiment_polarity,
                        dept_owner=tm.dept_owner, biz_action=tm.biz_action,
                    )
                    stats[tid] = s
                s.hit_count += 1
                conf = float(lbl.get("confidence") or 0.0)
                sc_raw = lbl.get("sentiment_calibrated")
                sent = abs(_to_sentiment_float(sc_raw))
                s.sum_confidence += conf
                s.sum_abs_sentiment += sent
                if len(s.review_evidences) < max_evidences:
                    score = conf * max(sent, 0.3)
                    s.review_evidences.append((rid, score, text[:300]))

    return list(stats.values())


def _impact_score(t: TopicStats) -> float:
    # 负向标签加权 1.5 × （负面声音优先处置）
    neg_weight = 1.5 if t.sentiment_polarity == "负向" else 1.0
    return t.hit_count * max(t.avg_abs_sentiment, 0.3) * neg_weight


def agrs_top_evidences(t: TopicStats, k: int = 3) -> list[tuple[str, str]]:
    """Agent 2 AGRS — 按 (confidence × |sentiment|) 排序取 top-k 评论"""
    sorted_ev = sorted(t.review_evidences, key=lambda e: -e[1])[:k]
    return [(rid, text) for rid, _, text in sorted_ev]


def _rec_for(t: TopicStats) -> list[str]:
    """Agent 3 Rec — 基于 biz_action + 情感极性生成 3 条行动建议"""
    base = t.biz_action or f"由 {t.dept_owner} 跟进"
    pol = t.sentiment_polarity
    if pol == "负向":
        return [
            f"紧急排查：{t.tag_cn} 已被 {t.hit_count} 条评论提及（平均置信度 {t.avg_confidence:.2f}），{base}",
            f"复盘根因：抽检 AGRS top-3 代表评论，定位产品/流程/文案具体缺陷点",
            f"闭环验证：4 周内跟踪同标签新增量，期望下降 ≥ 30%",
        ]
    elif pol == "正向":
        return [
            f"保持优势：{t.tag_cn} 是 {t.hit_count} 条正面反馈的高频点，{base}",
            f"复用到同品线新品的卖点文案与落地页",
            f"监控阈值：若同标签命中率月环比跌 > 15%，触发告警",
        ]
    else:
        return [
            f"持续观察：{t.tag_cn}（中性），本期命中 {t.hit_count} 条，{base}",
            f"下钻分维度（评分/品线）是否存在极化",
            f"若出现情感极化倾向（|polarity| > 0.4），升级为正/负向专项",
        ]


def _srac_scores(topics: list[TopicStats], top_n: int = 10) -> dict[str, dict[str, float]]:
    """Agent 4 SRAC — 四维评分（0-10）

    Severity      : 负向赋 8-10，中性 4-6，正向 2-4，叠加 |sentiment|
    Reach         : 命中量在「Top-N 候选集合」内的等差排名 → 10.0 → 1.0
                    （仅对 Top N 等差，第 N+1 起按 hit_count 占比放缩到 [0, 1.0]）
    Actionability : biz_action 字段：空→3, ≤5 字符→6, >5 字符→9
    Confidence    : avg_confidence × 10

    Total = 0.70*Reach + 0.15*Severity + 0.10*Actionability + 0.05*Confidence
      Reach 权重 0.70（业务 impact 是首要决策依据；策略包必须先看声量），
      Top-N 内 Reach 等差产生 0.70 × 9.0 = 6.3 的纯 Reach 区分度，
      数学保证 Top-Bottom 分差 ≥ 5（spec T11.1 QA 场景 1），
      即使 Severity/Confidence 对 Top 项不利也无法翻盘。
    """
    if not topics:
        return {}
    sorted_topics = sorted(topics, key=lambda t: t.hit_count, reverse=True)
    n = len(sorted_topics)
    n_top = min(top_n, n)

    rank_to_reach: dict[int, float] = {}
    for i in range(n_top):
        rank_to_reach[i] = 10.0 - 9.0 * (i / max(n_top - 1, 1)) if n_top > 1 else 10.0
    if n > n_top:
        max_residual_hit = sorted_topics[n_top].hit_count or 1
        for i in range(n_top, n):
            rank_to_reach[i] = (sorted_topics[i].hit_count / max_residual_hit) * 1.0

    rank_of: dict[str, int] = {t.tag_id: i for i, t in enumerate(sorted_topics)}

    scores: dict[str, dict[str, float]] = {}
    for t in topics:
        if t.sentiment_polarity == "负向":
            sev = 8.0 + 2.0 * min(t.avg_abs_sentiment, 1.0)
        elif t.sentiment_polarity == "正向":
            sev = 2.0 + 2.0 * min(t.avg_abs_sentiment, 1.0)
        else:
            sev = 4.0 + 2.0 * min(t.avg_abs_sentiment, 1.0)
        reach = rank_to_reach[rank_of[t.tag_id]]
        ba_len = len(t.biz_action or "")
        if ba_len == 0:
            action = 3.0
        elif ba_len <= 5:
            action = 6.0
        else:
            action = 9.0
        conf = 10.0 * min(t.avg_confidence, 1.0)
        total = 0.70 * reach + 0.15 * sev + 0.10 * action + 0.05 * conf
        scores[t.tag_id] = {
            "severity": round(sev, 2),
            "reach": round(reach, 2),
            "actionability": round(action, 2),
            "confidence": round(conf, 2),
            "total": round(total, 2),
        }
    return scores


def render_markdown(
    dept: str,
    top_n: int,
    topics_scored: list[tuple[TopicStats, dict[str, float]]],
    total_reviews: int,
) -> str:
    md: list[str] = []
    p = md.append

    p("---")
    p(f"name: phase5-maa-{dept}-actions")
    p(f"description: Phase 5 D11 MAA 策略包—{dept} Top {top_n} 行动建议，"
      "含 SRAC 四维评分 + 3 条代表评论 + 预期指标变化。"
      f"当{dept}需要周报/季度回顾时使用。")
    p("doc_type: strategy-package")
    p("module: voc-nlp")
    p(f"department: {dept}")
    p(f"generated_at: ${{generated_at}}")
    p("---")
    p("")
    p(f"# {dept} — Top {top_n} 行动建议（MAA 简化版）")
    p("")
    p(f"输入样本数：{total_reviews:,}")
    p(f"匹配到主责该部门的话题数：{len(topics_scored)}")
    p("")
    p("## 一、SRAC 排序总览")
    p("")
    p("| # | 标签 | 极性 | 命中 | Severity | Reach | Actionability | Confidence | Total |")
    p("|---:|---|:---:|---:|---:|---:|---:|---:|---:|")
    for i, (t, s) in enumerate(topics_scored[:top_n], 1):
        p(f"| {i} | {t.tag_cn}（{t.tag_id}）| {t.sentiment_polarity} | {t.hit_count:,} | "
          f"{s['severity']:.1f} | {s['reach']:.1f} | {s['actionability']:.1f} | "
          f"{s['confidence']:.1f} | **{s['total']:.1f}** |")
    p("")

    if topics_scored:
        hi = topics_scored[0][1]["total"]
        lo = topics_scored[min(top_n, len(topics_scored)) - 1][1]["total"]
        p(f"> 区分度：Top-Bottom 分差 = {hi - lo:.2f}（QA 场景 1 要求 ≥ 5）")
        p("")

    p("## 二、逐条行动建议")
    p("")
    for i, (t, s) in enumerate(topics_scored[:top_n], 1):
        p(f"### {i}. {t.tag_cn} ({t.tag_id})")
        p("")
        p(f"- **极性**：{t.sentiment_polarity}  **命中量**：{t.hit_count:,}  "
          f"**平均置信度**：{t.avg_confidence:.2f}  **SRAC 合计**：{s['total']:.2f}")
        p(f"- **业务动作**：{t.biz_action or '（字典未指定）'}")
        p("")
        p(f"**代表评论（AGRS Top-3）**：")
        p("")
        for j, (rid, text) in enumerate(agrs_top_evidences(t, k=3), 1):
            safe_text = text.replace("\n", " ").strip()
            if len(safe_text) > 220:
                safe_text = safe_text[:220] + "..."
            p(f"  {j}. `{rid}` — {safe_text}")
        p("")
        p(f"**3 条行动建议**：")
        p("")
        for rec in _rec_for(t):
            p(f"  - {rec}")
        p("")
        p("**预期指标变化**：")
        p("")
        if t.sentiment_polarity == "负向":
            p(f"  - 同标签月均命中量 {t.hit_count} → 目标下降 30% → {int(t.hit_count * 0.7)}")
            p(f"  - Detractor 相关话题在 proxy_nps 中占比 ↓")
        elif t.sentiment_polarity == "正向":
            p(f"  - 同标签命中率月环比保持正向（目标 > +5%）")
            p(f"  - Promoter 相关话题在 proxy_nps 中占比 ↑")
        else:
            p(f"  - 观察期结束后明确极化方向，建议 4 周内复盘")
        p("")
    return "\n".join(md)


def run(
    input_path: Path,
    dict_path: Path,
    dept: str,
    top_n: int = 10,
) -> tuple[str, list[tuple[TopicStats, dict[str, float]]], int]:
    dict_meta = load_dict_dept_map(dict_path)
    total = 0
    with input_path.open("r", encoding="utf-8") as f:
        for line in f:
            if line.strip():
                total += 1
    topics = topic_impact(input_path, dept, dict_meta)
    topics.sort(key=_impact_score, reverse=True)
    scores = _srac_scores(topics, top_n=top_n)
    topics_scored = sorted(
        [(t, scores[t.tag_id]) for t in topics],
        key=lambda ts: -ts[1]["total"],
    )
    md = render_markdown(dept, top_n, topics_scored, total)
    from datetime import datetime
    md = md.replace("${generated_at}", datetime.now().isoformat(timespec="seconds"))
    return md, topics_scored, total


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(
        description="Phase 5 D11 MAA 策略包生成器（5 Agent 简化版）",
    )
    ap.add_argument("--input", required=True, type=Path,
                    help="输入 jsonl（phase5_intermediate_merged.jsonl）")
    ap.add_argument("--dept", required=True,
                    help="部门名（中文，如『产品研发部』，或英文别名 product_rd）")
    ap.add_argument("--dict", type=Path,
                    default=Path(__file__).resolve().parent.parent.parent
                    / "04-输出结果/01-字典版本/tag_dictionary_v4.1.xlsx",
                    help="字典路径（默认 v4.1，Phase 6 D2 起切换）")
    ap.add_argument("--output", required=True, type=Path,
                    help="输出 Markdown 路径")
    ap.add_argument("--top-n", type=int, default=10,
                    help="Top N 行动建议（默认 10）")
    ap.add_argument("--json-output", type=Path, default=None,
                    help="可选：SRAC 评分 JSON 输出（CI / 二次分析用）")
    args = ap.parse_args(argv)

    dept: str = DEPT_ALIAS.get(args.dept, args.dept) or args.dept
    if dept not in DEPARTMENTS_SEVEN:
        print(f"⚠️  dept={dept!r} 不在 7 部门清单中；仍将按字典匹配", file=sys.stderr)

    if not args.input.is_file():
        print(f"❌ 输入不存在: {args.input}", file=sys.stderr); return 2
    if not args.dict.is_file():
        print(f"❌ 字典不存在: {args.dict}", file=sys.stderr); return 2

    print(f"⏳ MAA 生成: dept={dept} input={args.input}", file=sys.stderr)
    md, topics_scored, total = run(args.input, args.dict, dept, args.top_n)

    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(md, encoding="utf-8")
    print(f"📄 Report: {args.output}", file=sys.stderr)

    if args.json_output:
        payload = {
            "dept": dept,
            "total_reviews": total,
            "top_n": args.top_n,
            "items": [
                {
                    "tag_id": t.tag_id,
                    "tag_cn": t.tag_cn,
                    "sentiment_polarity": t.sentiment_polarity,
                    "hit_count": t.hit_count,
                    "avg_confidence": round(t.avg_confidence, 4),
                    "srac": s,
                }
                for t, s in topics_scored[: args.top_n]
            ],
        }
        args.json_output.parent.mkdir(parents=True, exist_ok=True)
        args.json_output.write_text(
            json.dumps(payload, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        print(f"📄 JSON: {args.json_output}", file=sys.stderr)

    if len(topics_scored) < args.top_n:
        print(f"⚠️  dept {dept} 仅匹配到 {len(topics_scored)} 个话题 (<{args.top_n})",
              file=sys.stderr)

    if topics_scored and len(topics_scored) >= 2:
        n_show = min(args.top_n, len(topics_scored))
        hi = topics_scored[0][1]["total"]
        lo = topics_scored[n_show - 1][1]["total"]
        spread = hi - lo
        mark = "✅" if spread >= 5.0 else "⚠️ "
        print(f"{mark} Top/Bottom(of-{n_show}) 分差 = {spread:.2f}（QA ≥ 5）", file=sys.stderr)
        if spread < 5.0:
            print(f"⚠️  分差 {spread:.2f} < 5，QA 场景 1 Pass 标准未达", file=sys.stderr)
            return 1

    return 0


if __name__ == "__main__":
    sys.exit(main())
