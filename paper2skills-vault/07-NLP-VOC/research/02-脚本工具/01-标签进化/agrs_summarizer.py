"""Phase 5 D11 — AGRS (Aspect-Guided Review Summarization)

按 aspect 或 tag 分组，对每组评论做抽取式摘要，供 BI 看板「评论快讯」模块使用。

输入：
  phase5_intermediate_merged.jsonl
  --group-by tag|aspect|aipl
  --filter-dept <部门名>（可选，限定主责部门 — 配合 v4.0 字典）

输出：
  Markdown 报告：每组 → 标题 / 关键统计 / Top-3 代表句（含情感）/ 1 句聚合摘要
  JSON：结构化数据，便于 BI 二次消费

设计：
  - 完全离线（简化版）：用句子层规则（情感强度 × 命中关键词数 × 长度归一）抽 top
  - 聚合摘要 = 用 Top-3 句子拼接 + 模板包装（"用户主要反馈... 同时... 但..."）
  - 与 maa_strategy_generator 互补：MAA 给"做什么"，AGRS 给"用户怎么说"
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl


SENTENCE_SPLIT_RE = re.compile(r"(?<=[.!?。！？])\s+|\n+")
MIN_SENT_LEN = 12
MAX_SENT_LEN = 240
TOP_K_SENTENCES = 3


@dataclass
class SentenceCandidate:
    review_id: str
    sentence: str
    polarity: float
    sentiment_label: str
    rank_score: float = 0.0


@dataclass
class GroupSummary:
    group_key: str
    group_label: str
    n_reviews: int = 0
    polarity_sum: float = 0.0
    polarity_pos: int = 0
    polarity_neg: int = 0
    polarity_neu: int = 0
    sentence_pool: list[SentenceCandidate] = field(default_factory=list)

    @property
    def avg_polarity(self) -> float:
        return self.polarity_sum / max(self.n_reviews, 1)

    @property
    def dominant_sentiment(self) -> str:
        if self.polarity_pos >= max(self.polarity_neg, self.polarity_neu):
            return "正向为主"
        if self.polarity_neg >= self.polarity_neu:
            return "负向为主"
        return "中性为主"


def split_sentences(text: str) -> list[str]:
    if not text:
        return []
    parts = SENTENCE_SPLIT_RE.split(text.strip())
    out = []
    for s in parts:
        s = s.strip()
        if MIN_SENT_LEN <= len(s) <= MAX_SENT_LEN:
            out.append(s)
    return out


def _polarity_label(p: float) -> str:
    if p >= 0.3:
        return "positive"
    if p <= -0.3:
        return "negative"
    return "neutral"


def _score_sentence(sent: str, polarity: float, group_dominant_sign: int) -> float:
    """规则评分：长度归一化（越接近 80 字越优）+ 情感与组主导方向一致性 + 否定/比较关键词加权"""
    length_norm = 1.0 - abs(len(sent) - 80) / 200
    length_norm = max(0.1, min(1.0, length_norm))

    sentiment_align = abs(polarity)
    if group_dominant_sign != 0:
        if (polarity > 0 and group_dominant_sign > 0) or (polarity < 0 and group_dominant_sign < 0):
            sentiment_align *= 1.4

    sl = sent.lower()
    keyword_boost = 1.0
    for kw in ("but ", "however", "compared", "vs ", "than ", "wish", "would have",
               "但是", "可惜", "如果", "比", "不如"):
        if kw in sl:
            keyword_boost += 0.15
    keyword_boost = min(keyword_boost, 1.6)

    return length_norm * sentiment_align * keyword_boost


def load_dept_filter(dict_path: Path, dept: str) -> set[str]:
    wb = openpyxl.load_workbook(str(dict_path), read_only=True, data_only=True)
    ws = wb["01_通用标签主表"]
    hdr = [c.value for c in next(ws.iter_rows(max_row=1))]
    ix_tag = hdr.index("标签ID")
    ix_dept = hdr.index("主责部门")
    out: set[str] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[ix_tag] and str(row[ix_dept] or "").strip() == dept:
            out.add(str(row[ix_tag]))
    return out


def build_groups(
    input_path: Path,
    group_by: str,
    tag_filter: set[str] | None = None,
    min_group_size: int = 5,
    max_groups: int = 20,
) -> dict[str, GroupSummary]:
    groups: dict[str, GroupSummary] = {}

    def get_keys(record: dict[str, Any]) -> list[tuple[str, str]]:
        if group_by == "tag":
            keys = []
            for lbl in record.get("labels") or []:
                tid = lbl.get("tag_id")
                if not tid:
                    continue
                if tag_filter is not None and tid not in tag_filter:
                    continue
                keys.append((str(tid), str(lbl.get("tag_cn") or lbl.get("tag_en") or tid)))
            return keys
        if group_by == "aipl":
            stage = (record.get("aipl_stage") or "Unknown")
            return [(stage, f"AIPL-{stage}")]
        if group_by == "persona":
            persona = record.get("persona_derived") or "unspecified"
            return [(persona, persona)]
        return []

    with input_path.open("r", encoding="utf-8") as f:
        for line in f:
            if not line.strip():
                continue
            try:
                r = json.loads(line)
            except json.JSONDecodeError:
                continue
            polarity = float(r.get("sentiment_polarity") or 0.0)
            text = r.get("text") or ""
            rid = r.get("review_id", "")

            for gk, gl in get_keys(r):
                g = groups.get(gk)
                if g is None:
                    g = GroupSummary(group_key=gk, group_label=gl)
                    groups[gk] = g
                g.n_reviews += 1
                g.polarity_sum += polarity
                if polarity >= 0.3:
                    g.polarity_pos += 1
                elif polarity <= -0.3:
                    g.polarity_neg += 1
                else:
                    g.polarity_neu += 1
                if len(g.sentence_pool) < 200:
                    for s in split_sentences(text):
                        if len(g.sentence_pool) >= 200:
                            break
                        g.sentence_pool.append(SentenceCandidate(
                            review_id=rid,
                            sentence=s,
                            polarity=polarity,
                            sentiment_label=_polarity_label(polarity),
                        ))

    filtered = {k: v for k, v in groups.items() if v.n_reviews >= min_group_size}
    sorted_groups = sorted(
        filtered.items(),
        key=lambda kv: -kv[1].n_reviews,
    )[:max_groups]
    return dict(sorted_groups)


def rank_top_sentences(g: GroupSummary, k: int = TOP_K_SENTENCES) -> list[SentenceCandidate]:
    if g.avg_polarity > 0.15:
        sign = 1
    elif g.avg_polarity < -0.15:
        sign = -1
    else:
        sign = 0
    for c in g.sentence_pool:
        c.rank_score = _score_sentence(c.sentence, c.polarity, sign)
    seen_rids: set[str] = set()
    out: list[SentenceCandidate] = []
    for c in sorted(g.sentence_pool, key=lambda x: -x.rank_score):
        if c.review_id in seen_rids:
            continue
        seen_rids.add(c.review_id)
        out.append(c)
        if len(out) >= k:
            break
    return out


def aggregate_summary(g: GroupSummary, top_sentences: list[SentenceCandidate]) -> str:
    pos = sum(1 for s in top_sentences if s.sentiment_label == "positive")
    neg = sum(1 for s in top_sentences if s.sentiment_label == "negative")
    if pos > neg:
        lead = "用户主要给出正面反馈"
    elif neg > pos:
        lead = "用户主要表达不满"
    else:
        lead = "用户反馈正负参半"
    snippets = "；".join(s.sentence[:60].rstrip(".!?") for s in top_sentences)
    return f"{lead}：{snippets}。"


def render_markdown(
    groups: dict[str, GroupSummary],
    group_by: str,
    dept: str | None,
    top_k: int,
) -> tuple[str, list[dict[str, Any]]]:
    md: list[str] = []
    p = md.append
    json_payload: list[dict[str, Any]] = []

    p("---")
    p("name: phase5-d11-agrs-summary")
    p(f"description: Phase 5 D11 AGRS 评论摘要——按 {group_by} 分组，"
      f"每组 Top-{top_k} 代表句 + 聚合摘要 + 情感分布。"
      "BI 看板「评论快讯」模块的离线版数据源。")
    p(f"generated_at: {datetime.now().isoformat(timespec='seconds')}")
    p(f"group_by: {group_by}")
    if dept:
        p(f"dept_filter: {dept}")
    p("doc_type: agrs-summary")
    p("module: voc-nlp")
    p("---")
    p("")
    p(f"# AGRS 评论摘要 — group_by={group_by}"
      + (f" / dept={dept}" if dept else ""))
    p("")
    p(f"分组数：{len(groups)}（min_group_size 过滤后保留）")
    p("")
    p("## 一、分组总览")
    p("")
    p("| 组 | 评论数 | 平均极性 | 主导情感 | 正/中/负 |")
    p("|---|---:|---:|:---:|---|")
    for k, g in groups.items():
        p(f"| {g.group_label} (`{k}`) | {g.n_reviews:,} | {g.avg_polarity:+.2f} | "
          f"{g.dominant_sentiment} | {g.polarity_pos}/{g.polarity_neu}/{g.polarity_neg} |")
    p("")

    p("## 二、各组代表评论 + 聚合摘要")
    p("")
    for k, g in groups.items():
        top = rank_top_sentences(g, k=top_k)
        summary = aggregate_summary(g, top)
        p(f"### {g.group_label} (`{k}`)")
        p("")
        p(f"- **评论数**：{g.n_reviews:,}  **平均极性**：{g.avg_polarity:+.2f}  "
          f"**主导**：{g.dominant_sentiment}")
        p(f"- **聚合摘要**：{summary}")
        p("")
        p(f"**Top-{top_k} 代表句**：")
        p("")
        for i, c in enumerate(top, 1):
            sent = c.sentence.replace("\n", " ").strip()
            p(f"  {i}. [{c.sentiment_label}] `{c.review_id}` — {sent}")
        p("")
        json_payload.append({
            "group_key": k,
            "group_label": g.group_label,
            "n_reviews": g.n_reviews,
            "avg_polarity": round(g.avg_polarity, 4),
            "dominant_sentiment": g.dominant_sentiment,
            "distribution": {
                "positive": g.polarity_pos,
                "neutral": g.polarity_neu,
                "negative": g.polarity_neg,
            },
            "top_sentences": [
                {"review_id": c.review_id, "sentence": c.sentence,
                 "polarity": c.polarity, "label": c.sentiment_label,
                 "rank_score": round(c.rank_score, 4)}
                for c in top
            ],
            "aggregate_summary": summary,
        })

    return "\n".join(md), json_payload


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(description="Phase 5 D11 AGRS 评论摘要生成器")
    ap.add_argument("--input", required=True, type=Path,
                    help="输入 jsonl（phase5_intermediate_merged.jsonl）")
    ap.add_argument("--group-by", choices=["tag", "aipl", "persona"], default="tag",
                    help="分组维度（默认 tag）")
    ap.add_argument("--filter-dept", default=None,
                    help="按主责部门过滤标签（仅 group_by=tag 有效）")
    ap.add_argument("--dict", type=Path,
                    default=Path(__file__).resolve().parent.parent.parent
                    / "04-输出结果/01-字典版本/tag_dictionary_v4.1.xlsx",
                    help="字典路径（默认 v4.1，Phase 6 D2 起切换；仅 --filter-dept 时使用）")
    ap.add_argument("--output", required=True, type=Path,
                    help="Markdown 输出路径")
    ap.add_argument("--json-output", type=Path, default=None,
                    help="可选 JSON 输出路径（BI 二次消费）")
    ap.add_argument("--max-groups", type=int, default=20,
                    help="最多保留多少组（按 n_reviews 降序）")
    ap.add_argument("--min-group-size", type=int, default=5,
                    help="最小组大小（评论数 < 此值被过滤）")
    ap.add_argument("--top-k", type=int, default=TOP_K_SENTENCES,
                    help="每组 Top-K 代表句")
    args = ap.parse_args(argv)

    if not args.input.is_file():
        print(f"❌ 输入不存在: {args.input}", file=sys.stderr); return 2

    tag_filter: set[str] | None = None
    if args.filter_dept and args.group_by == "tag":
        if not args.dict.is_file():
            print(f"❌ 字典不存在: {args.dict}", file=sys.stderr); return 2
        tag_filter = load_dept_filter(args.dict, args.filter_dept)
        print(f"⏳ dept={args.filter_dept} → {len(tag_filter)} 个标签", file=sys.stderr)

    print(f"⏳ AGRS 摘要: group_by={args.group_by} input={args.input}", file=sys.stderr)
    groups = build_groups(
        args.input,
        group_by=args.group_by,
        tag_filter=tag_filter,
        min_group_size=args.min_group_size,
        max_groups=args.max_groups,
    )
    print(f"✅ 保留 {len(groups)} 组", file=sys.stderr)

    md, payload = render_markdown(groups, args.group_by, args.filter_dept, args.top_k)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(md, encoding="utf-8")
    print(f"📄 Report: {args.output}", file=sys.stderr)

    if args.json_output:
        args.json_output.parent.mkdir(parents=True, exist_ok=True)
        args.json_output.write_text(
            json.dumps({
                "generated_at": datetime.now().isoformat(timespec="seconds"),
                "group_by": args.group_by,
                "dept_filter": args.filter_dept,
                "n_groups": len(groups),
                "groups": payload,
            }, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        print(f"📄 JSON: {args.json_output}", file=sys.stderr)

    if not groups:
        print("⚠️  无符合条件的组（min_group_size 可能过严）", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
